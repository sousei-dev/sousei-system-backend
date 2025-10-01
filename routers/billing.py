from fastapi import APIRouter, Depends, HTTPException, Query, status, Request
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import Optional, List
from database import SessionLocal, engine
from models import (
    BillingInvoice, BillingInvoiceItem, BillingMonthlyItem,
    Student, Company, Room, Building, Grade, Department
)
from schemas import (
    InvoiceCreate, InvoiceItemCreate, BillingInvoiceCreate, BillingInvoiceResponse,
    BillingInvoiceItemCreate, BillingInvoiceItemResponse
)
from datetime import datetime, date
import uuid
import calendar
from io import BytesIO
import io
from weasyprint import HTML
from fastapi.responses import StreamingResponse
from urllib.parse import quote
from database_log import create_database_log
from fastapi.templating import Jinja2Templates
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

router = APIRouter(prefix="/billing", tags=["청구 관리"])

# Jinja2Templates 설정
templates = Jinja2Templates(directory="templates")

# HTML to PDF 변환 함수 (weasyprint 66.0 사용)
def html_to_pdf_bytes(html_content: str) -> bytes:
    """
    HTML 내용을 PDF로 변환 (weasyprint 66.0 호환)
    """
    try:
        # weasyprint 66.0에서는 새로운 API 사용
        # 방법 1: HTML(string=html_content) - 최신 방식
        try:
            html_doc = HTML(string=html_content)
            pdf_bytes = html_doc.write_pdf()
            print(f"PDF 변환 성공 (weasyprint 66.0): {len(pdf_bytes)} bytes")
            return pdf_bytes
        except Exception as e1:
            print(f"방법 1 실패: {e1}")
            
            # 방법 2: HTML() 생성자에 직접 HTML 내용 전달
            try:
                html_doc = HTML(html_content)
                pdf_bytes = html_doc.write_pdf()
                print(f"PDF 변환 성공 (방법 2): {len(pdf_bytes)} bytes")
                return pdf_bytes
            except Exception as e2:
                print(f"방법 2 실패: {e2}")
                
                # 방법 3: from_string 메서드 사용
                try:
                    html_doc = HTML.from_string(html_content)
                    pdf_bytes = html_doc.write_pdf()
                    print(f"PDF 변환 성공 (방법 3): {len(pdf_bytes)} bytes")
                    return pdf_bytes
                except Exception as e3:
                    print(f"방법 3 실패: {e3}")
                    raise Exception(f"모든 PDF 변환 방법 실패: {str(e1)}")
                    
    except Exception as e:
        print(f"PDF 변환 중 오류: {e}")
        print(f"HTML 내용 길이: {len(html_content)}")
        print(f"HTML 내용 일부: {html_content[:200]}...")
        raise Exception(f"PDF 변환 실패: {str(e)}")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.get("/invoices")
def get_invoices(
    student_id: Optional[str] = Query(None, description="학생 ID로 필터링"),
    year: Optional[int] = Query(None, description="년도로 필터링"),
    month: Optional[int] = Query(None, description="월로 필터링"),
    page: int = Query(1, description="페이지 번호", ge=1),
    page_size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db)
):
    """청구서 목록 조회"""
    try:
        query = db.query(BillingInvoice).options(
            joinedload(BillingInvoice.student).joinedload(Student.company)
        )
        
        # 필터링 적용
        if student_id:
            query = query.filter(BillingInvoice.student_id == student_id)
        if year:
            query = query.filter(BillingInvoice.year == year)
        if month:
            query = query.filter(BillingInvoice.month == month)
        
        # 전체 항목 수 계산
        total_count = query.count()
        
        # 페이지네이션 적용
        invoices = query.offset((page - 1) * page_size).limit(page_size).all()
        
        # 전체 페이지 수 계산
        total_pages = (total_count + page_size - 1) // page_size
        
        # 응답 데이터 준비
        result = []
        for invoice in invoices:
            invoice_data = {
                "id": str(invoice.id),
                "student_id": str(invoice.student_id),
                "student_name": invoice.student.name if invoice.student else None,
                "company_name": invoice.student.company.name if invoice.student and invoice.student.company else None,
                "year": invoice.year,
                "month": invoice.month,
                "total_amount": invoice.total_amount,
                "status": invoice.status,
                "created_at": invoice.created_at
            }
            result.append(invoice_data)
        
        return {
            "items": result,
            "total": total_count,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_previous": page > 1
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"청구서 조회 중 오류가 발생했습니다: {str(e)}")

@router.post("/billing-invoices/generate/excel")
def generate_company_invoice_excel(
    department_id: str,
    year: int,
    month: int,
    student_type: Optional[str] = Query(None, description="학생 타입으로 필터링"),
    db: Session = Depends(get_db)
):
    """회사별 청구서 상세 데이터를 엑셀 파일로 다운로드 (부서 ID 기반)"""
    try:
        # 부서 존재 여부 확인 및 회사 정보 가져오기
        department = db.query(Department).filter(Department.id == department_id).first()
        if not department:
            raise HTTPException(status_code=404, detail="部署が見つかりません。")
        
        # 부서에서 회사 정보 가져오기
        company = department.company
        if not company:
            raise HTTPException(status_code=404, detail="会社が見つかりません。")
        
        print(f"company: {company.id}")
        company_id = str(company.id)
        
        # 해당 회사의 학생들 조회 (학생 타입 필터링)
        query = db.query(Student).filter(Student.company_id == company_id)
        if student_type:
            query = query.filter(Student.student_type == student_type)
        
        students = query.all()
        if not students:
            raise HTTPException(status_code=404, detail="該当会社の学生がありません。")
        
        # 모든 학생의 월별 항목을 수집
        all_invoice_items = []
        total_company_amount = 0
        
        for student in students:
            # 해당 학생의 해당 월 billing_monthly_items 조회
            monthly_items = db.query(BillingMonthlyItem).filter(
                BillingMonthlyItem.student_id == student.id,
                BillingMonthlyItem.year == year,
                BillingMonthlyItem.month == month
            ).all()
            
            if not monthly_items:
                continue  # 해당 월에 항목이 없으면 건너뛰기
            
            # 총 금액 계산
            total_amount = sum(float(item.amount) for item in monthly_items)
            total_company_amount += total_amount
            
            # 청구서 항목들 준비
            for item in monthly_items:
                invoice_item = {
                    "student_id": student.id,
                    "student_name": student.name,
                    "student_type": student.student_type,
                    "grade_name": student.grade.name if student.grade else None,
                    "item_name": item.item_name,
                    "amount": item.amount,
                    "memo": item.memo,
                    "sort_order": item.sort_order,
                    "original_item_id": item.id
                }
                all_invoice_items.append(invoice_item)
        
        if not all_invoice_items:
            raise HTTPException(status_code=404, detail="該当月の請求項目がありません。")
        
        # 엑셀 데이터 구성 (필요한 컬럼만, 0원 항목 제거)
        excel_data = []
        
        # 헤더 행 추가 (필요한 컬럼만)
        excel_data.append([
            "学生名", "項目名", "金額"
        ])
        
        # 학생별로 그룹화하여 데이터 정리
        student_groups = {}
        for item in all_invoice_items:
            student_id = item["student_id"]
            if student_id not in student_groups:
                student_groups[student_id] = {
                    "student_name": item["student_name"],
                    "items": []
                }
            
            # 0원이 아닌 항목만 추가
            if float(item["amount"]) > 0:
                student_groups[student_id]["items"].append({
                    "item_name": item["item_name"],
                    "amount": item["amount"]
                })
        
        print(f"엑셀 데이터: {student_groups}")
        
        # 그룹화된 데이터를 엑셀 형식으로 변환
        for student_id, student_data in student_groups.items():
            if student_data["items"]:  # 항목이 있는 경우만
                # 첫 번째 행: 학생명과 첫 번째 항목
                first_item = student_data["items"][0]
                excel_data.append([
                    student_data["student_name"],
                    first_item["item_name"],
                    first_item["amount"]
                ])
                
                # 나머지 항목들 (학생명은 빈 값으로)
                for item in student_data["items"][1:]:
                    excel_data.append([
                        "",  # 학생명 병합을 위해 빈 값
                        item["item_name"],
                        item["amount"]
                    ])
        
        # 요약 행 추가
        excel_data.append([])  # 빈 행
        excel_data.append([
            "合計", "", total_company_amount
        ])
        
        # 엑셀 파일 생성
        # DataFrame 생성
        df = pd.DataFrame(excel_data[1:], columns=excel_data[0])
        
        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="請求詳細", index=False)
            ws = writer.sheets["請求詳細"]
            
            # 열 너비 자동 조정
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
            
            # 학생명 셀 병합
            current_student_name = None
            merge_start_row = 2  # 헤더 다음 행부터 시작
            merge_end_row = 2
            
            for row_idx in range(2, len(excel_data) + 1):  # 헤더 제외하고 2행부터
                student_name = ws.cell(row=row_idx, column=1).value
                
                if student_name and student_name != "":  # 학생명이 있는 경우
                    if current_student_name is not None and merge_start_row < merge_end_row:
                        # 이전 학생의 셀 병합
                        ws.merge_cells(f'A{merge_start_row}:A{merge_end_row}')
                    
                    # 새로운 학생 시작
                    current_student_name = student_name
                    merge_start_row = row_idx
                    merge_end_row = row_idx
                else:
                    # 같은 학생의 추가 항목
                    merge_end_row = row_idx
            
            # 마지막 학생의 셀 병합
            if current_student_name is not None and merge_start_row < merge_end_row:
                ws.merge_cells(f'A{merge_start_row}:A{merge_end_row}')
            
            # 셀 정렬 설정
            for row in ws.iter_rows(min_row=2, max_row=len(excel_data), min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
        output.seek(0)

        # 파일명 생성 (슬래시 등 위험문자 제거 권장)
        student_type_filter = f"_{student_type}" if student_type else ""
        display_name = f"請求詳細_{company.name}_{year}年{month}月{student_type_filter}.xlsx"
        # OS/브라우저 안전용: ASCII 대체 이름
        ascii_fallback = "invoice_detail.xlsx"
        # RFC 5987: UTF-8 퍼센트 인코딩
        filename_star = quote(display_name)

        headers = {
            # ASCII 대체 + 실제 유니코드 파일명
            "Content-Disposition": f"attachment; filename={ascii_fallback}; filename*=UTF-8''{filename_star}"
        }

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)}")

# ===== Billing Invoices 관련 API들 =====

@router.post("/billing-invoices/generate")
def generate_company_invoice_pdfV2(
    department_id: str,
    year: int,
    month: int,
    student_type: Optional[str] = Query(None, description="학생 타입으로 필터링"),
    memo: Optional[str] = Query(None, description="비고 작성"),
    db: Session = Depends(get_db)
):
    """회사별 청구서 생성 (billing_monthly_items 기반) 및 PDF 생성 (부서 ID 기반)"""
    try:
        # 부서 존재 여부 확인 및 회사 정보 가져오기
        department = db.query(Department).filter(Department.id == department_id).first()
        if not department:
            raise HTTPException(status_code=404, detail="部署が見つかりません。")
        
        # 부서에서 회사 정보 가져오기
        company = department.company
        if not company:
            raise HTTPException(status_code=404, detail="会社が見つかりません。")
        
        company_id = str(company.id)
        company_name = company.name
        department_name = department.name
        company_address = company.address or ""
        
        # 해당 회사의 학생들 조회 (학생 타입 필터링)
        query = db.query(Student).filter(Student.company_id == company_id)
        if student_type:
            query = query.filter(Student.student_type == student_type)
        
        students = query.all()
        if not students:
            raise HTTPException(status_code=404, detail="該当会社の学生がありません。")
        
        # 회사별로 하나의 청구서 생성
        created_invoices = []
        invoice_list = []
        total_company_amount = 0
        all_invoice_items = []
        
        # 모든 학생의 월별 항목을 수집
        for student in students:
            # 해당 학생의 해당 월 billing_monthly_items 조회
            monthly_items = db.query(BillingMonthlyItem).filter(
                BillingMonthlyItem.student_id == student.id,
                BillingMonthlyItem.year == year,
                BillingMonthlyItem.month == month
            ).all()
            
            if not monthly_items:
                continue  # 해당 월에 항목이 없으면 건너뛰기
            
            # 총 금액 계산
            total_amount = sum(float(item.amount) for item in monthly_items)
            total_company_amount += total_amount
            
            # 청구서 항목들 준비 (스냅샷)
            for item in monthly_items:
                invoice_item = {
                    "student_id": student.id,
                    "item_name": item.item_name,
                    "amount": item.amount,
                    "memo": item.memo,
                    "sort_order": item.sort_order,
                    "original_item_id": item.id
                }
                all_invoice_items.append(invoice_item)
        
        # 회사별 총액이 있는 경우에만 청구서 생성
        if total_company_amount > 0:
            # 청구서 생성
            invoice = BillingInvoice(
                company_id=company.id,
                year=year,
                month=month,
                total_amount=total_company_amount,
                memo=memo or f"{year}年 {month}月 請求書"
            )
            db.add(invoice)
            db.flush()  # ID 생성
            
            # 청구서 항목들 생성
            invoice_items = []
            for item_data in all_invoice_items:
                invoice_item = BillingInvoiceItem(
                    invoice_id=invoice.id,
                    student_id=item_data["student_id"],
                    item_name=item_data["item_name"],
                    amount=item_data["amount"],
                    memo=item_data["memo"],
                    sort_order=item_data["sort_order"],
                    original_item_id=item_data["original_item_id"]
                )
                invoice_items.append(invoice_item)
            
            db.add_all(invoice_items)
            created_invoices.append(invoice)
            
            # PDF용 데이터 구성 (학생별로 그룹화)
            student_invoice_data = {}
            for item_data in all_invoice_items:
                student_id = item_data["student_id"]
                if student_id not in student_invoice_data:
                    student_invoice_data[student_id] = {
                        "student_name": next((s.name for s in students if s.id == student_id), "Unknown Student"),
                        "invoice_items": [],
                        "total_amount": 0
                    }
                
                amount = int(float(item_data["amount"]))
                if amount > 0:  # 0원이 아닌 항목만 추가
                    student_invoice_data[student_id]["invoice_items"].append({
                        "name": item_data["item_name"] or "Unknown Item",
                        "unit_price": amount,
                        "amount": amount,
                        "memo": item_data["memo"] or ""
                    })
                    student_invoice_data[student_id]["total_amount"] += amount
            
            # 0원이 아닌 항목이 있는 학생만 추가
            for student_data in student_invoice_data.values():
                if student_data["total_amount"] > 0:
                    invoice_data = {
                        "student_name": student_data["student_name"],
                        "invoice_number": str(invoice.id),
                        "invoice_items": student_data["invoice_items"],
                        "total_amount": student_data["total_amount"]
                    }
                    invoice_list.append(invoice_data)
        
        db.commit()
        
        # 학생 타입별 청구서 정보 설정
        def get_invoice_info(student_type):
            if student_type == "SPECIFIED":
                return {
                    "sender_name": "株式会社ワールドワーカー",
                    "sender_address": "〒546-0003 大阪市東住吉区今川四丁目5番9号",
                    "sender_tel": "TEL 06-6760-7830",
                    "sender_fax": "FAX -",
                    "registration_number": "登録番号 -",
                    "recipient_name": f"{company_name}　{department_name}　御中",
                    "recipient_address": company_address
                }
            elif student_type == "GENERAL":
                return {
                    "sender_name": "大阪医療介護協同組合",
                    "sender_address": "〒546-0023 大阪府大阪市東住吉区矢田1-26-7-1階",
                    "sender_tel": "TEL 06-6654-8836",
                    "sender_fax": "FAX 06-6654-8837",
                    "registration_number": "登録番号 T6120005018169",
                    "recipient_name": f"{company_name}　{department_name}　御中",
                    "recipient_address": company_address
                }
            else:  # 기본값 (기존 정보)
                return {
                    "sender_name": "大阪医療介護協同組合",
                    "sender_address": "〒546-0023 大阪府大阪市東住吉区矢田1-26-7-1階",
                    "sender_tel": "TEL 06-6654-8836",
                    "sender_fax": "FAX 06-6654-8837",
                    "registration_number": "登録番号 T6120005018169",
                    "recipient_name": f"{company_name}　{department_name}　御中",
                    "recipient_address": company_address
                }
        
        # 작성일 (오늘 날짜)
        invoice_date = datetime.now().strftime("%Y年%m月%d日")
        
        # 입금 기한 (해당 월의 마지막 날)
        last_day = calendar.monthrange(year, month)[1]
        payment_deadline = f"{year}年 {month}月 {last_day}日"
        
        # 학생 타입별 청구서 정보 가져오기
        invoice_info = get_invoice_info(student_type)
        
        # PDF 생성용 데이터 구성
        pdf_data = {
            "total_amount": int(total_company_amount),
            "invoice_list": invoice_list,
            "invoice_date": invoice_date,
            "payment_deadline": payment_deadline,
            "sender_name": invoice_info["sender_name"],
            "sender_address": invoice_info["sender_address"],
            "sender_tel": invoice_info["sender_tel"],
            "sender_fax": invoice_info["sender_fax"],
            "registration_number": invoice_info["registration_number"],
            "recipient_name": invoice_info["recipient_name"],
            "memo": memo or ""
        }
        
        # HTML 템플릿 렌더링
        try:
            template = templates.get_template("company_invoice.html")
            html_content = template.render(data=pdf_data)
            print(f"HTML 템플릿 렌더링 완료, 길이: {len(html_content)}")
        except Exception as template_error:
            print(f"HTML 템플릿 렌더링 실패: {template_error}")
            raise HTTPException(status_code=500, detail=f"HTML 템플릿 렌더링 실패: {str(template_error)}")
        
        # HTML 내용을 UTF-8로 인코딩하여 PDF 생성
        try:
            pdf_bytes = html_to_pdf_bytes(html_content)
            print(f"PDF 생성 성공, 크기: {len(pdf_bytes)} bytes")
        except UnicodeEncodeError as encode_error:
            print(f"UnicodeEncodeError 발생: {encode_error}")
            # 인코딩 에러 발생 시 HTML 내용을 안전하게 처리
            safe_html_content = html_content.encode('utf-8', errors='ignore').decode('utf-8')
            pdf_bytes = html_to_pdf_bytes(safe_html_content)
        except Exception as pdf_error:
            print(f"PDF 생성 실패: {pdf_error}")
            # PDF 생성 실패 시 HTML 파일로 대체
            filename = f"invoice_{company_id}_{year}_{month}.html"
            return StreamingResponse(
                BytesIO(html_content.encode('utf-8')),
                media_type="text/html; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
            )
        
        # 파일명 생성 (완전히 영어로 변경)
        filename = f"invoice_{company_id}_{year}_{month}.pdf"
        
        return StreamingResponse(
            BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"請求書作成中にエラーが発生しました: {str(e)}")

@router.get("/billing-invoices/{invoice_id}")
def get_billing_invoice(
    invoice_id: str,
    db: Session = Depends(get_db)
):
    """청구서 상세 조회"""
    try:
        invoice = db.query(BillingInvoice).options(
            joinedload(BillingInvoice.items)
        ).filter(BillingInvoice.id == invoice_id).first()
        
        if not invoice:
            raise HTTPException(status_code=404, detail="청구서를 찾을 수 없습니다.")
        
        # 회사 정보 조회
        company = db.query(Company).filter(Company.id == invoice.company_id).first()
        
        # 응답 데이터 구성
        invoice_data = {
            "id": str(invoice.id),
            "company_id": str(invoice.company_id),
            "company_name": company.name if company else "Unknown Company",
            "year": invoice.year,
            "month": invoice.year,
            "total_amount": float(invoice.total_amount) if invoice.total_amount else 0,
            "memo": invoice.memo,
            "status": invoice.status,
            "created_at": invoice.created_at.strftime("%Y-%m-%d %H:%M:%S") if invoice.created_at else None,
            "updated_at": invoice.updated_at.strftime("%Y-%m-%d %H:%M:%S") if invoice.updated_at else None,
            "items": []
        }
        
        # 청구서 항목들 추가
        for item in invoice.items:
            item_data = {
                "id": str(item.id),
                "student_id": str(item.student_id),
                "item_name": item.item_name,
                "amount": float(item.amount) if item.amount else 0,
                "memo": item.memo,
                "sort_order": item.sort_order,
                "original_item_id": str(item.original_item_id) if item.original_item_id else None
            }
            invoice_data["items"].append(item_data)
        
        return invoice_data
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"청구서 조회 중 오류가 발생했습니다: {str(e)}")

@router.get("/billing-invoices/student/{student_id}")
def get_student_billing_invoices(
    student_id: str,
    year: Optional[int] = Query(None, description="년도로 필터링"),
    month: Optional[int] = Query(None, description="월로 필터링"),
    db: Session = Depends(get_db)
):
    """학생별 청구서 조회"""
    try:
        # 학생 존재 여부 확인
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다.")
        
        # 학생이 속한 회사의 청구서들 조회
        query = db.query(BillingInvoice).filter(BillingInvoice.company_id == student.company_id)
        
        # 년도 필터링
        if year:
            query = query.filter(BillingInvoice.year == year)
        
        # 월 필터링
        if month:
            query = query.filter(BillingInvoice.month == month)
        
        invoices = query.order_by(BillingInvoice.created_at.desc()).all()
        
        # 각 청구서에서 해당 학생의 항목들만 추출
        result = []
        for invoice in invoices:
            student_items = db.query(BillingInvoiceItem).filter(
                BillingInvoiceItem.invoice_id == invoice.id,
                BillingInvoiceItem.original_item_id == student_id
            ).all()
            
            if student_items:
                invoice_data = {
                    "invoice_id": str(invoice.id),
                    "year": invoice.year,
                    "month": invoice.month,
                    "total_amount": float(invoice.total_amount) if invoice.total_amount else 0,
                    "memo": invoice.memo,
                    "created_at": invoice.created_at.strftime("%Y-%m-%d %H:%M:%S") if invoice.created_at else None,
                    "student_items": []
                }
                
                for item in student_items:
                    item_data = {
                        "id": str(item.id),
                        "item_name": item.item_name,
                        "amount": float(item.amount) if item.amount else 0,
                        "memo": item.memo,
                        "sort_order": item.sort_order
                    }
                    invoice_data["student_items"].append(item_data)
                
                result.append(invoice_data)
        
        return {
            "student_id": student_id,
            "student_name": student.name,
            "company_id": str(student.company_id),
            "company_name": student.company.name if student.company else "Unknown Company",
            "invoices": result,
            "total_invoices": len(result)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"학생별 청구서 조회 중 오류가 발생했습니다: {str(e)}")

@router.delete("/billing-invoices/{invoice_id}")
def delete_billing_invoice(
    invoice_id: str,
    db: Session = Depends(get_db)
):
    """청구서 삭제"""
    try:
        invoice = db.query(BillingInvoice).filter(BillingInvoice.id == invoice_id).first()
        
        if not invoice:
            raise HTTPException(status_code=404, detail="청구서를 찾을 수 없습니다.")
        
        # 청구서 항목들 먼저 삭제
        db.query(BillingInvoiceItem).filter(BillingInvoiceItem.invoice_id == invoice_id).delete()
        
        # 청구서 삭제
        db.delete(invoice)
        db.commit()
        
        return {
            "message": "청구서가 성공적으로 삭제되었습니다.",
            "deleted_invoice_id": invoice_id
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"청구서 삭제 중 오류가 발생했습니다: {str(e)}") 