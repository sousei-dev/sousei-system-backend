from fastapi import FastAPI, Depends, HTTPException, Query, status, UploadFile, File, Response
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm, HTTPBearer
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request
from jose import JWTError, jwt
from datetime import datetime, timedelta, date
from typing import Optional, List
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from database import SessionLocal, engine
from models import Base, User, Company, Student, Grade, Invoice, InvoiceItem, BillingItem, Building, Room, Resident, RoomLog, RoomCharge, ChargeItem, ChargeItemAllocation, RoomUtility, ResidenceCardHistory, DatabaseLog, Elderly, ElderlyCategories, BuildingCategoriesRent, ElderlyContract, ElderlyInvoiceItem, ElderlyInvoice, CareItem, CareMealPrice, CareUtilityPrice, BillingMonthlyItem, BillingInvoice, BillingInvoiceItem, ElderlyMealRecord, ElderlyHospitalization
from schemas import UserCreate, UserLogin, StudentUpdate, StudentResponse, InvoiceCreate, InvoiceResponse, StudentCreate, InvoiceUpdate, BuildingCreate, BuildingUpdate, BuildingResponse, RoomCreate, RoomUpdate, RoomResponse, ChangeResidenceRequest, CheckInRequest, CheckOutRequest, AssignRoomRequest, NewResidenceRequest, RoomLogResponse, ResidentResponse, RoomCapacityStatus, EmptyRoomOption, BuildingOption, AvailableRoom, RoomChargeCreate, RoomChargeUpdate, RoomChargeResponse, ChargeItemCreate, ChargeItemUpdate, ChargeItemResponse, ChargeItemAllocationCreate, ChargeItemAllocationUpdate, ChargeItemAllocationResponse, RoomUtilityCreate, RoomUtilityUpdate, RoomUtilityResponse, ResidenceCardHistoryCreate, ResidenceCardHistoryUpdate, ResidenceCardHistoryResponse, VisaInfoUpdate, ElderlyCreate, ElderlyUpdate, ElderlyResponse, ElderlyContractCreate, ElderlyContractUpdate, ElderlyContractResponse, ElderlyInvoiceItemCreate, ElderlyInvoiceItemUpdate, ElderlyInvoiceItemResponse, ElderlyInvoiceCreate, ElderlyInvoiceUpdate, ElderlyInvoiceResponse, CareMealPriceCreate, CareMealPriceUpdate, CareMealPriceResponse, CareUtilityPriceCreate, CareUtilityPriceUpdate, CareUtilityPriceResponse, CareItemResponse, BillingMonthlyItemCreate, BillingMonthlyItemUpdate, BillingMonthlyItemResponse, BillingInvoiceCreate, BillingInvoiceResponse, BillingInvoiceItemCreate, BillingInvoiceItemResponse, ElderlyMealRecordCreate, ElderlyMealRecordUpdate, ElderlyMealRecordResponse, ElderlyHospitalizationCreate, ElderlyHospitalizationResponse, MonthlyItemSortOrderUpdate
import uuid
import json
from uuid import UUID
from passlib.hash import bcrypt
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
import os
from supabase import create_client
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
from io import BytesIO
import io
from weasyprint import HTML
from pydantic import BaseModel
import time
import calendar
from urllib.parse import quote
import pandas as pd
import openpyxl

templates = Jinja2Templates(directory="templates")

# .env 파일 로드
load_dotenv()

# JWT 설정
SECRET_KEY = os.getenv("SECRET_KEY")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 7  # 7일로 변경

# .env 파일에서 Supabase 설정 로드
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY")  # anon key 사용 (로그인용)
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")  # service role key 사용 (Storage용)

supabase = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
# Storage 작업을 위한 service role 클라이언트
supabase_storage = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY) if SUPABASE_SERVICE_KEY else supabase

app = FastAPI()
Base.metadata.create_all(bind=engine)


# CORS 설정
origins = [
    "http://localhost:5173",
    "http://localhost:5174",
    "http://localhost:5175",    
    "http://localhost:8000",     
    "http://127.0.0.1:3000",
    "http://127.0.0.1:8000",
    "http://127.0.0.1:8080",
    "https://system.sousei-group.com",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  # 위에서 정의한 origins 리스트
    allow_credentials=True,  # 쿠키를 포함한 요청을 허용
    allow_methods=["*"],    # 모든 HTTP 메서드 허용
    allow_headers=["*"],    # 모든 HTTP 헤더 허용
)

# 로그 유틸리티 함수들
def create_database_log(
    db: Session,
    table_name: str,
    record_id: str,
    action: str,
    user_id: Optional[str] = None,
    old_values: Optional[dict] = None,
    new_values: Optional[dict] = None,
    changed_fields: Optional[list] = None,
    ip_address: Optional[str] = None,
    user_agent: Optional[str] = None,
    note: Optional[str] = None
):
    """데이터베이스 로그 생성"""
    try:
        log = DatabaseLog(
            id=str(uuid.uuid4()),
            table_name=table_name,
            record_id=record_id,
            action=action,
            user_id=user_id,
            old_values=json.dumps(old_values, ensure_ascii=False, default=str) if old_values else None,
            new_values=json.dumps(new_values, ensure_ascii=False, default=str) if new_values else None,
            changed_fields=json.dumps(changed_fields, ensure_ascii=False) if changed_fields else None,
            ip_address=ip_address,
            user_agent=user_agent,
            note=note
        )
        db.add(log)
        db.commit()  # 로그를 즉시 커밋
        print(f"로그 생성 완료: {table_name} - {action} - {record_id}")
        return log
    except Exception as e:
        print(f"로그 생성 중 오류: {str(e)}")
        db.rollback()
        return None

def get_changed_fields(old_values: dict, new_values: dict) -> list:
    """변경된 필드들을 찾아서 반환"""
    changed = []
    if not old_values or not new_values:
        return changed
    
    for key in new_values:
        if key in old_values:
            if old_values[key] != new_values[key]:
                changed.append(key)
        else:
            changed.append(key)
    
    return changed

def get_client_ip(request):
    """클라이언트 IP 주소 가져오기"""
    if hasattr(request, 'client'):
        return request.client.host
    return None

def get_user_agent(request):
    """User Agent 가져오기"""
    if hasattr(request, 'headers'):
        return request.headers.get('user-agent')
    return None

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# 토큰 생성 함수
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# OAuth2 설정
def get_token_header(token: str = Depends(HTTPBearer())):
    return token.credentials

# 현재 사용자 가져오기 (Supabase Auth 사용)
async def get_current_user(token: str = Depends(get_token_header)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        # Supabase Auth 토큰 검증
        user = supabase.auth.get_user(token)
        
        # 사용자 권한 정보 조회
        try:
            profile_response = supabase.table("profiles").select("name, role").eq("id", user.user.id).execute()
            profile_data = profile_response.data[0] if profile_response.data else {}
            
            # 사용자 정보를 딕셔너리로 변환하여 권한 정보 추가
            user_dict = {
                "id": user.user.id,
                "email": user.user.email,
                "name": profile_data.get("name"),
                "role": profile_data.get("role", "manager")
            }
            
        except Exception as profile_error:
            # 프로필 조회 실패 시 기본값 설정
            user_dict = {
                "id": user.user.id,
                "email": user.user.email,
                "name": None,
                "role": "manager"
            }
        
        return user_dict
    except Exception as e:
        print(f"토큰 검증 에러: {e}")
        raise credentials_exception

# 로그인 엔드포인트
@app.post("/login")
async def login(user: UserLogin):
    try:
        
        # Supabase Auth를 사용한 로그인
        auth_response = supabase.auth.sign_in_with_password({
            "email": user.email,
            "password": user.password
        })
        
        # 로그인 성공 시 사용자 정보 반환
        user_data = auth_response.user
        session = auth_response.session
        
        # 사용자 권한 정보 조회
        try:
            profile_response = supabase.table("profiles").select("name, role").eq("id", user_data.id).execute()
            profile_data = profile_response.data[0] if profile_response.data else {}
            
            return {
                "message": "ログインに成功しました",
                "user_id": user_data.id,
                "email": user_data.email,
                "name": profile_data.get("name"),
                "role": profile_data.get("role", "manager"),
                "access_token": session.access_token,
                "refresh_token": session.refresh_token,
                "token_type": "bearer"
            }
        except Exception as profile_error:
            # 프로필 조회 실패 시 기본 정보만 반환
            return {
                "message": "ログインに成功しました",
                "user_id": user_data.id,
                "email": user_data.email,
                "name": None,
                "role": "manager",
                "access_token": session.access_token,
                "refresh_token": session.refresh_token,
                "token_type": "bearer"
            }
        
    except Exception as e:
        # Supabase Auth 에러 처리
        error_message = str(e)
        
        if "Invalid login credentials" in error_message:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="이메일 또는 비밀번호가 올바르지 않습니다."
            )
        elif "Email not confirmed" in error_message:
            # 이메일 인증 없이 로그인 허용 (개발 환경용)
            try:
                # Supabase 대시보드에서 이메일 인증을 비활성화하거나
                # 해당 사용자의 이메일 인증 상태를 수동으로 true로 변경해야 합니다.
                
                # 임시 해결책: 에러 메시지에 해결 방법 안내
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="이메일 인증이 필요합니다. 해결 방법: 1) Supabase 대시보드 → Authentication → Settings → 'Enable email confirmations' 체크 해제, 2) 또는 Authentication → Users에서 해당 사용자의 'Email confirmed'를 true로 변경"
                )
            except Exception as retry_error:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="이메일 인증이 필요합니다. Supabase 대시보드에서 이메일 인증 설정을 확인해주세요."
                )
        elif "Invalid API key" in error_message:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="API 키가 올바르지 않습니다. Supabase 설정을 확인해주세요."
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"ログイン中にエラーが発生しました: {error_message}"
            )

# 로그아웃 엔드포인트
@app.post("/logout")
async def logout():
    try:
        # Supabase Auth 로그아웃
        supabase.auth.sign_out()
        return {"message": "ログアウトに成功しました"}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"ログアウト中にエラーが発生しました: {str(e)}"
        )

# 토큰 갱신 엔드포인트
@app.post("/refresh-token")
async def refresh_token(refresh_token: str):
    try:
        # Supabase Auth 토큰 갱신
        auth_response = supabase.auth.refresh_session(refresh_token)
        
        return {
            "message": "トークン更新に成功しました",
            "access_token": auth_response.session.access_token,
            "refresh_token": auth_response.session.refresh_token,
            "token_type": "bearer"
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"토큰 갱신 실패: {str(e)}"
        )

# 사용자 프로필 조회 엔드포인트
@app.get("/profile")
async def get_profile(current_user: dict = Depends(get_current_user)):
    try:
        profile_response = supabase.table("profiles").select("name, role").eq("id", current_user.id).execute()
        profile_data = profile_response.data[0] if profile_response.data else {}
        
        return {
            "user_id": current_user["id"],
            "email": current_user["email"],
            "name": profile_data.get("name"),
            "role": profile_data.get("role", "manager")
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"프로필 조회 중 오류가 발생했습니다: {str(e)}"
        )

# 사용자 프로필 업데이트 엔드포인트
@app.put("/profile")
async def update_profile(
    profile_update: dict,
    current_user: dict = Depends(get_current_user)
):
    try:
        # 업데이트할 필드 검증
        allowed_fields = ["name", "role"]
        update_data = {k: v for k, v in profile_update.items() if k in allowed_fields}
        
        if "role" in update_data and update_data["role"] not in ["admin", "manager"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="role은 'admin' 또는 'manager'만 허용됩니다."
            )
        
        # 프로필 업데이트
        profile_response = supabase.table("profiles").update(update_data).eq("id", current_user.id).execute()
        
        return {
            "message": "프로필이 성공적으로 업데이트되었습니다.",
            "user_id": current_user["id"],
            "email": current_user["email"],
            "name": update_data.get("name"),
            "role": update_data.get("role", "manager")
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"프로필 업데이트 중 오류가 발생했습니다: {str(e)}"
        )

# 회원가입 엔드포인트
@app.post("/signup")
async def signup(user: UserCreate):
    try:
        # Supabase Auth를 사용한 회원가입 (이메일 인증 없이)
        auth_response = supabase.auth.sign_up({
            "email": user.email,
            "password": user.password,
            "options": {
                "email_confirm": False  # 이메일 인증 없이 가입
            }
        })
        
        # 회원가입 성공 시 사용자 정보 반환
        user_data = auth_response.user
        
        # 프로필 테이블에 기본 레코드 생성
        try:
            profile_data = {
                "id": user_data.id,
                "name": None,
                "role": "manager"
            }
            supabase.table("profiles").insert(profile_data).execute()
        except Exception as profile_error:
            # 프로필 생성 실패 시에도 회원가입은 성공으로 처리
            print(f"프로필 생성 실패: {profile_error}")
        
        return {
            "message": "会員登録に成功しました",
            "user_id": user_data.id,
            "email": user_data.email,
            "name": None,
            "role": "manager"
        }
        
    except Exception as e:
        # Supabase Auth 에러 처리
        error_message = str(e)
        if "User already registered" in error_message:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="이미 가입된 이메일입니다."
            )
        elif "Password should be at least" in error_message:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="비밀번호는 최소 6자 이상이어야 합니다."
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"会員登録中にエラーが発生しました: {error_message}"
            )

# 학생 목록 조회 (인증 필요)
@app.get("/students")
def get_students(
    name: Optional[str] = Query(None, description="학생 이름으로 검색"),
    name_katakana: Optional[str] = Query(None, description="학생 이름 카타카나로 검색"),
    nationality: Optional[str] = Query(None, description="학생 국적으로 검색"),
    company: Optional[str] = Query(None, description="회사 이름으로 검색"),
    consultant: Optional[int] = Query(None, description="컨설턴트 번호 이상으로 검색"),
    email: Optional[str] = Query(None, description="이메일로 검색"),
    student_type: Optional[str] = Query(None, description="학생 유형으로 검색"),
    building_name: Optional[str] = Query(None, description="건물 이름으로 검색"),
    room_number: Optional[str] = Query(None, description="방 번호로 검색"),
    status: Optional[str] = Query(None, description="상태로 검색"),
    grade: Optional[str] = Query(None, description="등급으로 검색"),
    sort_by: Optional[str] = Query(None, description="정렬 필드 (nationality 또는 grade)"),
    sort_desc: Optional[bool] = Query(False, description="내림차순 정렬 여부 (true: 내림차순, false: 오름차순)"),
    page: int = Query(1, description="페이지 번호", ge=1),  # 1 이상의 정수
    page_size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),  # 1~100 사이의 정수
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)  # 인증 추가
):
    # 기본 쿼리 생성
    query = db.query(Student).options(
        joinedload(Student.company),
        joinedload(Student.grade),
        joinedload(Student.current_room).joinedload(Room.building)
    ).outerjoin(Company).outerjoin(Grade)

    # 방 관련 필터링이 있는 경우 Room과 Building 테이블과 조인
    if building_name or room_number:
        query = query.outerjoin(Room, Student.current_room_id == Room.id).outerjoin(Building, Room.building_id == Building.id)

    # 필터 조건 추가
    if student_type and student_type != 'ALL':
        query = query.filter(Student.student_type.ilike(f"%{student_type}%"))
    if name:
        query = query.filter(Student.name.ilike(f"%{name}%"))
    if nationality:
        query = query.filter(Student.nationality.ilike(f"%{nationality}%"))
    if name_katakana:
        query = query.filter(Student.name_katakana.ilike(f"%{name_katakana}%"))
    if company:
        query = query.filter(Company.name.ilike(f"%{company}%"))
    if consultant is not None:
        query = query.filter(Student.consultant >= consultant)
    if email:
        query = query.filter(Student.email.ilike(f"%{email}%"))
    if building_name:
        query = query.filter(Building.name.ilike(f"%{building_name}%"))
    if room_number:
        query = query.filter(Room.room_number.ilike(f"%{room_number}%"))
    if status:
        query = query.filter(Student.status == status)
    if grade:
        query = query.filter(Grade.name.ilike(f"%{grade}%"))
    
    # 정렬 적용
    if sort_by:
        if sort_by == "nationality":
            if sort_desc:
                query = query.order_by(Student.nationality.desc())
            else:
                query = query.order_by(Student.nationality.asc())
        elif sort_by == "grade.name":
            if sort_desc:
                query = query.order_by(Grade.name.desc())
            else:
                query = query.order_by(Grade.name.asc())
    
    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    students = query.offset((page - 1) * page_size).limit(page_size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + page_size - 1) // page_size

    # 응답 데이터 준비
    result = []
    for student in students:
        student_data = {
            "id": str(student.id),
            "name": student.name,
            "email": student.email if student.email else "",
            "created_at": student.created_at,
            "company_id": str(student.company_id) if student.company_id else None,
            "consultant": student.consultant,
            "phone": student.phone,
            "avatar": student.avatar,
            "grade_id": str(student.grade_id) if student.grade_id else None,
            "cooperation_submitted_date": student.cooperation_submitted_date,
            "cooperation_submitted_place": student.cooperation_submitted_place,
            "assignment_date": student.assignment_date,
            "ward": student.ward,
            "name_katakana": student.name_katakana,
            "gender": student.gender,
            "birth_date": student.birth_date,
            "nationality": student.nationality,
            "has_spouse": student.has_spouse,
            "japanese_level": student.japanese_level,
            "passport_number": student.passport_number,
            "residence_card_number": student.residence_card_number,
            "residence_card_start": student.residence_card_start,
            "residence_card_expiry": student.residence_card_expiry,
            "resignation_date": student.resignation_date,
            "local_address": student.local_address,
            "address": student.address,
            "experience_over_2_years": student.experience_over_2_years,
            "status": student.status,
            "arrival_type": student.arrival_type,
            "company": {
                "name": student.company.name
            } if student.company else None,
            "grade": {
                "name": student.grade.name
            } if student.grade else None,
            "current_room": {
                "room_number": student.current_room.room_number,
                "building_name": student.current_room.building.name
            } if student.current_room and student.current_room.building else None
        }
        result.append(student_data)

    return {
        "items": result,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_previous": page > 1
    }

@app.get("/students/{student_id}")
def get_student(student_id: str, db: Session = Depends(get_db)):
    student = db.query(Student).options(
        joinedload(Student.company), 
        joinedload(Student.grade),
        joinedload(Student.current_room).joinedload(Room.building)
    ).filter(Student.id == student_id).first()
    
    if student is None:
        raise HTTPException(status_code=404, detail="学生が見つかりません")
    
    # 응답 데이터 준비
    student_data = {
        "id": str(student.id),
        "name": student.name,
        "email": student.email if student.email else "",
        "created_at": student.created_at,
        "company_id": str(student.company_id) if student.company_id else None,
        "consultant": student.consultant,
        "phone": student.phone,
        "avatar": student.avatar,
        "grade_id": str(student.grade_id) if student.grade_id else None,
        "cooperation_submitted_date": student.cooperation_submitted_date,
        "cooperation_submitted_place": student.cooperation_submitted_place,
        "assignment_date": student.assignment_date,
        "ward": student.ward,
        "name_katakana": student.name_katakana,
        "gender": student.gender,
        "birth_date": student.birth_date,
        "nationality": student.nationality,
        "has_spouse": student.has_spouse,
        "japanese_level": student.japanese_level,
        "passport_number": student.passport_number,
        "residence_card_number": student.residence_card_number,
        "residence_card_start": student.residence_card_start,
        "residence_card_expiry": student.residence_card_expiry,
        "resignation_date": student.resignation_date,
        "local_address": student.local_address,
        "address": student.address,
        "experience_over_2_years": student.experience_over_2_years,
        "status": student.status,
        "arrival_type": student.arrival_type,
        "entry_date": student.entry_date,
        "interview_date": student.interview_date,
        "pre_guidance_date": student.pre_guidance_date,
        "orientation_date": student.orientation_date,
        "certification_application_date": student.certification_application_date,
        "visa_application_date": student.visa_application_date,
        "passport_expiration_date": student.passport_expiration_date,
        "student_type": student.student_type,
        "current_room_id": str(student.current_room_id) if student.current_room_id else None,
        "facebook_name": student.facebook_name,
        "visa_year": student.visa_year,
        "company": {
            "id": str(student.company.id),
            "name": student.company.name
        } if student.company else None,
        "grade": {
            "id": str(student.grade.id),
            "name": student.grade.name
        } if student.grade else None,
        "current_room": {
          "room_number": student.current_room.room_number,
          "building_name": student.current_room.building.name
        } if student.current_room and student.current_room.building else None
    }
    
    return student_data

@app.post("/students", status_code=201)
def create_student(
    student: StudentCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    try:
        # 빈 문자열을 None으로 변환하는 헬퍼 함수
        def clean_value(value):
            if value == "" or value is None:
                return None
            return value
        
        # UUID 변환 헬퍼 함수
        def parse_uuid(value):
            if value is None or value == "":
                return None
            try:
                return str(value) if isinstance(value, UUID) else value
            except:
                return None
        
        # 날짜 변환 헬퍼 함수
        def parse_date(value):
            if value is None or value == "":
                return None
            if isinstance(value, date):
                return value
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except:
                return None

        # 학생 생성
        new_student = Student(
            id=str(uuid.uuid4()),
            name=student.name,
            company_id=parse_uuid(student.company_id),
            grade_id=parse_uuid(student.grade_id),
            consultant=student.consultant,
            cooperation_submitted_date=parse_date(student.cooperation_submitted_date),
            cooperation_submitted_place=clean_value(student.cooperation_submitted_place),
            assignment_date=parse_date(student.assignment_date),
            ward=clean_value(student.ward),
            name_katakana=clean_value(student.name_katakana),
            gender=clean_value(student.gender),
            birth_date=parse_date(student.birth_date),
            nationality=clean_value(student.nationality),
            has_spouse=student.has_spouse,
            japanese_level=clean_value(student.japanese_level),
            passport_number=clean_value(student.passport_number),
            phone=clean_value(student.phone),
            address=clean_value(student.address),
            arrival_type=clean_value(student.arrival_type),
            residence_card_number=clean_value(student.residence_card_number),
            residence_card_start=parse_date(student.residence_card_start),
            residence_card_expiry=parse_date(student.residence_card_expiry),
            experience_over_2_years=student.experience_over_2_years,
            entry_date=parse_date(student.entry_date),
            local_address=clean_value(student.local_address),
            interview_date=parse_date(student.interview_date),
            pre_guidance_date=parse_date(student.pre_guidance_date),
            orientation_date=parse_date(student.orientation_date),
            certification_application_date=parse_date(student.certification_application_date),
            visa_application_date=parse_date(student.visa_application_date),
            passport_expiration_date=parse_date(student.passport_expiration_date),
            student_type=clean_value(student.student_type),
            current_room_id=parse_uuid(student.current_room_id),
            facebook_name=clean_value(student.facebook_name),
            visa_year=clean_value(student.visa_year),
        )
        db.add(new_student)
        db.flush()  # ID 생성을 위해 flush

        # Residence Card 히스토리 저장 (residence card 정보가 있는 경우)
        if student.residence_card_number and student.residence_card_start and student.residence_card_expiry:
            # 같은 년차인지 확인
            if student.visa_year:
                existing_history = db.query(ResidenceCardHistory).filter(
                    ResidenceCardHistory.student_id == new_student.id,
                    ResidenceCardHistory.year == student.visa_year
                ).first()
                
                if existing_history:
                    raise HTTPException(
                        status_code=400,
                        detail=f"{student.visa_year}년차의 residence card가 이미 존재합니다. 다른 년차를 입력해주세요."
                    )
            
            residence_card_history = ResidenceCardHistory(
                id=str(uuid.uuid4()),
                student_id=new_student.id,
                card_number=student.residence_card_number,
                start_date=parse_date(student.residence_card_start),
                expiry_date=parse_date(student.residence_card_expiry),
                year=student.visa_year,
                note="新規登録"
            )
            db.add(residence_card_history)

        # 방 배정 및 입주 처리
        room_info = None
        if student.room_id:
            try:
                print(f"DEBUG: 방 배정 시작 - room_id: {student.room_id}")
                
                # 방 존재 여부 확인
                room = db.query(Room).filter(Room.id == student.room_id).first()
                if not room:
                    raise HTTPException(status_code=404, detail="指定された部屋が見つかりません")

                print(f"DEBUG: 방 확인 완료 - room_number: {room.room_number}")

                # 방이 사용 가능한지 확인
                if not room.is_available:
                    raise HTTPException(status_code=400, detail="該当の部屋は現在使用できません")

                # 방의 정원 확인
                current_residents = db.query(Resident).filter(
                    Resident.room_id == student.room_id,
                    Resident.is_active == True,
                    Resident.check_out_date.is_(None)
                ).count()
                
                print(f"DEBUG: 현재 거주자 수: {current_residents}, 정원: {room.capacity}")
                
                if room.capacity and current_residents >= room.capacity:
                    raise HTTPException(status_code=400, detail="該当の部屋は定員が超過しているため入居できません")

                # 입주일 설정 (기본값: 오늘)
                check_in_date = datetime.now().date()
                if student.check_in_date:
                    check_in_date = datetime.strptime(student.check_in_date, "%Y-%m-%d").date()

                print(f"DEBUG: 입주일 설정: {check_in_date}")

                # 학생의 current_room_id 설정
                new_student.current_room_id = student.room_id

                # 입주 기록 생성 (Resident 테이블)
                new_resident = Resident(
                    id=str(uuid.uuid4()),
                    room_id=student.room_id,
                    student_id=new_student.id,
                    check_in_date=check_in_date,
                    note=clean_value(student.room_note)
                )
                db.add(new_resident)
                print(f"DEBUG: Resident 레코드 생성 완료 - id: {new_resident.id}")

                # 방 로그 기록 (RoomLog 테이블)
                room_log = RoomLog(
                    id=str(uuid.uuid4()),
                    room_id=student.room_id,
                    student_id=new_student.id,
                    action="CHECK_IN",
                    action_date=check_in_date,
                    note=f"新規学生入居 - {clean_value(student.room_note)}" if clean_value(student.room_note) else "新規学生入居"
                )
                db.add(room_log)
                print(f"DEBUG: RoomLog 레코드 생성 완료 - id: {room_log.id}")

                room_info = {
                    "room_id": str(room.id),
                    "room_number": room.room_number,
                    "building_name": room.building.name if room.building else None,
                    "check_in_date": check_in_date.strftime("%Y-%m-%d")
                }
                
                print(f"DEBUG: 방 배정 완료 - 학생: {new_student.id}, 방: {student.room_id}")
                
            except Exception as e:
                print(f"DEBUG: 방 배정 중 오류 - {str(e)}")
                db.rollback()
                raise HTTPException(
                    status_code=500,
                    detail=f"방 배정 중 오류가 발생했습니다: {str(e)}"
                )

        try:
            db.commit()
            print(f"DEBUG: 데이터베이스 커밋 완료")
            db.refresh(new_student)
            
        except Exception as e:
            print(f"DEBUG: 커밋 중 오류 - {str(e)}")
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"데이터베이스 저장 중 오류가 발생했습니다: {str(e)}"
            )
        
        # 로그 생성 (별도 트랜잭션)
        try:
            create_database_log(
                db=db,
                table_name="students",
                record_id=str(new_student.id),
                action="CREATE",
                user_id=current_user["id"] if current_user else None,
                new_values={
                    "name": new_student.name,
                    "email": new_student.email,
                    "company_id": str(new_student.company_id) if new_student.company_id else None,
                    "consultant": new_student.consultant,
                    "phone": new_student.phone,
                    "grade_id": str(new_student.grade_id) if new_student.grade_id else None,
                    "residence_card_number": new_student.residence_card_number,
                    "residence_card_start": str(new_student.residence_card_start) if new_student.residence_card_start else None,
                    "residence_card_expiry": str(new_student.residence_card_expiry) if new_student.residence_card_expiry else None,
                    "visa_year": new_student.visa_year
                },
                note="学生新規登録"
            )
        except Exception as log_error:
            print(f"로그 생성 중 오류: {str(log_error)}")
            # 로그 생성 실패는 학생 생성에 영향을 주지 않도록 함

        return {
            "message": "学生が正常に作成されました",
            "student": {
                "id": str(new_student.id),
                "name": new_student.name,
                "company_id": str(new_student.company_id) if new_student.company_id else None,
                "consultant": new_student.consultant,
                "grade_id": str(new_student.grade_id) if new_student.grade_id else None,
                "current_room_id": str(new_student.current_room_id) if new_student.current_room_id else None,
                "created_at": new_student.created_at
            },
            "room_assignment": room_info
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"학생 생성 중 오류가 발생했습니다: {str(e)}"
        )

@app.put("/students/{student_id}", response_model=StudentResponse)
def update_student(
    student_id: str,
    student_update: StudentUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 헬퍼 함수
    def parse_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None
    
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(
            status_code=404,
            detail="학생을 찾을 수 없습니다."
        )

    # 회사 존재 여부 확인 (회사가 변경되는 경우)
    if student_update.company_id:
        company = db.query(Company).filter(Company.id == student_update.company_id).first()
        if not company:
            raise HTTPException(
                status_code=404,
                detail="존재하지 않는 회사입니다."
            )

    # Residence Card 정보가 변경되었는지 확인
    residence_card_changed = False
    if (student_update.residence_card_number and 
        student_update.residence_card_start and 
        student_update.residence_card_expiry):
        
        # 기존 정보와 비교
        if (student.residence_card_number != student_update.residence_card_number or
            student.residence_card_start != student_update.residence_card_start or
            student.residence_card_expiry != student_update.residence_card_expiry):
            residence_card_changed = True
            
            # 같은 년차인지 확인
            if student_update.visa_year:
                # 기존 히스토리에서 같은 년차가 있는지 확인
                existing_history = db.query(ResidenceCardHistory).filter(
                    ResidenceCardHistory.student_id == student_id,
                    ResidenceCardHistory.year == student_update.visa_year
                ).first()
                
                if existing_history:
                    raise HTTPException(
                        status_code=400,
                        detail=f"{student_update.visa_year}년차의 residence card가 이미 존재합니다. 다른 년차를 입력해주세요."
                    )

    # 학생 정보 업데이트
    update_data = student_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(student, field, value)

    # Residence Card 히스토리 저장 (변경된 경우)
    if residence_card_changed:
        residence_card_history = ResidenceCardHistory(
            id=str(uuid.uuid4()),
            student_id=student.id,
            card_number=student_update.residence_card_number,
            start_date=parse_date(student_update.residence_card_start),
            expiry_date=parse_date(student_update.residence_card_expiry),
            year=student_update.visa_year,
            note="学生情報更新"
        )
        db.add(residence_card_history)
        print(f"DEBUG: Residence Card 히스토리 생성 완료 (업데이트) - student_id: {student.id}, year: {student_update.visa_year}")

    try:
        db.commit()
        db.refresh(student)
        
        # 로그 생성 (변경된 경우에만)
        create_database_log(
            db=db,
            table_name="students",
            record_id=str(student.id),
            action="UPDATE",
            user_id=current_user.get("id") if current_user else None,
            old_values={
                "residence_card_number": student.residence_card_number,
                "residence_card_start": str(student.residence_card_start) if student.residence_card_start else None,
                "residence_card_expiry": str(student.residence_card_expiry) if student.residence_card_expiry else None,
                "visa_year": student.visa_year
            },
            new_values={
                "residence_card_number": student_update.residence_card_number,
                "residence_card_start": student_update.residence_card_start,
                "residence_card_expiry": student_update.residence_card_expiry,
                "visa_year": student_update.visa_year
            },
            changed_fields=["residence_card_number", "residence_card_start", "residence_card_expiry", "visa_year"],
            note="学生情報更新"
        )
      
        # UUID를 문자열로 변환
        student_dict = {
            "id": str(student.id),
            "name": student.name,
            "email": student.email if student.email else "",
            "company_id": str(student.company_id) if student.company_id else None,
            "consultant": student.consultant,
            "avatar": student.avatar,
            "created_at": student.created_at
        }
        return student_dict
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"학생 정보 수정 중 오류가 발생했습니다: {str(e)}"
        )

# Residence Card 히스토리 조회
@app.get("/students/{student_id}/residence-card-history")
def get_student_residence_card_history(
    student_id: str,
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(
            status_code=404,
            detail="학생을 찾을 수 없습니다."
        )

    # Residence Card 히스토리 조회
    query = db.query(ResidenceCardHistory).filter(
        ResidenceCardHistory.student_id == student_id
    ).order_by(ResidenceCardHistory.registered_at.desc())

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    histories = query.offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for history in histories:
        history_data = {
            "id": str(history.id),
            "student_id": str(history.student_id),
            "card_number": history.card_number,
            "start_date": history.start_date,
            "expiry_date": history.expiry_date,
            "year": history.year,
            "registered_at": history.registered_at,
            "note": history.note
        }
        result.append(history_data)

    return {
        "items": result,
        "total": total_count,
        "page": page,
        "size": size,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_previous": page > 1
    }

# 비자 정보만 변경하는 API
@app.put("/students/{student_id}/visa-info")
def update_student_visa_info(
    student_id: str,
    visa_update: VisaInfoUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(
            status_code=404,
            detail="학생을 찾을 수 없습니다."
        )

    # 헬퍼 함수들
    def clean_value(value):
        if value == "" or value is None:
            return None
        return value
    
    def parse_date(value):
        if value is None or value == "":
            return None
        if isinstance(value, date):
            return value
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except:
            return None

    # Residence Card 정보가 변경되었는지 확인
    residence_card_changed = False
    if (visa_update.residence_card_number and 
        visa_update.residence_card_start and 
        visa_update.residence_card_expiry):
        
        # 기존 정보와 비교
        if (student.residence_card_number != visa_update.residence_card_number or
            student.residence_card_start != parse_date(visa_update.residence_card_start) or
            student.residence_card_expiry != parse_date(visa_update.residence_card_expiry)):
            residence_card_changed = True
            
            # 같은 년차인지 확인
            if visa_update.year:
                # 기존 히스토리에서 같은 년차가 있는지 확인
                existing_history = db.query(ResidenceCardHistory).filter(
                    ResidenceCardHistory.student_id == student_id,
                    ResidenceCardHistory.year == visa_update.year
                ).first()
                
                if existing_history:
                    raise HTTPException(
                        status_code=400,  
                        detail=f"{visa_update.year}년차의 residence card가 이미 존재합니다. 다른 년차를 입력해주세요."
                    )

    # 학생 비자 정보 업데이트
    update_data = visa_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        if field == "note":  # note는 학생 테이블에 저장하지 않음
            continue
        if field in ["residence_card_start", "residence_card_expiry", "passport_expiration_date", "visa_application_date"]:
            setattr(student, field, parse_date(value))
        elif field == "residence_card_number":
            setattr(student, field, clean_value(value))
        elif field == "year":
            setattr(student, field, clean_value(value))
        else:
            setattr(student, field, clean_value(value))

    # Residence Card 히스토리 저장 (변경된 경우)
    if residence_card_changed:
        start_date = parse_date(visa_update.residence_card_start)
        
        residence_card_history = ResidenceCardHistory(
            id=str(uuid.uuid4()),
            student_id=student.id,
            card_number=visa_update.residence_card_number,
            start_date=start_date,
            expiry_date=parse_date(visa_update.residence_card_expiry),
            year=visa_update.year,
            note=visa_update.note if visa_update.note else "ビザ情報更新"
        )
        db.add(residence_card_history)
        print(f"DEBUG: Residence Card 히스토리 생성 완료 (비자 정보 업데이트) - student_id: {student.id}, year: {visa_update.year}")

    try:
        db.commit()
        db.refresh(student)
        
    except Exception as e:
        print(f"DEBUG: 커밋 중 오류 - {str(e)}")
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"데이터베이스 저장 중 오류가 발생했습니다: {str(e)}"
        )
    
    # 로그 생성 (별도 트랜잭션, 변경된 경우에만)
    if residence_card_changed:
        try:
            create_database_log(
                db=db,
                table_name="students",
                record_id=str(student.id),
                action="UPDATE",
                user_id=current_user.id if current_user else None,
                old_values={
                    "residence_card_number": student.residence_card_number,
                    "residence_card_start": str(student.residence_card_start) if student.residence_card_start else None,
                    "residence_card_expiry": str(student.residence_card_expiry) if student.residence_card_expiry else None,
                    "visa_year": student.visa_year,
                    "passport_number": student.passport_number,
                    "passport_expiration_date": str(student.passport_expiration_date) if student.passport_expiration_date else None,
                    "visa_application_date": str(student.visa_application_date) if student.visa_application_date else None
                },
                new_values={
                    "residence_card_number": visa_update.residence_card_number,
                    "residence_card_start": visa_update.residence_card_start,
                    "residence_card_expiry": visa_update.residence_card_expiry,
                    "visa_year": visa_update.year,
                    "passport_number": visa_update.passport_number if hasattr(visa_update, 'passport_number') else student.passport_number,
                    "passport_expiration_date": visa_update.passport_expiration_date if hasattr(visa_update, 'passport_expiration_date') else student.passport_expiration_date,
                    "visa_application_date": visa_update.visa_application_date
                },
                changed_fields=["residence_card_number", "residence_card_start", "residence_card_expiry", "visa_year", "passport_number", "passport_expiration_date", "visa_application_date"],
                note="ビザ情報更新"
            )
        except Exception as log_error:
            print(f"로그 생성 중 오류: {str(log_error)}")
            # 로그 생성 실패는 학생 수정에 영향을 주지 않도록 함
    
    # 응답 데이터 준비
    response_data = {
        "id": str(student.id),
        "name": student.name,
        "residence_card_number": student.residence_card_number,
        "residence_card_start": student.residence_card_start,
        "residence_card_expiry": student.residence_card_expiry,
        "passport_number": student.passport_number,
        "passport_expiration_date": student.passport_expiration_date,
        "visa_application_date": student.visa_application_date,
        "message": "ビザ情報が正常に更新されました。"
    }
    
    if residence_card_changed:
        response_data["history_created"] = True
        response_data["history_note"] = visa_update.note if visa_update.note else ""
    
    return response_data

# 데이터베이스 로그 조회 API
@app.get("/database-logs")
def get_database_logs(
    table_name: Optional[str] = Query(None, description="테이블 이름으로 필터링"),
    action: Optional[str] = Query(None, description="작업 유형으로 필터링 (CREATE, UPDATE, DELETE)"),
    user_id: Optional[str] = Query(None, description="사용자 ID로 필터링"),
    record_id: Optional[str] = Query(None, description="레코드 ID로 필터링"),
    start_date: Optional[str] = Query(None, description="시작 날짜 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="종료 날짜 (YYYY-MM-DD)"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """데이터베이스 로그 조회"""
    # 기본 쿼리 생성
    query = db.query(DatabaseLog)
    
    # 필터링 적용
    if table_name:
        query = query.filter(DatabaseLog.table_name.ilike(f"%{table_name}%"))
    
    if action:
        query = query.filter(DatabaseLog.action == action)
    
    if user_id:
        query = query.filter(DatabaseLog.user_id == user_id)
    
    if record_id:
        query = query.filter(DatabaseLog.record_id == record_id)
    
    if start_date:
        try:
            start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(DatabaseLog.created_at >= start_datetime)
        except ValueError:
            raise HTTPException(status_code=400, detail="시작 날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식으로 입력해주세요.")
    
    if end_date:
        try:
            end_datetime = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(DatabaseLog.created_at < end_datetime)
        except ValueError:
            raise HTTPException(status_code=400, detail="종료 날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식으로 입력해주세요.")
    
    # 최신순으로 정렬
    query = query.order_by(DatabaseLog.created_at.desc())
    
    # 전체 항목 수 계산
    total_count = query.count()
    
    # 페이지네이션 적용
    logs = query.offset((page - 1) * size).limit(size).all()
    
    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size
    
    # 응답 데이터 준비
    result = []
    for log in logs:
        log_data = {
            "id": str(log.id),
            "table_name": log.table_name,
            "record_id": log.record_id,
            "action": log.action,
            "user_id": log.user_id,
            "old_values": log.old_values,
            "new_values": log.new_values,
            "changed_fields": log.changed_fields,
            "ip_address": log.ip_address,
            "user_agent": log.user_agent,
            "created_at": log.created_at,
            "note": log.note
        }
        result.append(log_data)
    
    return {
        "items": result,
        "total": total_count,
        "page": page,
        "size": size,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_previous": page > 1
    }

@app.get("/users")
def read_users(db: Session = Depends(get_db)):
    return db.query(User).all()

@app.post("/users")
def create_user(
    user: dict, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    new_user = User(**user)
    db.add(new_user)
    
    try:
        db.commit()
        db.refresh(new_user)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="users",
            record_id=str(new_user.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "username": new_user.username,
                "email": new_user.email,
                "full_name": new_user.full_name,
                "is_active": new_user.is_active,
                "is_superuser": new_user.is_superuser
            },
            note="ユーザー新規登録"
        )
        
        return new_user
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"사용자 생성 중 오류가 발생했습니다: {str(e)}"
        )

# 모든 회사 목록 조회
@app.get("/companies")
def get_companies(db: Session = Depends(get_db)):
    companies = db.query(Company).all()
    return companies

# 특정 회사 조회
@app.get("/companies/{company_id}")
def get_company(company_id: str, db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if company is None:
        raise HTTPException(status_code=404, detail="회사를 찾을 수 없습니다")
    return company

# 회사 검색
@app.get("/companies/search/{keyword}")
def search_companies(keyword: str, db: Session = Depends(get_db)):
    companies = db.query(Company).filter(
        Company.name.ilike(f"%{keyword}%")
    ).all()
    return companies

# 새 회사 생성
@app.post("/companies")
def create_company(
    company: dict, 
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    new_company = Company(**company)
    db.add(new_company)
    
    try:
        db.commit()
        db.refresh(new_company)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="companies",
            record_id=str(new_company.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "name": new_company.name,
                "address": new_company.address,
                "phone": new_company.phone,
                "email": new_company.email,
                "contact_person": new_company.contact_person
            },
            note="会社新規登録"
        )
        
        return new_company
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"회사 생성 중 오류가 발생했습니다: {str(e)}"
        )

# 파일 업로드 엔드포인트
@app.post("/upload-avatar")
async def upload_avatar(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # 파일 확장자 검사
        file_extension = file.filename.split(".")[-1].lower()
        if file_extension not in ["jpg", "jpeg", "png", "gif"]:
            raise HTTPException(
                status_code=400,
                detail="지원하지 않는 파일 형식입니다. jpg, jpeg, png, gif 파일만 업로드 가능합니다."
            )

        # 파일 크기 검사 (예: 5MB)
        file_size = 0
        chunk_size = 1024 * 1024  # 1MB
        while chunk := await file.read(chunk_size):
            file_size += len(chunk)
            if file_size > 5 * 1024 * 1024:  # 5MB
                raise HTTPException(
                    status_code=400,
                    detail="파일 크기는 5MB를 초과할 수 없습니다."
                )

        # 파일을 다시 처음으로 되돌림
        await file.seek(0)

        # Supabase Storage에 파일 업로드
        file_path = f"avatars/{current_user.id}/{file.filename}"
        file_content = await file.read()
        
        try:
            print(f"file_path: {file_path}")
            print(f"file_content type: {type(file_content)}, size: {len(file_content)}")
            print(f"file content-type: {file.content_type}")
            result = supabase.storage.from_("avatars").upload(
                file_path,
                file_content,
                {"content-type": file.content_type}
            )
            print(f"Storage upload result: {result}")
        except Exception as storage_error:
            print(f"Storage 업로드 에러: {storage_error}")
            raise HTTPException(
                status_code=500,
                detail=f"Storage 연결 오류: {str(storage_error)}"
            )

        # 파일 URL 가져오기
        file_url = supabase.storage.from_("avatars").get_public_url(file_path)
        return {
            "message": "プロフィール画像が正常にアップロードされました。",
            "avatar_url": file_url
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"파일 업로드 중 오류가 발생했습니다: {str(e)}"
        )

@app.post("/students/{student_id}/changeAvatar")
async def upload_student_avatar(
    student_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # 학생 존재 여부 확인
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(
                status_code=404,
                detail="학생을 찾을 수 없습니다."
            )

        # 파일 확장자 검사
        file_extension = file.filename.split(".")[-1].lower()
        if file_extension not in ["jpg", "jpeg", "png", "gif"]:
            raise HTTPException(
                status_code=400,
                detail="지원하지 않는 파일 형식입니다. jpg, jpeg, png, gif 파일만 업로드 가능합니다."
            )

        # 파일 크기 검사 (예: 5MB)
        file_size = 0
        chunk_size = 1024 * 1024  # 1MB
        while chunk := await file.read(chunk_size):
            file_size += len(chunk)
            if file_size > 5 * 1024 * 1024:  # 5MB
                raise HTTPException(
                    status_code=400,
                    detail="파일 크기는 5MB를 초과할 수 없습니다."
                )

        # 파일을 다시 처음으로 되돌림
        await file.seek(0)

        # Supabase Storage에 파일 업로드
        file_path = f"student_avatars/{student_id}/{file.filename}"
        file_content = await file.read()
        
        try:
            print(f"file_path: {file_path}")
            print(f"file_content type: {type(file_content)}, size: {len(file_content)}")
            print(f"file content-type: {file.content_type}")
            
            # service role 클라이언트를 사용하여 Storage 업로드
            result = supabase_storage.storage.from_("avatars").upload(
                file_path,
                file_content,
                {"content-type": file.content_type}
            )
            print(f"Storage upload result: {result}")
            
            # 파일 URL 생성 (anon 클라이언트 사용)
            file_url = supabase.storage.from_("avatars").get_public_url(file_path)
            
            # 학생의 avatar 필드 업데이트
            student.avatar = file_url
            db.commit()
            
        except Exception as storage_error:
            print(f"Storage 업로드 에러: {storage_error}")
            raise HTTPException(
                status_code=500,
                detail=f"Storage 연결 오류: {str(storage_error)}"
            )
        return {
            "message": "学生のプロフィール画像が正常にアップロードされました。",
            "avatar_url": file_url
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"파일 업로드 중 오류가 발생했습니다: {str(e)}"
        )

# 모든 등급 목록 조회
@app.get("/grades")
def get_grades(db: Session = Depends(get_db)):
    grades = db.query(Grade).all()
    return grades

@app.post("/invoices", response_model=InvoiceResponse)
async def create_invoice(
    invoice: InvoiceCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == invoice.student_id).first()
    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="학생을 찾을 수 없습니다."
        )

    # 해당 월의 청구서가 이미 존재하는지 확인
    existing_invoice = db.query(Invoice).filter(
        Invoice.student_id == invoice.student_id,
        Invoice.year == invoice.year,
        Invoice.month == invoice.month
    ).first()
    
    if existing_invoice:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="해당 월의 청구서가 이미 존재합니다."
        )

    try:
        # 청구서 항목들의 amount 합계 계산
        total_amount = sum(item.amount for item in invoice.items)

        # 해당 월의 마지막 invoice 번호 조회
        last_invoice_count = db.query(Invoice).filter(
            Invoice.year == invoice.year,
            Invoice.month == invoice.month
        ).count()
        
        # 청구서 생성
        new_invoice = Invoice(
            student_id=invoice.student_id,
            year=invoice.year,
            month=invoice.month,
            total_amount=total_amount,  # 계산된 총액 사용
            invoice_number=generate_invoice_number(datetime(invoice.year, invoice.month, 1), last_invoice_count),
            status="SAVED"
        )
        db.add(new_invoice)
        db.flush()  # ID 생성을 위해 flush

        # 청구서 항목 생성
        for item in invoice.items:
            new_item = InvoiceItem(
                invoice_id=new_invoice.id,
                name=item.name,
                unit_price=item.unit_price,
                quantity=item.quantity,  # quantity 필드 추가
                amount=item.amount,
                sort_order=item.sort_order,
                memo=item.memo,
                type=item.type
            )
            db.add(new_item)
        
        db.commit()
        db.refresh(new_invoice)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="invoices",
            record_id=str(new_invoice.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "student_id": str(new_invoice.student_id),
                "year": new_invoice.year,
                "month": new_invoice.month,
                "total_amount": new_invoice.total_amount,
                "invoice_number": new_invoice.invoice_number,
                "status": new_invoice.status
            },
            note="請求書新規作成"
        )
        
        return new_invoice

    except IntegrityError as e:
      db.rollback()
      print("DEBUG: IntegrityError:", e)
      raise HTTPException(
          status_code=status.HTTP_400_BAD_REQUEST,
          detail=f"청구서 생성 중 오류: {str(e.orig)}"  # 실제 DB에서 온 메시지
      )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )
    
@app.get("/invoices/{invoice_id}/preview")
def get_invoice_preview(
    invoice_id: str, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """청구서 미리보기 데이터 조회 (PDF 다운로드 전 프론트 표시용)"""
    # 청구서 존재 여부 확인
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="해당 청구서를 찾을 수 없습니다.")
    
    # 청구서 아이템 조회
    invoice_items = db.query(InvoiceItem).filter(
        InvoiceItem.invoice_id == invoice_id
    ).order_by(InvoiceItem.sort_order).all()
    
    # 학생 정보 조회
    student = db.query(Student).filter(Student.id == invoice.student_id).first()
    
    # 응답 데이터 구성
    invoice_data = {
        "id": str(invoice.id),
        "invoice_number": invoice.invoice_number,
        "student_id": str(invoice.student_id),
        "student_name": student.name if student else "알 수 없음",
        "student_name_katakana": student.name_katakana if student else "",
        "total_amount": invoice.total_amount,
        "note": getattr(invoice, "note", ""),
        "created_at": invoice.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "issue_date": datetime.now().strftime("%Y年%m月%d日 %H:%M"),
        "items": [
            {
                "id": str(item.id),
                "name": item.name,
                "unit_price": item.unit_price,
                "quantity": item.quantity,
                "amount": item.amount,
                "sort_order": item.sort_order
            }
            for item in invoice_items
        ]
    }
    
    return invoice_data

@app.get("/invoices/{invoice_id}/pdf")
def download_invoice_pdf(invoice_id: str, db: Session = Depends(get_db)):
    # 1. DB에서 invoice, invoice_items 조회 (이전 답변 참고)
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="해당 청구서를 찾을 수 없습니다.")
    invoice_items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice_id).order_by(InvoiceItem.sort_order).all()

    invoice_dict = {
        "customer_name": '테스트',
        "date": invoice.created_at.strftime("%Y年%m月%d日"),
        "total": invoice.total_amount,
        "note": getattr(invoice, "note", "")
    }
    items_list = [
        {
            "name": item.name,
            "unit_price": item.unit_price,
            "quantity": item.quantity,
            "amount": item.amount
        }
        for item in invoice_items
    ]
    issue_date = datetime.now().strftime("%Y年%m月%d日 %H:%M")
    # 2. HTML 렌더링
    html_content = templates.get_template("invoice.html").render(invoice=invoice_dict, items=items_list, issue_date=issue_date)

    # 3. HTML → PDF 변환
    pdf_bytes = html_to_pdf_bytes(html_content)

    # 4. PDF 파일로 다운로드
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice_{invoice_id}.pdf"}
    )

def html_to_pdf_bytes(html_content: str) -> bytes:
    pdf = HTML(string=html_content).write_pdf()
    return pdf

@app.get("/billing-options")
def billing_options(db: Session = Depends(get_db)):
    return get_billing_options(db)

def get_billing_options(db: Session):
    # 데이터 조회
    res = db.query(BillingItem).all()

    # groupOptions, individualItems 분리
    groupOptions = {}
    individualItems = []

    for row in res:
        item = {
            "name": row.name,
            "unit": row.unit,
            "qna": row.qna,
            "billing_type": row.billing_type
        }
        # value 필드 있으면 추가
        if row.value:
            item["value"] = row.value
        if row.group_type:
            group = row.group_type.lower()
            if group not in groupOptions:
                groupOptions[group] = []
            groupOptions[group].append(item)
        else:
            # individual 항목은 amount 필드 추가
            item["amount"] = row.unit
            individualItems.append(item)
    return {
        "groupOptions": groupOptions,
        "individualItems": individualItems
    }

def generate_invoice_number(date, last_num):
    # date: 2024-06-16, last_num: 7 (이번 달 마지막 번호)
    ym = date.strftime("%Y%m")
    return f"{ym}-{last_num+1:03d}"

# 예) 2024년 6월 8번째 발급이면: 202406-008

@app.get("/invoices")
def get_invoices(
    student_id: Optional[str] = Query(None, description="학생 ID로 검색"),
    page: int = Query(1, description="페이지 번호", ge=1),
    page_size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 기본 쿼리 생성
    query = db.query(Invoice).options(
        joinedload(Invoice.student)
    )

    # student_id가 제공된 경우 해당 학생의 청구서만 필터링
    if student_id:
        query = query.filter(Invoice.student_id == student_id)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    invoices = query.order_by(Invoice.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + page_size - 1) // page_size

    return {
        "items": invoices,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_previous": page > 1
    }

@app.get("/company-invoice-pdf/{company_id}/{year}/{month}")
def generate_company_invoice_pdf(
    company_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 회사 정보 조회
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="会社が見つかりません")

    # 해당 회사 소속 학생들의 청구서 조회
    invoices = db.query(Invoice).join(Student).options(
        joinedload(Invoice.student),
        joinedload(Invoice.items)  # InvoiceItem 정보도 함께 로드
    ).filter(
        Student.company_id == company_id,
        Invoice.year == year,
        Invoice.month == month
    ).all()

    if not invoices:
        raise HTTPException(status_code=404, detail="該当月の請求書がありません")

    # 총 금액 계산
    total_amount = sum(invoice.total_amount for invoice in invoices)

    # 청구서 데이터 준비
    invoice_data = {
        "company_name": company.name,
        "year": year,
        "month": month,
        "total_amount": total_amount,
        "vat_total_amount": int(total_amount + total_amount * 0.1),
        "vat": int(total_amount * 0.1),
        "issue_date": datetime.now().strftime("%Y年%m月%d日"),
        "invoice_list": [{
            "student_name": invoice.student.name,
            "amount": invoice.total_amount,
            "invoice_number": invoice.invoice_number,
            "invoice_items": [{
                "name": item.name,
                "unit_price": item.unit_price,
                "amount": item.amount,
                "memo": item.memo,
                "type": item.type
            } for item in sorted(invoice.items, key=lambda x: x.sort_order)]
        } for invoice in invoices]
    }

    # HTML 템플릿 렌더링
    html_content = templates.get_template("company_invoice.html").render(
        data=invoice_data
    )

    # PDF 생성
    pdf_bytes = html_to_pdf_bytes(html_content)

    # PDF 파일 반환
    filename = f"company_invoice_{company_id}_{year}{month:02d}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/pdf"
        }
    )

@app.get("/invoice/{student_id}/{year}/{month}")
def get_student_invoice_by_year_month(
    student_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 청구서 조회
    invoice = db.query(Invoice).options(
        joinedload(Invoice.student).joinedload(Student.company),
        joinedload(Invoice.items)
    ).filter(
        Invoice.student_id == student_id,
        Invoice.year == year,
        Invoice.month == month
    ).first()

    if not invoice:
        result = {
            "is_data": False
        }
    else:
        # 응답 데이터 준비
        result = {
          "id": str(invoice.id),
          "is_data": True,
          "invoice_number": invoice.invoice_number,
          "year": invoice.year,
          "month": invoice.month,
          "total_amount": invoice.total_amount,
          "created_at": invoice.created_at,
          "student": {
              "id": str(invoice.student.id),
              "name": invoice.student.name,
              "company": {
                  "id": str(invoice.student.company.id),
                  "name": invoice.student.company.name
              } if invoice.student.company else None
          },
          "items": [{
              "name": item.name,
              "unit_price": item.unit_price,
              "amount": item.amount,
              "quantity": item.quantity,
              "memo": item.memo,
              "type": item.type
          } for item in sorted(invoice.items, key=lambda x: x.sort_order)]
      }

    return result

@app.get("/invoice/{invoice_id}")
def get_invoice_by_id(
    invoice_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
   
    # 청구서 조회
    invoice = db.query(Invoice).options(
        joinedload(Invoice.student).joinedload(Student.company),
        joinedload(Invoice.items)
    ).filter(
        Invoice.id == invoice_id,
    ).first()

    if not invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="청구서를 찾을 수 없습니다."
        )
    else:
        # 응답 데이터 준비
        result = {
          "id": str(invoice.id),
          "is_data": True,
          "invoice_number": invoice.invoice_number,
          "year": invoice.year,
          "month": invoice.month,
          "total_amount": invoice.total_amount,
          "created_at": invoice.created_at,
          "student": {
              "id": str(invoice.student.id),
              "name": invoice.student.name,
              "company": {
                  "id": str(invoice.student.company.id),
                  "name": invoice.student.company.name
              } if invoice.student.company else None
          },
          "items": [{
              "name": item.name,
              "unit_price": item.unit_price,
              "amount": item.amount,
              "quantity": item.quantity,
              "memo": item.memo,
              "type": item.type
          } for item in sorted(invoice.items, key=lambda x: x.sort_order)]
      }

    return result

@app.put("/invoice")
async def update_invoice(
    invoice_update: InvoiceUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 청구서 존재 여부 확인
    existing_invoice = db.query(Invoice).filter(Invoice.id == invoice_update.invoice_id).first()
    if not existing_invoice:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="청구서를 찾을 수 없습니다."
        )

    # 기존 값 저장 (로그용)
    old_values = {
        "student_id": str(existing_invoice.student_id),
        "year": existing_invoice.year,
        "month": existing_invoice.month,
        "total_amount": existing_invoice.total_amount,
        "invoice_number": existing_invoice.invoice_number,
        "status": existing_invoice.status
    }
    
    try:
        # 청구서 항목들의 amount 합계 계산
        total_amount = sum(item.amount for item in invoice_update.items)

        # 기존 invoice_items 삭제
        db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice_update.invoice_id).delete()

        # 청구서 정보 업데이트
        existing_invoice.year = invoice_update.year
        existing_invoice.month = invoice_update.month
        existing_invoice.total_amount = total_amount

        # 새로운 청구서 항목 생성
        for item in invoice_update.items:
            new_item = InvoiceItem(
                invoice_id=invoice_update.invoice_id,
                name=item.name,
                unit_price=item.unit_price,
                quantity=item.quantity,
                amount=item.amount,
                sort_order=item.sort_order,
                memo=item.memo,
                type=item.type
            )
            db.add(new_item)
        
        db.commit()
        db.refresh(existing_invoice)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="invoices",
            record_id=str(existing_invoice.id),
            action="UPDATE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            new_values={
                "student_id": str(existing_invoice.student_id),
                "year": existing_invoice.year,
                "month": existing_invoice.month,
                "total_amount": existing_invoice.total_amount,
                "invoice_number": existing_invoice.invoice_number,
                "status": existing_invoice.status
            },
            changed_fields=["year", "month", "total_amount"],
            note="請求書情報更新"
        )
        
        return existing_invoice

    except IntegrityError as e:
        db.rollback()
        print("DEBUG: IntegrityError:", e)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"청구서 수정 중 오류: {str(e.orig)}"
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )

@app.get("/buildings")
def get_buildings(
    name: Optional[str] = Query(None, description="빌딩 이름으로 검색"),
    address: Optional[str] = Query(None, description="주소로 검색"),
    resident_type: Optional[str] = Query(None, description="거주자 타입으로 검색"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 기본 쿼리 생성
    query = db.query(Building)

    # 필터 조건 추가
    if name:
        query = query.filter(Building.name.ilike(f"%{name}%"))
    if address:
        query = query.filter(Building.address.ilike(f"%{address}%"))
    if resident_type:
        query = query.filter(Building.resident_type == resident_type)
    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    buildings = query.offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    return {
        "items": buildings,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/buildings/options")
def get_building_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """빌딩 목록을 셀렉트 옵션용으로 조회"""
    buildings = db.query(Building).order_by(Building.name).all()
    
    options = []
    for building in buildings:
        # 각 빌딩의 빈 호실 수 계산
        empty_rooms_count = db.query(Room).filter(
            Room.building_id == building.id,
            Room.is_available == True
        ).count()
        
        building_option = {
            "value": str(building.id),
            "label": building.name,
            "address": building.address,
            "empty_rooms_count": empty_rooms_count
        }
        options.append(building_option)
    
    return {
        "options": options,
        "total": len(options)
    }

@app.get("/buildings/{building_id}")
def get_building(
    building_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="빌딩을 찾을 수 없습니다")
    return building

@app.post("/buildings", response_model=BuildingResponse)
def create_building(
    building: BuildingCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    new_building = Building(
        id=str(uuid.uuid4()),
        name=building.name,
        address=building.address,
        building_type=building.building_type,
        total_rooms=building.total_rooms,
        note=building.note
    )
    db.add(new_building)
    
    try:
        db.commit()
        db.refresh(new_building)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="buildings",
            record_id=str(new_building.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "name": new_building.name,
                "address": new_building.address,
                "building_type": new_building.building_type,
                "total_rooms": new_building.total_rooms,
                "note": new_building.note
            },
            note="ビル新規登録"
        )
        
        return new_building
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"빌딩 생성 중 오류가 발생했습니다: {str(e)}"
        )

@app.put("/buildings/{building_id}", response_model=BuildingResponse)
def update_building(
    building_id: str,
    building_update: BuildingUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 빌딩 존재 여부 확인
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="빌딩을 찾을 수 없습니다")

    # 기존 값 저장 (로그용)
    old_values = {
        "name": building.name,
        "address": building.address,
        "building_type": building.building_type,
        "total_rooms": building.total_rooms,
        "note": building.note
    }
    
    # 빌딩 정보 업데이트
    update_data = building_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(building, field, value)

    try:
        db.commit()
        db.refresh(building)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="buildings",
            record_id=str(building.id),
            action="UPDATE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            new_values={
                "name": building.name,
                "address": building.address,
                "building_type": building.building_type,
                "total_rooms": building.total_rooms,
                "note": building.note
            },
            changed_fields=list(update_data.keys()),
            note="ビル情報更新"
        )
        
        return building
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"빌딩 정보 수정 중 오류가 발생했습니다: {str(e)}"
        )

@app.delete("/buildings/{building_id}")
def delete_building(
    building_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 빌딩 존재 여부 확인
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="빌딩을 찾을 수 없습니다")

    # 삭제 전 값 저장 (로그용)
    old_values = {
        "name": building.name,
        "address": building.address,
        "total_rooms": building.total_rooms,
        "note": building.note
    }
    
    try:
        db.delete(building)
        db.commit()
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="buildings",
            record_id=str(building.id),
            action="DELETE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            note="ビル削除"
        )
        
        return {"message": "部屋が正常に削除されました"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"빌딩 삭제 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/buildings/{building_id}/rooms")
def get_rooms_by_building(
    building_id: str,
    room_number: Optional[str] = Query(None, description="방 번호로 검색"),
    is_available: Optional[bool] = Query(None, description="사용 가능 여부로 필터링"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(100, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 빌딩 존재 여부 확인
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="빌딩을 찾을 수 없습니다")

    # 기본 쿼리 생성
    query = db.query(Room).filter(Room.building_id == building_id)

    # 필터 조건 추가
    if room_number:
        query = query.filter(Room.room_number.ilike(f"%{room_number}%"))
    if is_available is not None:
        query = query.filter(Room.is_available == is_available)

    # room_number 기준으로 정렬 (숫자 정렬을 위해 CAST 사용)
    query = query.order_by(Room.room_number)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    rooms = query.offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 방 정보와 함께 BuildingCategoriesRent 정보 포함
    room_data = []
    for room in rooms:
        room_dict = {
            "id": str(room.id),
            "building_id": str(room.building_id),
            "room_number": room.room_number,
            "rent": room.rent,
            "maintenance": room.maintenance,
            "service": room.service,
            "floor": room.floor,
            "capacity": room.capacity,
            "is_available": room.is_available,
            "note": room.note,
            "security_deposit": room.security_deposit,
            "monthly_rent": None
        }
        
        # 해당 빌딩의 BuildingCategoriesRent 정보 조회
        building_status_rent = db.query(BuildingCategoriesRent).filter(
            BuildingCategoriesRent.building_id == room.building_id,
            BuildingCategoriesRent.categories_id == 1
        ).first()
        
        if building_status_rent:
            room_dict["monthly_rent"] = building_status_rent.monthly_rent
        
        room_data.append(room_dict)

    return {
        "items": room_data,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/rooms/{room_id}")
def get_room(
    room_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    room = db.query(Room).options(
        joinedload(Room.building)
    ).filter(Room.id == room_id).first()
    
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")
    
    # 방 정보와 함께 빌딩의 resident_type 포함
    room_data = {
        "id": str(room.id),
        "building_id": str(room.building_id),
        "room_number": room.room_number,
        "rent": room.rent,
        "maintenance": room.maintenance,
        "service": room.service,
        "floor": room.floor,
        "capacity": room.capacity,
        "is_available": room.is_available,
        "note": room.note,
        "building": {
            "id": str(room.building.id),
            "name": room.building.name,
            "address": room.building.address,
            "total_rooms": room.building.total_rooms,
            "note": room.building.note,
            "resident_type": room.building.resident_type
        }
    }
    
    return room_data

@app.post("/rooms", response_model=RoomResponse)
def create_room(
    room: RoomCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 빌딩 존재 여부 확인
    building = db.query(Building).filter(Building.id == room.building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="빌딩을 찾을 수 없습니다")

    # 같은 빌딩 내에서 방 번호 중복 확인
    existing_room = db.query(Room).filter(
        Room.building_id == room.building_id,
        Room.room_number == room.room_number
    ).first()
    
    if existing_room:
        raise HTTPException(status_code=400, detail="해당 빌딩에 같은 방 번호가 이미 존재합니다")

    # 정원 유효성 검사
    if room.capacity is not None and room.capacity <= 0:
        raise HTTPException(status_code=400, detail="정원은 1명 이상이어야 합니다")

    new_room = Room(
        id=str(uuid.uuid4()),
        building_id=room.building_id,
        room_number=room.room_number,
        rent=room.rent,
        floor=room.floor,
        capacity=room.capacity,
        is_available=room.is_available,
        note=room.note
    )
    db.add(new_room)
    
    try:
        db.commit()
        db.refresh(new_room)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="rooms",
            record_id=str(new_room.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "building_id": str(new_room.building_id),
                "room_number": new_room.room_number,
                "rent": new_room.rent,
                "floor": new_room.floor,
                "capacity": new_room.capacity,
                "is_available": new_room.is_available,
                "note": new_room.note
            },
            note="部屋新規登録"
        )
        
        return new_room
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"방 생성 중 오류가 발생했습니다: {str(e)}"
        )

@app.put("/rooms/{room_id}", response_model=RoomResponse)
def update_room(
    room_id: str,
    room_update: RoomUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 방 번호가 변경되는 경우 중복 확인
    if room_update.room_number and room_update.room_number != room.room_number:
        existing_room = db.query(Room).filter(
            Room.building_id == room.building_id,
            Room.room_number == room_update.room_number,
            Room.id != room_id
        ).first()
        
        if existing_room:
            raise HTTPException(status_code=400, detail="해당 빌딩에 같은 방 번호가 이미 존재합니다")

    # 정원 유효성 검사
    if room_update.capacity is not None and room_update.capacity <= 0:
        raise HTTPException(status_code=400, detail="정원은 1명 이상이어야 합니다")

    # 기존 값 저장 (로그용)
    old_values = {
        "building_id": str(room.building_id),
        "room_number": room.room_number,
        "rent": room.rent,
        "floor": room.floor,
        "capacity": room.capacity,
        "is_available": room.is_available,
        "note": room.note
    }
    
    # 방 정보 업데이트
    update_data = room_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(room, field, value)

    try:
        db.commit()
        db.refresh(room)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="rooms",
            record_id=str(room.id),
            action="UPDATE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            new_values={
                "building_id": str(room.building_id),
                "room_number": room.room_number,
                "rent": room.rent,
                "floor": room.floor,
                "capacity": room.capacity,
                "is_available": room.is_available,
                "note": room.note
            },
            changed_fields=list(update_data.keys()),
            note="部屋情報更新"
        )
        
        return room
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"방 정보 수정 중 오류가 발생했습니다: {str(e)}"
        )

@app.delete("/rooms/{room_id}")
def delete_room(
    room_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 삭제 전 값 저장 (로그용)
    old_values = {
        "building_id": str(room.building_id),
        "room_number": room.room_number,
        "rent": room.rent,
        "floor": room.floor,
        "capacity": room.capacity,
        "is_available": room.is_available,
        "note": room.note
    }
    
    try:
        db.delete(room)
        db.commit()
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="rooms",
            record_id=str(room.id),
            action="DELETE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            note="部屋削除"
        )
        
        return {"message": "部屋が正常に削除されました"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"방 삭제 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/rooms/{room_id}/residents")
def get_room_residents(
    room_id: str,
    resident_type: Optional[str] = Query(None, description="거주자 타입으로 필터링 (student 또는 elderly)"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 해당 방에 실제 입주 중인 거주자들 조회 (Resident 테이블 기준)
    query = db.query(Resident).filter(
        Resident.room_id == room_id,
        Resident.is_active == True,
        Resident.check_out_date.is_(None)  # 퇴실일이 없는 경우 (현재 입주 중)
    )
    
    # 거주자 타입에 따라 조인 설정
    if resident_type == "elderly":
        query = query.options(joinedload(Resident.elderly))
    elif resident_type == "student":
        query = query.options(joinedload(Resident.student))
    else:  # 모든 타입 조회
        query = query.options(joinedload(Resident.student), joinedload(Resident.elderly))

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    residents = query.offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for resident in residents:
        resident_data = {
            "id": str(resident.id),
            "room_id": str(resident.room_id),
            "resident_id": str(resident.resident_id),
            "resident_type": resident.resident_type,
            "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
            "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
            "is_active": resident.is_active,
            "note": resident.note,
            "created_at": resident.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": resident.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "room": {
                "id": str(room.id),
                "room_number": room.room_number,
                "building_id": str(room.building_id)
            }
        }
        
        # 거주자 타입에 따라 다른 데이터 추가
        if resident.resident_type == "elderly" and resident.elderly:
            resident_data["elderly"] = {
                "id": str(resident.elderly.id),
                "name": resident.elderly.name,
                "name_katakana": resident.elderly.name_katakana,
                "gender": resident.elderly.gender,
                "birth_date": resident.elderly.birth_date.strftime("%Y-%m-%d") if resident.elderly.birth_date else None,
                "phone": resident.elderly.phone,
                "care_level": resident.elderly.care_level,
                "status": resident.elderly.status,
                "avatar": resident.elderly.avatar,
            }
        elif resident.resident_type == "student" and resident.student:
            resident_data["student"] = {
                "id": str(resident.student.id),
                "name": resident.student.name,
                "name_katakana": resident.student.name_katakana,
                "nationality": resident.student.nationality,
                "phone": resident.student.phone,
                "email": resident.student.email if resident.student.email else "",
                "avatar": resident.student.avatar,
                "gender": resident.student.gender,
                "birth_date": resident.student.birth_date.strftime("%Y-%m-%d") if resident.student.birth_date else None,
                "japanese_level": resident.student.japanese_level,
                "local_address": resident.student.local_address
            }
        
        result.append(resident_data)

    return {
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/rooms/{room_id}/residents/count")
def get_room_residents_count(
    room_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 해당 방에 실제 입주 중인 학생 수 조회 (Resident 테이블 기준)
    resident_count = db.query(Resident).filter(
        Resident.room_id == room_id,
        Resident.is_active == True,
        Resident.check_out_date.is_(None)  # 퇴실일이 없는 경우 (현재 입주 중)
    ).count()

    return {
        "room_id": room_id,
        "resident_count": resident_count
    }

@app.get("/rooms/{room_id}/residence-history")
def get_room_residence_history(
    room_id: str,
    resident_type: Optional[str] = Query(None, description="거주자 타입으로 필터링 (student 또는 elderly)"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    is_active: Optional[bool] = Query(None, description="활성 상태로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 해당 방의 모든 입주 기록 조회
    query = db.query(Resident)
    
    # 거주자 타입에 따라 조인 설정
    if resident_type == "elderly":
        query = query.options(joinedload(Resident.elderly))
    else:  # student 또는 기본값
        query = query.options(joinedload(Resident.student))
    
    query = query.filter(Resident.room_id == room_id)

    # 활성 상태 필터링 (선택사항)
    if is_active is not None:
        query = query.filter(Resident.is_active == is_active)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용 (최신 기록부터)
    residents = query.order_by(Resident.created_at.desc()).offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for resident in residents:
        resident_data = {
            "id": str(resident.id),
            "room_id": str(resident.room_id),
            "resident_id": str(resident.resident_id),
            "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
            "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
            "is_active": resident.is_active,
            "note": resident.note,
            "created_at": resident.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": resident.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "room": {
                "id": str(room.id),
                "room_number": room.room_number,
                "building_id": str(room.building_id)
            },
            "resident_type": resident.resident_type
        }
        
        # 거주자 타입에 따라 다른 데이터 추가
        if resident_type == "elderly" and resident.elderly:
            resident_data["elderly"] = {
                "id": str(resident.elderly.id),
                "name": resident.elderly.name,
                "name_katakana": resident.elderly.name_katakana,
                "gender": resident.elderly.gender,
                "birth_date": resident.elderly.birth_date.strftime("%Y-%m-%d") if resident.elderly.birth_date else None,
                "phone": resident.elderly.phone,
                "care_level": resident.elderly.care_level,
                "status": resident.elderly.status,
                "avatar": resident.elderly.avatar,
            }
        elif (resident_type == "student" or resident_type is None) and resident.student:
            resident_data["student"] = {
                "id": str(resident.student.id),
                "name": resident.student.name,
                "name_katakana": resident.student.name_katakana,
                "nationality": resident.student.nationality,
                "phone": resident.student.phone,
                "email": resident.student.email if resident.student.email else "",
                "avatar": resident.student.avatar,
                "gender": resident.student.gender,
                "birth_date": resident.student.birth_date.strftime("%Y-%m-%d") if resident.student.birth_date else None,
                "japanese_level": resident.student.japanese_level,
                "local_address": resident.student.local_address
            }
        
        result.append(resident_data)

    return {
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.put("/students/{student_id}/assign-room")
def assign_student_to_room(
    student_id: str,
    request: AssignRoomRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 방 배정 해제인 경우
    if request.room_id is None:
        student.current_room_id = None
        db.commit()
        return {"message": "学生の部屋割り当てが解除されました"}

    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == request.room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 방이 사용 가능한지 확인
    if not room.is_available:
        raise HTTPException(status_code=400, detail="해당 방은 현재 사용할 수 없습니다")

    # 기존 값 저장 (로그용)
    old_values = {
        "current_room_id": str(student.current_room_id) if student.current_room_id else None
    }
    
    # 학생의 방 배정 업데이트
    student.current_room_id = request.room_id
    
    try:
        db.commit()
        db.refresh(student)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="students",
            record_id=str(student.id),
            action="UPDATE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            new_values={
                "current_room_id": str(request.room_id) if request.room_id else None
            },
            changed_fields=["current_room_id"],
            note="学生部屋割り当て" if request.room_id else "学生部屋割り当て解除"
        )
        
        return {
            "message": "学生が正常に部屋に割り当てられました",
            "student_id": str(student.id),
            "room_id": str(request.room_id)
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"방 배정 중 오류가 발생했습니다: {str(e)}"
        )

@app.post("/rooms/{room_id}/check-in")
def check_in_student(
    room_id: str,
    request: CheckInRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == request.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 방이 사용 가능한지 확인
    if not room.is_available:
        raise HTTPException(status_code=400, detail="해당 방은 현재 사용할 수 없습니다")

    # 이미 해당 방에 입주 중인지 확인
    existing_resident = db.query(Resident).filter(
        Resident.resident_id == request.student_id,
        Resident.room_id == room_id,
        Resident.is_active == True,
        Resident.check_out_date.is_(None)
    ).first()
    
    if existing_resident:
        raise HTTPException(status_code=400, detail="該当の学生は既にこの部屋に入居中です")

    try:
        # 입주 기록 생성
        new_resident = Resident(
            id=str(uuid.uuid4()),
            room_id=room_id,
            student_id=request.student_id,
            check_in_date=datetime.strptime(request.check_in_date, "%Y-%m-%d").date(),
            note=request.note
        )
        db.add(new_resident)

        # 학생의 current_room_id 업데이트
        student.current_room_id = room_id
        
        # 방 로그 기록
        room_log = RoomLog(
            id=str(uuid.uuid4()),
            room_id=room_id,
            student_id=request.student_id,
            action="CHECK_IN",
            action_date=datetime.strptime(request.check_in_date, "%Y-%m-%d").date(),
            note=f"入居 - {request.note}" if request.note else "入居"
        )
        db.add(room_log)
        
        db.commit()
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="residents",
            record_id=str(new_resident.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "room_id": str(new_resident.room_id),
                "student_id": str(new_resident.resident_id),
                "check_in_date": str(new_resident.check_in_date),
                "is_active": new_resident.is_active,
                "note": new_resident.note
            },
            note="学生入居"
        )
        
        return {
            "message": "学生が正常に入居しました",
            "student_id": str(request.student_id),
            "room_id": str(room_id),
            "check_in_date": request.check_in_date
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"입주 처리 중 오류가 발생했습니다: {str(e)}"
        )

@app.post("/rooms/{room_id}/check-out")
def check_out_student(
    room_id: str,
    request: CheckOutRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == request.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 해당 방에 입주 중인지 확인
    resident = db.query(Resident).filter(
        Resident.resident_id == request.student_id,
        Resident.room_id == room_id,
        Resident.is_active == True,
        Resident.check_out_date.is_(None)
    ).first()
    
    if not resident:
        raise HTTPException(status_code=400, detail="해당 학생은 이 방에 입주 중이 아닙니다")

    try:
        # 퇴실 처리
        resident.check_out_date = datetime.strptime(request.check_out_date, "%Y-%m-%d").date()
        resident.is_active = False
        if request.note:
            resident.note = request.note

        # 학생의 current_room_id 초기화
        if student.current_room_id == room_id:
            student.current_room_id = None
        
        # 방 로그 기록
        room_log = RoomLog(
            id=str(uuid.uuid4()),
            room_id=room_id,
            student_id=request.student_id,
            action="CHECK_OUT",
            action_date=datetime.strptime(request.check_out_date, "%Y-%m-%d").date(),
            note=f"退去 - {request.note}" if request.note else "退去"
        )
        db.add(room_log)
        
        db.commit()
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="residents",
            record_id=str(resident.id),
            action="UPDATE",
            user_id=current_user["id"] if current_user else None,
            old_values={
                "room_id": str(resident.room_id),
                "student_id": str(resident.resident_id),
                "check_in_date": str(resident.check_in_date),
                "is_active": True,
                "check_out_date": None,
                "note": resident.note
            },
            new_values={
                "room_id": str(resident.room_id),
                "student_id": str(resident.resident_id),
                "check_in_date": str(resident.check_in_date),
                "is_active": False,
                "check_out_date": str(resident.check_out_date),
                "note": resident.note
            },
            changed_fields=["is_active", "check_out_date"],
            note="学生退去"
        )
        
        return {
            "message": "学生が正常に退去しました",
            "student_id": str(request.student_id),
            "room_id": str(room_id),
            "check_out_date": request.check_out_date
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"퇴실 처리 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/rooms/{room_id}/capacity-status")
def get_room_capacity_status(
    room_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 현재 거주자 수 조회
    current_residents = db.query(Resident).filter(
        Resident.room_id == room_id,
        Resident.is_active == True,
        Resident.check_out_date.is_(None)
    ).count()

    # 정원 대비 사용률 계산
    capacity_usage = {
        "room_id": str(room_id),
        "room_number": room.room_number,
        "capacity": room.capacity,
        "current_residents": current_residents,
        "available_spots": room.capacity - current_residents if room.capacity else None,
        "usage_percentage": round((current_residents / room.capacity) * 100, 1) if room.capacity else None,
        "is_full": room.capacity and current_residents >= room.capacity,
        "can_accept_more": room.capacity and current_residents < room.capacity
    }

    return capacity_usage

@app.get("/buildings/{building_id}/empty-rooms")
def get_empty_rooms_by_building(
    building_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 빌딩의 빈 호실 목록을 조회"""
    # 빌딩 존재 여부 확인
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="빌딩을 찾을 수 없습니다")

    # 빈 호실 조회 (사용 가능하고 정원이 남은 방들)
    empty_rooms = db.query(Room).filter(
        Room.building_id == building_id,
        Room.is_available == True
    ).order_by(Room.room_number).all()
    
    options = []
    for room in empty_rooms:
        # 현재 거주자 수 조회
        current_residents = db.query(Resident).filter(
            Resident.room_id == room.id,
            Resident.is_active == True,
            Resident.check_out_date.is_(None)
        ).count()
        
        # 정원 대비 사용 가능 여부 확인
        is_available_for_checkin = room.capacity is None or current_residents < room.capacity
        
        if is_available_for_checkin:
            room_option = {
                "value": str(room.id),
                "label": f"{room.room_number}",
                "room_number": room.room_number,
                "floor": room.floor,
                "capacity": room.capacity,
                "current_residents": current_residents,
                "available_spots": room.capacity - current_residents if room.capacity else None,
                "rent": room.rent,
                "note": room.note
            }
            options.append(room_option)
    
    return {
        "building": {
            "id": str(building.id),
            "name": building.name,
            "address": building.address
        },
        "options": options,
        "total": len(options)
    }

@app.get("/rooms/available")
def get_all_available_rooms(
    building_id: Optional[str] = Query(None, description="특정 빌딩으로 필터링"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """모든 빈 호실 목록을 조회 (페이지네이션 포함)"""
    # 기본 쿼리 생성
    query = db.query(Room).options(
        joinedload(Room.building)
    ).filter(Room.is_available == True)
    
    # 빌딩 필터링
    if building_id:
        query = query.filter(Room.building_id == building_id)
    
    # 전체 항목 수 계산
    total_count = query.count()
    
    # 페이지네이션 적용
    rooms = query.order_by(Room.building_id, Room.room_number).offset((page - 1) * size).limit(size).all()
    
    # 응답 데이터 준비
    result = []
    for room in rooms:
        # 현재 거주자 수 조회
        current_residents = db.query(Resident).filter(
            Resident.room_id == room.id,
            Resident.is_active == True,
            Resident.check_out_date.is_(None)
        ).count()
        
        # 정원 대비 사용 가능 여부 확인
        is_available_for_checkin = room.capacity is None or current_residents < room.capacity
        
        room_data = {
            "id": str(room.id),
            "room_number": room.room_number,
            "floor": room.floor,
            "capacity": room.capacity,
            "current_residents": current_residents,
            "available_spots": room.capacity - current_residents if room.capacity else None,
            "rent": room.rent,
            "note": room.note,
            "is_available_for_checkin": is_available_for_checkin,
            "building": {
                "id": str(room.building.id),
                "name": room.building.name,
                "address": room.building.address
            }
        }
        result.append(room_data)
    
    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size
    
    return {
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/students/{student_id}/residence-history")
def get_student_residence_history(
    student_id: str,
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    is_active: Optional[bool] = Query(None, description="활성 상태로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 해당 학생의 모든 거주 기록 조회
    query = db.query(Resident).options(
        joinedload(Resident.room).joinedload(Room.building)
    ).filter(Resident.resident_id == student_id)

    # 활성 상태 필터링 (선택사항)
    if is_active is not None:
        query = query.filter(Resident.is_active == is_active)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용 (최신 기록부터)
    residents = query.order_by(Resident.created_at.desc()).offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for resident in residents:
        resident_data = {
            "id": str(resident.id),
            "room_id": str(resident.room_id),
            "student_id": str(resident.resident_id),
            "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
            "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
            "is_active": resident.is_active,
            "note": resident.note,
            "created_at": resident.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": resident.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "room": {
                "id": str(resident.room.id),
                "room_number": resident.room.room_number,
                "building_id": str(resident.room.building_id),
                "floor": resident.room.floor,
                "rent": resident.room.rent
            },
            "building": {
                "id": str(resident.room.building.id),
                "name": resident.room.building.name,
                "address": resident.room.building.address
            },
            "student": {
                "id": str(student.id),
                "name": student.name,
                "name_katakana": student.name_katakana,
                "nationality": student.nationality,
                "phone": student.phone,
                "email": student.email,
                "avatar": student.avatar,
                "gender": student.gender,
                "birth_date": student.birth_date.strftime("%Y-%m-%d") if student.birth_date else None,
                "japanese_level": student.japanese_level,
                "local_address": student.local_address
            }
        }
        result.append(resident_data)

    return {
        "student": {
            "id": str(student.id),
            "name": student.name,
            "current_room_id": str(student.current_room_id) if student.current_room_id else None
        },
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/elderly/{elderly_id}/residence-history")
def get_elderly_residence_history(
    elderly_id: str,
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    is_active: Optional[bool] = Query(None, description="활성 상태로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 노인자 존재 여부 확인
    elderly = db.query(Elderly).filter(Elderly.id == elderly_id).first()
    if not elderly:
        raise HTTPException(status_code=404, detail="高齢者が見つかりません")

    # 해당 노인자의 모든 거주 기록 조회
    query = db.query(Resident).options(
        joinedload(Resident.room).joinedload(Room.building)
    ).filter(Resident.resident_id == elderly_id)

    # 활성 상태 필터링 (선택사항)
    if is_active is not None:
        query = query.filter(Resident.is_active == is_active)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용 (최신 기록부터)
    residents = query.order_by(Resident.created_at.desc()).offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for resident in residents:
        resident_data = {
            "id": str(resident.id),
            "room_id": str(resident.room_id),
            "elderly_id": str(resident.resident_id),
            "resident_type": resident.resident_type,
            "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
            "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
            "is_active": resident.is_active,
            "note": resident.note,
            "created_at": resident.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "updated_at": resident.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "room": {
                "id": str(resident.room.id),
                "room_number": resident.room.room_number,
                "building_id": str(resident.room.building_id),
                "floor": resident.room.floor,
                "rent": resident.room.rent
            },
            "building": {
                "id": str(resident.room.building.id),
                "name": resident.room.building.name,
                "address": resident.room.building.address
            },
            "elderly": {
                "id": str(elderly.id),
                "name": elderly.name,
                "name_katakana": elderly.name_katakana,
                "gender": elderly.gender,
                "birth_date": elderly.birth_date.strftime("%Y-%m-%d") if elderly.birth_date else None,
                "phone": elderly.phone,
                "care_level": elderly.care_level,
                "status": elderly.status,
                "avatar": elderly.avatar
            }
        }
        result.append(resident_data)

    return {
        "elderly": {
            "id": str(elderly.id),
            "name": elderly.name,
            "current_room_id": str(elderly.current_room_id) if elderly.current_room_id else None
        },
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/residents/{resident_id}")
def get_resident(
    resident_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 거주자의 상세 정보를 조회합니다."""
    # 거주자 존재 여부 확인
    resident = db.query(Resident).options(
        joinedload(Resident.room).joinedload(Room.building),
        joinedload(Resident.student),
        joinedload(Resident.elderly)
    ).filter(Resident.id == resident_id).first()
    
    if not resident:
        raise HTTPException(status_code=404, detail="거주자를 찾을 수 없습니다")

    # 기본 거주자 정보
    resident_data = {
        "id": str(resident.id),
        "room_id": str(resident.room_id),
        "resident_id": str(resident.resident_id),
        "resident_type": resident.resident_type,
        "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
        "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
        "is_active": resident.is_active,
        "note": resident.note,
        "created_at": resident.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "updated_at": resident.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
        "room": {
            "id": str(resident.room.id),
            "room_number": resident.room.room_number,
            "building_id": str(resident.room.building_id),
            "floor": resident.room.floor,
            "rent": resident.room.rent,
            "maintenance": resident.room.maintenance,
            "service": resident.room.service,
            "capacity": resident.room.capacity,
            "is_available": resident.room.is_available,
            "note": resident.room.note
        },
        "building": {
            "id": str(resident.room.building.id),
            "name": resident.room.building.name,
            "address": resident.room.building.address,
            "total_rooms": resident.room.building.total_rooms,
            "note": resident.room.building.note,
            "resident_type": resident.room.building.resident_type
        }
    }
    
    # 거주자 타입에 따라 다른 정보 추가
    if resident.resident_type == "student" and resident.student:
        resident_data["student"] = {
            "id": str(resident.student.id),
            "name": resident.student.name,
            "name_katakana": resident.student.name_katakana,
            "nationality": resident.student.nationality,
            "phone": resident.student.phone,
            "email": resident.student.email if resident.student.email else "",
            "avatar": resident.student.avatar,
            "gender": resident.student.gender,
            "birth_date": resident.student.birth_date.strftime("%Y-%m-%d") if resident.student.birth_date else None,
            "japanese_level": resident.student.japanese_level,
            "local_address": resident.student.local_address,
            "status": resident.student.status,
            "student_type": resident.student.student_type,
            "consultant": resident.student.consultant,
            "current_room_id": str(resident.student.current_room_id) if resident.student.current_room_id else None
        }
        
        # 회사 정보 추가 (있는 경우)
        if resident.student.company:
            resident_data["student"]["company"] = {
                "id": str(resident.student.company.id),
                "name": resident.student.company.name,
                "address": resident.student.company.address,
            }
        
        # 등급 정보 추가 (있는 경우)
        if resident.student.grade:
            resident_data["student"]["grade"] = {
                "id": str(resident.student.grade.id),
                "name": resident.student.grade.name,
            }
            
    elif resident.resident_type == "elderly" and resident.elderly:
        resident_data["elderly"] = {
            "id": str(resident.elderly.id),
            "name": resident.elderly.name,
            "name_katakana": resident.elderly.name_katakana,
            "gender": resident.elderly.gender,
            "birth_date": resident.elderly.birth_date.strftime("%Y-%m-%d") if resident.elderly.birth_date else None,
            "phone": resident.elderly.phone,
            "care_level": resident.elderly.care_level,
            "status": resident.elderly.status,
            "avatar": resident.elderly.avatar,
            "email": resident.elderly.email,
            "current_room_id": str(resident.elderly.current_room_id) if resident.elderly.current_room_id else None,
            "note": resident.elderly.note
        }
        
        # 카테고리 정보 추가 (있는 경우)
        if resident.elderly.category:
            resident_data["elderly"]["category"] = {
                "id": str(resident.elderly.category.id),
                "category": resident.elderly.category.category,
                "label": resident.elderly.category.label
            }

    return resident_data

@app.get("/students/{student_id}/residence-history/monthly")
def get_student_residence_history_monthly(
    student_id: str,
    year: int = Query(..., description="조회할 년도"),
    month: int = Query(..., description="조회할 월 (1-12)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """학생의 특정 월 거주 이력을 조회"""
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 해당 월의 시작일과 종료일 계산
    from datetime import date
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # 해당 월에 거주 기록이 있는 모든 레코드 조회
    # (입주일이 해당 월에 있거나, 퇴실일이 해당 월에 있거나, 또는 해당 월 전체를 거주한 경우)
    residents = db.query(Resident).options(
        joinedload(Resident.room).joinedload(Room.building)
    ).filter(
        Resident.resident_id == student_id,
        # 입주일이 해당 월에 있거나
        (Resident.check_in_date <= end_date) &
        # 퇴실일이 없거나(현재 거주 중) 퇴실일이 해당 월 이후이거나
        ((Resident.check_out_date.is_(None)) | (Resident.check_out_date >= start_date))
    ).order_by(Resident.check_in_date.desc()).all()

    # 월별 거주 이력 정리
    monthly_history = []
    
    for resident in residents:
        # 해당 월에 실제로 거주했는지 확인
        check_in = resident.check_in_date
        check_out = resident.check_out_date or end_date  # 퇴실일이 없으면 월말까지
        
        # 거주 기간이 해당 월과 겹치는지 확인
        if check_in <= end_date and check_out >= start_date:
            # 실제 거주 시작일과 종료일 계산
            actual_check_in = max(check_in, start_date)
            actual_check_out = min(check_out, end_date)
            
            # 거주 일수 계산
            days_resided = (actual_check_out - actual_check_in).days + 1
            
            resident_data = {
                "id": str(resident.id),
                "room_id": str(resident.room_id),
                "student_id": str(resident.resident_id),
                "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
                "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
                "is_active": resident.is_active,
                "note": resident.note,
                "created_at": resident.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "updated_at": resident.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                "room": {
                    "id": str(resident.room.id),
                    "room_number": resident.room.room_number,
                    "building_id": str(resident.room.building_id),
                    "floor": resident.room.floor,
                    "rent": resident.room.rent
                },
                "building": {
                    "id": str(resident.room.building.id),
                    "name": resident.room.building.name,
                    "address": resident.room.building.address
                },
                "student": {
                    "id": str(student.id),
                    "name": student.name,
                    "name_katakana": student.name_katakana,
                    "nationality": student.nationality,
                    "phone": student.phone,
                    "email": student.email if student.email else "",
                    "avatar": student.avatar,
                    "gender": student.gender,
                    "birth_date": student.birth_date.strftime("%Y-%m-%d") if student.birth_date else None,
                    "japanese_level": student.japanese_level,
                    "local_address": student.local_address
                }
            }
            monthly_history.append(resident_data)

    return {
        "student": {
            "id": str(student.id),
            "name": student.name,
            "current_room_id": str(student.current_room_id) if student.current_room_id else None
        },
        "items": monthly_history,
        "total": len(monthly_history),
        "total_pages": 1,
        "current_page": 1
    }

@app.post("/students/{student_id}/change-residence")
def change_student_residence(
    student_id: str,
    request: ChangeResidenceRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 현재 거주 중인 방 확인
    current_residence = db.query(Resident).filter(
        Resident.resident_id == student_id,
        Resident.is_active == True,
        Resident.check_out_date.is_(None)
    ).first()

    if not current_residence:
        raise HTTPException(status_code=400, detail="해당 학생은 현재 거주 중인 방이 없습니다")

    try:
        # 퇴실만 처리하는 경우 (new_room_id가 None)
        if request.new_room_id is None:
            # 현재 거주지에서 퇴실 처리
            current_residence.check_out_date = datetime.strptime(request.change_date, "%Y-%m-%d").date()
            current_residence.is_active = False
            if request.note:
                current_residence.note = f"退去 - {request.note}"
            else:
                current_residence.note = "退去"

            # 학생의 current_room_id 초기화
            student.current_room_id = None
            
            # 방 로그 기록
            room_log = RoomLog(
                id=str(uuid.uuid4()),
                room_id=current_residence.room_id,
                student_id=student_id,
                action="CHECK_OUT",
                action_date=datetime.strptime(request.change_date, "%Y-%m-%d").date(),
                note=f"퇴실 - {request.note}" if request.note else "퇴실"
            )
            db.add(room_log)
            
            db.commit()
            
            # 로그 생성
            create_database_log(
                db=db,
                table_name="residents",
                record_id=str(current_residence.id),
                action="UPDATE",
                user_id=current_user.id if current_user else None,
                old_values={
                    "room_id": str(current_residence.room_id),
                    "student_id": str(current_residence.resident_id),
                    "check_in_date": str(current_residence.check_in_date),
                    "is_active": True,
                    "check_out_date": None,
                    "note": current_residence.note
                },
                new_values={
                    "room_id": str(current_residence.room_id),
                    "student_id": str(current_residence.resident_id),
                    "check_in_date": str(current_residence.check_in_date),
                    "is_active": False,
                    "check_out_date": str(current_residence.check_out_date),
                    "note": current_residence.note
                },
                changed_fields=["is_active", "check_out_date"],
                note="居住地変更 - 退去"
            )
            
            return {
                "message": "학생이 성공적으로 퇴실했습니다",
                "student_id": str(student_id),
                "old_room_id": str(current_residence.room_id),
                "new_room_id": None,
                "change_date": request.change_date,
                "action": "CHECK_OUT"
            }

        # 이사 처리하는 경우 (new_room_id가 제공됨)
        else:
            # 새로운 방 존재 여부 확인
            new_room = db.query(Room).filter(Room.id == request.new_room_id).first()
            if not new_room:
                raise HTTPException(status_code=404, detail="새로운 방을 찾을 수 없습니다")

            # 새로운 방이 사용 가능한지 확인
            if not new_room.is_available:
                raise HTTPException(status_code=400, detail="해당 방은 현재 사용할 수 없습니다")

            # 새로운 방의 정원 확인
            current_residents_in_new_room = db.query(Resident).filter(
                Resident.room_id == request.new_room_id,
                Resident.is_active == True,
                Resident.check_out_date.is_(None)
            ).count()
            
            if new_room.capacity and current_residents_in_new_room >= new_room.capacity:
                raise HTTPException(status_code=400, detail="해당 방은 정원이 초과되어 입주할 수 없습니다")

            # 1. 현재 거주지에서 퇴실 처리 (이사할 때는 전날에 퇴실)
            current_residence.check_out_date = datetime.strptime(request.change_date, "%Y-%m-%d").date() - timedelta(days=1)
            current_residence.is_active = False
            if request.note:
                current_residence.note = f"이사로 인한 퇴실 - {request.note}"

            # 2. 새로운 방에 입주 기록 생성
            change_date_obj = datetime.strptime(request.change_date, "%Y-%m-%d").date()
            check_in_date = change_date_obj  # 변경날짜로 설정
            
            new_residence = Resident(
                id=str(uuid.uuid4()),
                room_id=request.new_room_id,
                student_id=student_id,
                check_in_date=check_in_date,
                note=f"이사로 인한 입주 - {request.note}" if request.note else "이사로 인한 입주"
            )
            db.add(new_residence)

            # 3. 학생의 current_room_id 업데이트
            student.current_room_id = request.new_room_id
            
            # 4. 방 로그 기록
            room_log = RoomLog(
                id=str(uuid.uuid4()),
                room_id=request.new_room_id,
                student_id=student_id,
                action="MOVE",
                action_date=datetime.strptime(request.change_date, "%Y-%m-%d").date(),
                note=f"거주지 변경 - {request.note}" if request.note else "거주지 변경"
            )
            db.add(room_log)
            
            db.commit()
            
            # 로그 생성 - 기존 거주지 퇴실
            create_database_log(
                db=db,
                table_name="residents",
                record_id=str(current_residence.id),
                action="UPDATE",
                user_id=current_user.id if current_user else None,
                old_values={
                    "room_id": str(current_residence.room_id),
                    "student_id": str(current_residence.resident_id),
                    "check_in_date": str(current_residence.check_in_date),
                    "is_active": True,
                    "check_out_date": None,
                    "note": current_residence.note
                },
                new_values={
                    "room_id": str(current_residence.room_id),
                    "student_id": str(current_residence.resident_id),
                    "check_in_date": str(current_residence.check_in_date),
                    "is_active": False,
                    "check_out_date": str(current_residence.check_out_date),
                    "note": current_residence.note
                },
                changed_fields=["is_active", "check_out_date"],
                note="居住地変更 - 既存部屋退去"
            )
            
            # 로그 생성 - 새로운 거주지 입주
            create_database_log(
                db=db,
                table_name="residents",
                record_id=str(new_residence.id),
                action="CREATE",
                user_id=current_user.id if current_user else None,
                new_values={
                    "room_id": str(new_residence.room_id),
                    "student_id": str(new_residence.resident_id),
                    "check_in_date": str(new_residence.check_in_date),
                    "is_active": new_residence.is_active,
                    "note": new_residence.note
                },
                note="居住地変更 - 新部屋入居"
            )
            
            return {
                "message": "学生の居住地が正常に変更されました",
                "student_id": str(student_id),
                "old_room_id": str(current_residence.room_id),
                "new_room_id": str(request.new_room_id),
                "change_date": request.change_date,
                "action": "MOVE"
            }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"거주지 변경 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/rooms")
def get_all_rooms(
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """모든 방 목록을 조회"""
    # 기본 쿼리 생성
    query = db.query(Room).options(
        joinedload(Room.building)
    )
    
    # 전체 항목 수 계산
    total_count = query.count()
    
    # 페이지네이션 적용
    rooms = query.order_by(Room.building_id, Room.room_number).offset((page - 1) * size).limit(size).all()
    
    # 응답 데이터 준비
    result = []
    for room in rooms:
        # 현재 거주자 수 조회
        current_residents = db.query(Resident).filter(
            Resident.room_id == room.id,
            Resident.is_active == True,
            Resident.check_out_date.is_(None)
        ).count()
        
        room_data = {
            "id": str(room.id),
            "room_number": room.room_number,
            "floor": room.floor,
            "capacity": room.capacity,
            "current_residents": current_residents,
            "available_spots": room.capacity - current_residents if room.capacity else None,
            "rent": room.rent,
            "note": room.note,
            "is_available": room.is_available,
            "building": {
                "id": str(room.building.id),
                "name": room.building.name,
                "address": room.building.address
            }
        }
        result.append(room_data)
    
    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size
    
    return {
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.post("/students/{student_id}/new-residence")
def create_new_residence(
    student_id: str,
    request: NewResidenceRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """학생에게 새로운 거주 기록을 추가"""
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == request.new_room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 방이 사용 가능한지 확인
    if not room.is_available:
        raise HTTPException(status_code=400, detail="해당 방은 현재 사용할 수 없습니다")

    # 입주일과 퇴실일 유효성 검사
    try:
        check_in_date = datetime.strptime(request.change_date, "%Y-%m-%d").date()
        check_out_date = None
        if request.check_out_date:
            check_out_date = datetime.strptime(request.check_out_date, "%Y-%m-%d").date()
            if check_out_date <= check_in_date:
                raise HTTPException(status_code=400, detail="퇴실일은 입주일보다 늦어야 합니다")
    except ValueError:
        raise HTTPException(status_code=400, detail="날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)")

    # 현재 거주 중인지 확인 (퇴실일이 없는 경우)
    is_currently_residing = check_out_date is None
    
    # 현재 거주 중인 경우, 해당 방의 정원 확인
    if is_currently_residing:
        current_residents = db.query(Resident).filter(
            Resident.room_id == request.new_room_id,
            Resident.is_active == True,
            Resident.check_out_date.is_(None)
        ).count()
        
        # 방에 정원이 설정되어 있고, 현재 거주자 수가 정원에 도달한 경우
        if room.capacity and current_residents >= room.capacity:
            raise HTTPException(status_code=400, detail=f"해당 방은 정원({room.capacity}명)이 초과되어 입주할 수 없습니다. 현재 거주자: {current_residents}명")

    try:
        # 거주 기록 생성
        new_residence = Resident(
            id=str(uuid.uuid4()),
            room_id=request.new_room_id,
            resident_id=student_id,
            resident_type="student",  # 명시적으로 student 타입 설정
            check_in_date=check_in_date,
            check_out_date=check_out_date,
            is_active=is_currently_residing,
            note=request.note
        )
        db.add(new_residence)

        # 현재 거주 중인 경우 학생의 current_room_id와 address 업데이트
        if is_currently_residing:
            student.current_room_id = request.new_room_id
            
            # 학생의 주소 업데이트 (빌딩 타입에 따라 다르게 설정)
            building_address = room.building.address if room.building else ""
            building_type = room.building.building_type if room.building else None
            
            if building_type == "mansion":
                # mansion 타입: 빌딩주소 + 방번호
                room_number = room.room_number
                student.address = f"{building_address} {room_number}".strip()
            elif building_type == "house":
                # house 타입: 빌딩주소만
                student.address = building_address.strip()
            else:
                # 기본값: 빌딩주소 + 방번호
                room_number = room.room_number
                student.address = f"{building_address} {room_number}".strip()
        
        # 방 로그 기록
        action = "CHECK_IN" if is_currently_residing else "HISTORICAL_ENTRY"
        room_log = RoomLog(
            id=str(uuid.uuid4()),
            room_id=request.new_room_id,
            student_id=student_id,
            action=action,
            action_date=check_in_date,
            note=f"신규 거주 기록 추가 - {request.note}" if request.note else "신규 거주 기록 추가"
        )
        db.add(room_log)
        
        db.commit()
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="residents",
            record_id=str(new_residence.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "room_id": str(new_residence.room_id),
                "student_id": str(new_residence.resident_id),
                "check_in_date": str(new_residence.check_in_date),
                "check_out_date": str(new_residence.check_out_date) if new_residence.check_out_date else None,
                "is_active": new_residence.is_active,
                "note": new_residence.note
            },
            note="新居住記録追加"
        )
        
        return {
            "message": "新居住記録が正常に追加されました",
            "student_id": str(student_id),
            "room_id": str(request.new_room_id),
            "check_in_date": request.change_date,
            "check_out_date": request.check_out_date,
            "is_currently_residing": is_currently_residing,
            "residence_id": str(new_residence.id)
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"거주 기록 추가 중 오류가 발생했습니다: {str(e)}"
        )

# 광열비 관련 API들
@app.post("/room-charges", response_model=RoomChargeResponse)
def create_room_charge(
    charge: RoomChargeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """광열비 등록 (새로운 구조)"""
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == charge.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == charge.room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 날짜 형식 검증
    try:
        charge_month = datetime.strptime(charge.charge_month, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)")

    # 해당 월에 이미 광열비가 등록되어 있는지 확인
    existing_charge = db.query(RoomCharge).filter(
        RoomCharge.student_id == charge.student_id,
        RoomCharge.room_id == charge.room_id,
        RoomCharge.charge_month == charge_month
    ).first()
    
    if existing_charge:
        raise HTTPException(status_code=400, detail="해당 월에 이미 광열비가 등록되어 있습니다")

    try:
        # 광열비 등록
        new_charge = RoomCharge(
            id=str(uuid.uuid4()),
            student_id=charge.student_id,
            room_id=charge.room_id,
            charge_month=charge_month,
            total_amount=charge.total_amount,
            note=charge.note
        )
        db.add(new_charge)
        db.flush()  # ID 생성을 위해 flush

        # 청구 항목들 생성
        if charge.charge_items:
            for item_data in charge.charge_items:
                # 날짜 변환
                period_start = None
                period_end = None
                if item_data.period_start:
                    period_start = datetime.strptime(item_data.period_start, "%Y-%m-%d").date()
                if item_data.period_end:
                    period_end = datetime.strptime(item_data.period_end, "%Y-%m-%d").date()

                charge_item = ChargeItem(
                    id=str(uuid.uuid4()),
                    room_charge_id=new_charge.id,
                    charge_type=item_data.charge_type,
                    period_start=period_start,
                    period_end=period_end,
                    amount=item_data.amount,
                    unit_price=item_data.unit_price,
                    quantity=item_data.quantity,
                    memo=item_data.memo
                )
                db.add(charge_item)

        db.commit()
        db.refresh(new_charge)

        # 응답 데이터 준비
        response_data = {
            "id": str(new_charge.id),
            "student_id": str(new_charge.student_id),
            "room_id": str(new_charge.room_id),
            "charge_month": new_charge.charge_month,
            "total_amount": float(new_charge.total_amount) if new_charge.total_amount else None,
            "created_at": new_charge.created_at,
            "note": new_charge.note,
            "student": {
                "id": str(student.id),
                "name": student.name,
                "name_katakana": student.name_katakana
            },
            "room": {
                "id": str(room.id),
                "room_number": room.room_number,
                "building_name": room.building.name if room.building else None
            },
            "charge_items": []
        }

        # 청구 항목 정보 추가
        if new_charge.charge_items:
            for item in new_charge.charge_items:
                item_data = {
                    "id": str(item.id),
                    "room_charge_id": str(item.room_charge_id),
                    "charge_type": item.charge_type,
                    "period_start": item.period_start.strftime("%Y-%m-%d") if item.period_start else None,
                    "period_end": item.period_end.strftime("%Y-%m-%d") if item.period_end else None,
                    "amount": float(item.amount) if item.amount else None,
                    "unit_price": float(item.unit_price) if item.unit_price else None,
                    "quantity": float(item.quantity) if item.quantity else None,
                    "memo": item.memo,
                    "allocations": []
                }
                response_data["charge_items"].append(item_data)

        return response_data

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"광열비 등록 중 오류가 발생했습니다: {str(e)}"
        )

@app.post("/charge-items/{charge_item_id}/allocations")
def create_charge_item_allocation(
    charge_item_id: str,
    allocation: ChargeItemAllocationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """청구 항목 배분 등록"""
    # 청구 항목 존재 여부 확인
    charge_item = db.query(ChargeItem).filter(ChargeItem.id == charge_item_id).first()
    if not charge_item:
        raise HTTPException(status_code=404, detail="청구 항목을 찾을 수 없습니다")

    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == allocation.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 이미 배분이 등록되어 있는지 확인
    existing_allocation = db.query(ChargeItemAllocation).filter(
        ChargeItemAllocation.charge_item_id == charge_item_id,
        ChargeItemAllocation.student_id == allocation.student_id
    ).first()
    
    if existing_allocation:
        raise HTTPException(status_code=400, detail="해당 학생에게 이미 배분이 등록되어 있습니다")

    try:
        # 배분 등록
        new_allocation = ChargeItemAllocation(
            id=str(uuid.uuid4()),
            charge_item_id=charge_item_id,
            student_id=allocation.student_id,
            amount=allocation.amount,
            days_used=allocation.days_used,
            memo=allocation.memo
        )
        db.add(new_allocation)
        db.commit()
        db.refresh(new_allocation)

        return {
            "message": "請求項目配分が正常に登録されました",
            "id": str(new_allocation.id),
            "charge_item_id": str(charge_item_id),
            "student_id": str(allocation.student_id),
            "amount": float(new_allocation.amount),
            "days_used": new_allocation.days_used,
            "memo": new_allocation.memo
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"청구 항목 배분 등록 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/charge-items/{charge_item_id}/allocations")
def get_charge_item_allocations(
    charge_item_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """청구 항목의 배분 목록 조회"""
    # 청구 항목 존재 여부 확인
    charge_item = db.query(ChargeItem).filter(ChargeItem.id == charge_item_id).first()
    if not charge_item:
        raise HTTPException(status_code=404, detail="청구 항목을 찾을 수 없습니다")

    # 배분 목록 조회
    allocations = db.query(ChargeItemAllocation).options(
        joinedload(ChargeItemAllocation.student)
    ).filter(ChargeItemAllocation.charge_item_id == charge_item_id).all()

    result = []
    for allocation in allocations:
        allocation_data = {
            "id": str(allocation.id),
            "charge_item_id": str(allocation.charge_item_id),
            "student_id": str(allocation.student_id),
            "amount": float(allocation.amount),
            "days_used": allocation.days_used,
            "memo": allocation.memo,
            "student": {
                "id": str(allocation.student.id),
                "name": allocation.student.name,
                "name_katakana": allocation.student.name_katakana
            }
        }
        result.append(allocation_data)

    return {
        "charge_item": {
            "id": str(charge_item.id),
            "charge_type": charge_item.charge_type,
            "amount": float(charge_item.amount) if charge_item.amount else None
        },
        "allocations": result,
        "total": len(result)
    }

@app.put("/charge-items/{charge_item_id}/allocations/{allocation_id}")
def update_charge_item_allocation(
    charge_item_id: str,
    allocation_id: str,
    allocation_update: ChargeItemAllocationUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """청구 항목 배분 수정"""
    # 배분 존재 여부 확인
    allocation = db.query(ChargeItemAllocation).filter(
        ChargeItemAllocation.id == allocation_id,
        ChargeItemAllocation.charge_item_id == charge_item_id
    ).first()
    
    if not allocation:
        raise HTTPException(status_code=404, detail="배분을 찾을 수 없습니다")

    try:
        # 업데이트할 데이터 준비
        update_data = allocation_update.dict(exclude_unset=True)
        
        # 금액이 제공된 경우 float로 변환
        if "amount" in update_data and update_data["amount"] is not None:
            update_data["amount"] = float(update_data["amount"])

        # 데이터 업데이트
        for field, value in update_data.items():
            setattr(allocation, field, value)

        db.commit()
        db.refresh(allocation)

        return {
            "message": "請求項目配分が正常に修正されました",
            "id": str(allocation.id),
            "amount": float(allocation.amount),
            "days_used": allocation.days_used,
            "memo": allocation.memo
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"청구 항목 배분 수정 중 오류가 발생했습니다: {str(e)}"
        )

@app.delete("/charge-items/{charge_item_id}/allocations/{allocation_id}")
def delete_charge_item_allocation(
    charge_item_id: str,
    allocation_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """청구 항목 배분 삭제"""
    # 배분 존재 여부 확인
    allocation = db.query(ChargeItemAllocation).filter(
        ChargeItemAllocation.id == allocation_id,
        ChargeItemAllocation.charge_item_id == charge_item_id
    ).first()
    
    if not allocation:
        raise HTTPException(status_code=404, detail="배분을 찾을 수 없습니다")

    try:
        db.delete(allocation)
        db.commit()
        return {"message": "請求項目配分が正常に削除されました"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"청구 항목 배분 삭제 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/students/{student_id}/room-charges")
def get_student_room_charges(
    student_id: str,
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 학생의 광열비 목록 조회"""
    # 학생 존재 여부 확인
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 해당 학생의 광열비 조회
    query = db.query(RoomCharge).options(
        joinedload(RoomCharge.room).joinedload(Room.building)
    ).filter(RoomCharge.student_id == student_id)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용 (최신순)
    charges = query.order_by(RoomCharge.charge_month.desc()).offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for charge in charges:
        charge_data = {
            "id": str(charge.id),
            "room_id": str(charge.room_id),
            "charge_month": charge.charge_month.strftime("%Y-%m-%d"),
            "total_amount": float(charge.total_amount) if charge.total_amount else None,
            "created_at": charge.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "note": charge.note,
            "room": {
                "id": str(charge.room.id),
                "room_number": charge.room.room_number,
                "building_name": charge.room.building.name if charge.room.building else None
            }
        }
        result.append(charge_data)

    return {
        "student": {
            "id": str(student.id),
            "name": student.name,
            "name_katakana": student.name_katakana
        },
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/room-charges")
def get_room_charges(
    student_id: Optional[str] = Query(None, description="학생 ID로 필터링"),
    room_id: Optional[str] = Query(None, description="방 ID로 필터링"),
    charge_month: Optional[str] = Query(None, description="청구 월로 필터링 (YYYY-MM)"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """광열비 목록 조회 (새로운 구조)"""
    # 기본 쿼리 생성
    query = db.query(RoomCharge).options(
        joinedload(RoomCharge.student),
        joinedload(RoomCharge.room).joinedload(Room.building),
        joinedload(RoomCharge.charge_items)
    )

    # 필터 조건 추가
    if student_id:
        query = query.filter(RoomCharge.student_id == student_id)
    if room_id:
        query = query.filter(RoomCharge.room_id == room_id)
    if charge_month:
        try:
            # YYYY-MM 형식으로 받아서 해당 월의 첫날로 변환
            year, month = charge_month.split("-")
            start_date = datetime(int(year), int(month), 1).date()
            if int(month) == 12:
                end_date = datetime(int(year) + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(int(year), int(month) + 1, 1).date() - timedelta(days=1)
            
            query = query.filter(
                RoomCharge.charge_month >= start_date,
                RoomCharge.charge_month <= end_date
            )
        except ValueError:
            raise HTTPException(status_code=400, detail="청구 월 형식이 올바르지 않습니다 (YYYY-MM)")
    
    # 전체 항목 수 계산
    total_count = query.count()
    
    # 페이지네이션 적용 (최신순)
    charges = query.order_by(RoomCharge.charge_month.desc()).offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size
    
    # 응답 데이터 준비
    result = []
    for charge in charges:
        charge_data = {
            "id": str(charge.id),
            "student_id": str(charge.student_id),
            "room_id": str(charge.room_id),
            "charge_month": charge.charge_month.strftime("%Y-%m-%d"),
            "total_amount": float(charge.total_amount) if charge.total_amount else None,
            "created_at": charge.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "note": charge.note,
            "student": {
                "id": str(charge.student.id),
                "name": charge.student.name,
                "name_katakana": charge.student.name_katakana
            },
            "room": {
                "id": str(charge.room.id),
                "room_number": charge.room.room_number,
                "building_name": charge.room.building.name if charge.room.building else None
            },
            "charge_items": []
        }

        # 청구 항목 정보 추가
        for item in charge.charge_items:
            item_data = {
                "id": str(item.id),
                "charge_type": item.charge_type,
                "period_start": item.period_start.strftime("%Y-%m-%d") if item.period_start else None,
                "period_end": item.period_end.strftime("%Y-%m-%d") if item.period_end else None,
                "amount": float(item.amount) if item.amount else None,
                "unit_price": float(item.unit_price) if item.unit_price else None,
                "quantity": float(item.quantity) if item.quantity else None,
                "memo": item.memo
            }
            charge_data["charge_items"].append(item_data)

        result.append(charge_data)
    
    return {
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/room-charges/{charge_id}")
def get_room_charge(
    charge_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 광열비 조회 (새로운 구조)"""
    charge = db.query(RoomCharge).options(
        joinedload(RoomCharge.student),
        joinedload(RoomCharge.room).joinedload(Room.building),
        joinedload(RoomCharge.charge_items).joinedload(ChargeItem.allocations).joinedload(ChargeItemAllocation.student)
    ).filter(RoomCharge.id == charge_id).first()
    
    if not charge:
        raise HTTPException(status_code=404, detail="광열비를 찾을 수 없습니다")

    # 청구 항목 정보 준비
    charge_items = []
    for item in charge.charge_items:
        # 배분 정보 준비
        allocations = []
        for allocation in item.allocations:
            allocation_data = {
                "id": str(allocation.id),
                "student_id": str(allocation.student_id),
                "amount": float(allocation.amount),
                "days_used": allocation.days_used,
                "memo": allocation.memo,
                "student": {
                    "id": str(allocation.student.id),
                    "name": allocation.student.name,
                    "name_katakana": allocation.student.name_katakana
                }
            }
            allocations.append(allocation_data)

        item_data = {
            "id": str(item.id),
            "charge_type": item.charge_type,
            "period_start": item.period_start.strftime("%Y-%m-%d") if item.period_start else None,
            "period_end": item.period_end.strftime("%Y-%m-%d") if item.period_end else None,
            "amount": float(item.amount) if item.amount else None,
            "unit_price": float(item.unit_price) if item.unit_price else None,
            "quantity": float(item.quantity) if item.quantity else None,
            "memo": item.memo,
            "allocations": allocations
        }
        charge_items.append(item_data)

    return {
        "id": str(charge.id),
        "student_id": str(charge.student_id),
        "room_id": str(charge.room_id),
        "charge_month": charge.charge_month.strftime("%Y-%m-%d"),
        "total_amount": float(charge.total_amount) if charge.total_amount else None,
        "created_at": charge.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "note": charge.note,
        "student": {
            "id": str(charge.student.id),
            "name": charge.student.name,
            "name_katakana": charge.student.name_katakana
        },
        "room": {
            "id": str(charge.room.id),
            "room_number": charge.room.room_number,
            "building_name": charge.room.building.name if charge.room.building else None
        },
        "charge_items": charge_items
    }

@app.put("/room-charges/{charge_id}")
def update_room_charge(
    charge_id: str,
    charge_update: RoomChargeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """광열비 수정"""
    # 광열비 존재 여부 확인
    charge = db.query(RoomCharge).filter(RoomCharge.id == charge_id).first()
    if not charge:
        raise HTTPException(status_code=404, detail="광열비를 찾을 수 없습니다")

    try:
        # 업데이트할 데이터 준비
        update_data = charge_update.dict(exclude_unset=True)
        
        # 금액이 제공된 경우 float로 변환
        if "total_amount" in update_data and update_data["total_amount"] is not None:
            update_data["total_amount"] = float(update_data["total_amount"])

        # 데이터 업데이트
        for field, value in update_data.items():
            setattr(charge, field, value)

        db.commit()
        db.refresh(charge)

        return {
            "message": "光熱費が正常に修正されました",
            "id": str(charge.id),
            "total_amount": float(charge.total_amount) if charge.total_amount else None,
            "note": charge.note
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"광열비 수정 중 오류가 발생했습니다: {str(e)}"
        )

@app.delete("/room-charges/{charge_id}")
def delete_room_charge(
    charge_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """광열비 삭제"""
    # 광열비 존재 여부 확인
    charge = db.query(RoomCharge).filter(RoomCharge.id == charge_id).first()
    if not charge:
        raise HTTPException(status_code=404, detail="광열비를 찾을 수 없습니다")

    try:
        db.delete(charge)
        db.commit()
        return {"message": "光熱費が正常に削除されました"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"광열비 삭제 중 오류가 발생했습니다: {str(e)}"
        )

@app.post("/room-charges/{charge_id}/charge-items")
def create_charge_item(
    charge_id: str,
    charge_item: ChargeItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """청구 항목 추가"""
    # 광열비 존재 여부 확인
    room_charge = db.query(RoomCharge).filter(RoomCharge.id == charge_id).first()
    if not room_charge:
        raise HTTPException(status_code=404, detail="광열비를 찾을 수 없습니다")

    # 날짜 변환
    period_start = None
    period_end = None
    if charge_item.period_start:
        try:
            period_start = datetime.strptime(charge_item.period_start, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="기간 시작일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
    
    if charge_item.period_end:
        try:
            period_end = datetime.strptime(charge_item.period_end, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="기간 종료일 형식이 올바르지 않습니다 (YYYY-MM-DD)")

    try:
        # 청구 항목 생성
        new_charge_item = ChargeItem(
            id=str(uuid.uuid4()),
            room_charge_id=charge_id,
            charge_type=charge_item.charge_type,
            period_start=period_start,
            period_end=period_end,
            amount=charge_item.amount,
            unit_price=charge_item.unit_price,
            quantity=charge_item.quantity,
            memo=charge_item.memo
        )
        db.add(new_charge_item)
        db.commit()
        db.refresh(new_charge_item)

        return {
            "message": "請求項目が正常に追加されました",
            "id": str(new_charge_item.id),
            "charge_type": new_charge_item.charge_type,
            "amount": float(new_charge_item.amount) if new_charge_item.amount else None
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"청구 항목 추가 중 오류가 발생했습니다: {str(e)}"
        )

@app.put("/charge-items/{charge_item_id}")
def update_charge_item(
    charge_item_id: str,
    charge_item_update: ChargeItemUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """청구 항목 수정"""
    # 청구 항목 존재 여부 확인
    charge_item = db.query(ChargeItem).filter(ChargeItem.id == charge_item_id).first()
    if not charge_item:
        raise HTTPException(status_code=404, detail="청구 항목을 찾을 수 없습니다")

    # 기존 값 저장 (로그용)
    old_values = {
        "charge_id": str(charge_item.charge_id),
        "charge_type": charge_item.charge_type,
        "amount": float(charge_item.amount) if charge_item.amount else None,
        "period_start": str(charge_item.period_start) if charge_item.period_start else None,
        "period_end": str(charge_item.period_end) if charge_item.period_end else None,
        "memo": charge_item.memo
    }
    
    try:
        # 업데이트할 데이터 준비
        update_data = charge_item_update.dict(exclude_unset=True)
        
        # 날짜 변환
        if "period_start" in update_data and update_data["period_start"]:
            try:
                update_data["period_start"] = datetime.strptime(update_data["period_start"], "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="기간 시작일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        
        if "period_end" in update_data and update_data["period_end"]:
            try:
                update_data["period_end"] = datetime.strptime(update_data["period_end"], "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="기간 종료일 형식이 올바르지 않습니다 (YYYY-MM-DD)")

        # 데이터 업데이트
        for field, value in update_data.items():
            setattr(charge_item, field, value)

        db.commit()
        db.refresh(charge_item)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="charge_items",
            record_id=str(charge_item.id),
            action="UPDATE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            new_values={
                "charge_id": str(charge_item.charge_id),
                "charge_type": charge_item.charge_type,
                "amount": float(charge_item.amount) if charge_item.amount else None,
                "period_start": str(charge_item.period_start) if charge_item.period_start else None,
                "period_end": str(charge_item.period_end) if charge_item.period_end else None,
                "memo": charge_item.memo
            },
            changed_fields=list(update_data.keys()),
            note="請求項目修正"
        )

        return {
            "message": "請求項目が正常に修正されました",
            "id": str(charge_item.id),
            "charge_type": charge_item.charge_type,
            "amount": float(charge_item.amount) if charge_item.amount else None
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"청구 항목 수정 중 오류가 발생했습니다: {str(e)}"
        )

@app.delete("/charge-items/{charge_item_id}")
def delete_charge_item(
    charge_item_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """청구 항목 삭제"""
    # 청구 항목 존재 여부 확인
    charge_item = db.query(ChargeItem).filter(ChargeItem.id == charge_item_id).first()
    if not charge_item:
        raise HTTPException(status_code=404, detail="청구 항목을 찾을 수 없습니다")

    # 삭제 전 값 저장 (로그용)
    old_values = {
        "charge_id": str(charge_item.charge_id),
        "charge_type": charge_item.charge_type,
        "amount": float(charge_item.amount) if charge_item.amount else None,
        "period_start": str(charge_item.period_start) if charge_item.period_start else None,
        "period_end": str(charge_item.period_end) if charge_item.period_end else None,
        "memo": charge_item.memo
    }
    
    try:
        db.delete(charge_item)
        db.commit()
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="charge_items",
            record_id=str(charge_item.id),
            action="DELETE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            note="請求項目削除"
        )
        
        return {"message": "請求項目が正常に削除されました"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"청구 항목 삭제 중 오류가 발생했습니다: {str(e)}"
        )

# 방 공과금 관련 API들
@app.post("/room-utilities", response_model=RoomUtilityResponse)
def create_room_utility(
    utility: RoomUtilityCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """방 공과금 등록"""
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == utility.room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 날짜 형식 검증
    try:
        period_start = datetime.strptime(utility.period_start, "%Y-%m-%d").date()
        period_end = datetime.strptime(utility.period_end, "%Y-%m-%d").date()
        charge_month = datetime.strptime(utility.charge_month, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)")

    # 기간 유효성 검사
    if period_end <= period_start:
        raise HTTPException(status_code=400, detail="검침 종료일은 시작일보다 늦어야 합니다")

    # 해당 방의 같은 유형, 같은 청구 월에 이미 등록되어 있는지 확인하고 덮어쓰기
    existing_utility = db.query(RoomUtility).filter(
        RoomUtility.room_id == utility.room_id,
        RoomUtility.utility_type == utility.utility_type,
        RoomUtility.charge_month == charge_month
    ).first()
    
    if existing_utility:
        # 기존 데이터 업데이트
        existing_utility.period_start = period_start
        existing_utility.period_end = period_end
        existing_utility.usage = utility.usage
        existing_utility.unit_price = utility.unit_price
        existing_utility.total_amount = utility.total_amount
        existing_utility.memo = utility.memo
        
        db.commit()
        db.refresh(existing_utility)

        # 응답 데이터 준비
        response_data = {
          "id": str(existing_utility.id),
          "room_id": str(existing_utility.room_id),
          "utility_type": existing_utility.utility_type,
          "period_start": existing_utility.period_start,
          "period_end": existing_utility.period_end,
          "usage": float(existing_utility.usage) if existing_utility.usage else None,
          "unit_price": float(existing_utility.unit_price) if existing_utility.unit_price else None,
          "total_amount": float(existing_utility.total_amount) if existing_utility.total_amount else None,
          "charge_month": existing_utility.charge_month,
          "memo": existing_utility.memo,
          "created_at": existing_utility.created_at,
          "room": {
              "id": str(room.id),
              "room_number": room.room_number,
              "building_name": room.building.name if room.building else None
          }
        }
        
        return response_data

    try:
        # 공과금 등록
        new_utility = RoomUtility(
            id=str(uuid.uuid4()),
            room_id=utility.room_id,
            utility_type=utility.utility_type,
            period_start=period_start,
            period_end=period_end,
            usage=utility.usage,
            unit_price=utility.unit_price,
            total_amount=utility.total_amount,
            charge_month=charge_month,
            memo=utility.memo
        )
        db.add(new_utility)
        db.commit()
        db.refresh(new_utility)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="room_utilities",
            record_id=str(new_utility.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "room_id": str(new_utility.room_id),
                "utility_type": new_utility.utility_type,
                "period_start": str(new_utility.period_start),
                "period_end": str(new_utility.period_end),
                "usage": float(new_utility.usage) if new_utility.usage else None,
                "unit_price": float(new_utility.unit_price) if new_utility.unit_price else None,
                "total_amount": float(new_utility.total_amount) if new_utility.total_amount else None,
                "charge_month": str(new_utility.charge_month),
                "memo": new_utility.memo
            },
            note="部屋光熱費登録"
        )

        # 응답 데이터 준비
        response_data = {
            "id": str(new_utility.id),
            "room_id": str(new_utility.room_id),
            "utility_type": new_utility.utility_type,
            "period_start": new_utility.period_start,
            "period_end": new_utility.period_end,
            "usage": float(new_utility.usage) if new_utility.usage else None,
            "unit_price": float(new_utility.unit_price) if new_utility.unit_price else None,
            "total_amount": float(new_utility.total_amount) if new_utility.total_amount else None,
            "charge_month": new_utility.charge_month,
            "memo": new_utility.memo,
            "created_at": new_utility.created_at,
            "room": {
                "id": str(room.id),
                "room_number": room.room_number,
                "building_name": room.building.name if room.building else None
            }
        }

        return response_data

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"공과금 등록 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/room-utilities")
def get_room_utilities(
    room_id: Optional[str] = Query(None, description="방 ID로 필터링"),
    utility_type: Optional[str] = Query(None, description="공과금 유형으로 필터링"),
    charge_month: Optional[str] = Query(None, description="청구 월로 필터링 (YYYY-MM)"),
    page: int = Query(1, description="페이지 번호", ge=1),
    size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """방 공과금 목록 조회"""
    # 기본 쿼리 생성
    query = db.query(RoomUtility).options(
        joinedload(RoomUtility.room).joinedload(Room.building)
    )

    # 필터 조건 추가
    if room_id:
        query = query.filter(RoomUtility.room_id == room_id)
    if utility_type:
        query = query.filter(RoomUtility.utility_type == utility_type)
    if charge_month:
        try:
            # YYYY-MM 형식으로 받아서 해당 월의 첫날로 변환
            year, month = charge_month.split("-")
            start_date = datetime(int(year), int(month), 1).date()
            if int(month) == 12:
                end_date = datetime(int(year) + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(int(year), int(month) + 1, 1).date() - timedelta(days=1)
            
            query = query.filter(
                RoomUtility.charge_month >= start_date,
                RoomUtility.charge_month <= end_date
            )
        except ValueError:
            raise HTTPException(status_code=400, detail="청구 월 형식이 올바르지 않습니다 (YYYY-MM)")

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용 (최신순)
    utilities = query.order_by(RoomUtility.charge_month.desc(), RoomUtility.created_at.desc()).offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for utility in utilities:
        utility_data = {
            "id": str(utility.id),
            "room_id": str(utility.room_id),
            "utility_type": utility.utility_type,
            "period_start": utility.period_start.strftime("%Y-%m-%d"),
            "period_end": utility.period_end.strftime("%Y-%m-%d"),
            "usage": float(utility.usage) if utility.usage else None,
            "unit_price": float(utility.unit_price) if utility.unit_price else None,
            "total_amount": float(utility.total_amount) if utility.total_amount else None,
            "charge_month": utility.charge_month.strftime("%Y-%m-%d"),
            "memo": utility.memo,
            "created_at": utility.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "room": {
                "id": str(utility.room.id),
                "room_number": utility.room.room_number,
                "building_name": utility.room.building.name if utility.room.building else None
            }
        }
        result.append(utility_data)

    return {
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/room-utilities/{utility_id}")
def get_room_utility(
    utility_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 방 공과금 조회"""
    utility = db.query(RoomUtility).options(
        joinedload(RoomUtility.room).joinedload(Room.building)
    ).filter(RoomUtility.id == utility_id).first()
    
    if not utility:
        raise HTTPException(status_code=404, detail="공과금을 찾을 수 없습니다")

    return {
        "id": str(utility.id),
        "room_id": str(utility.room_id),
        "utility_type": utility.utility_type,
        "period_start": utility.period_start.strftime("%Y-%m-%d"),
        "period_end": utility.period_end.strftime("%Y-%m-%d"),
        "usage": float(utility.usage) if utility.usage else None,
        "unit_price": float(utility.unit_price) if utility.unit_price else None,
        "total_amount": float(utility.total_amount) if utility.total_amount else None,
        "charge_month": utility.charge_month.strftime("%Y-%m-%d"),
        "memo": utility.memo,
        "created_at": utility.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "room": {
            "id": str(utility.room.id),
            "room_number": utility.room.room_number,
            "building_name": utility.room.building.name if utility.room.building else None
        }
    }

@app.put("/room-utilities/{utility_id}")
def update_room_utility(
    utility_id: str,
    utility_update: RoomUtilityUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """방 공과금 수정"""
    # 공과금 존재 여부 확인
    utility = db.query(RoomUtility).filter(RoomUtility.id == utility_id).first()
    if not utility:
        raise HTTPException(status_code=404, detail="공과금을 찾을 수 없습니다")

    # 기존 값 저장 (로그용)
    old_values = {
        "room_id": str(utility.room_id),
        "utility_type": utility.utility_type,
        "period_start": str(utility.period_start),
        "period_end": str(utility.period_end),
        "usage": float(utility.usage) if utility.usage else None,
        "unit_price": float(utility.unit_price) if utility.unit_price else None,
        "total_amount": float(utility.total_amount) if utility.total_amount else None,
        "charge_month": str(utility.charge_month),
        "memo": utility.memo
    }

    try:
        # 업데이트할 데이터 준비
        update_data = utility_update.dict(exclude_unset=True)
        
        # 날짜 변환
        if "period_start" in update_data and update_data["period_start"]:
            try:
                update_data["period_start"] = datetime.strptime(update_data["period_start"], "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="검침 시작일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        
        if "period_end" in update_data and update_data["period_end"]:
            try:
                update_data["period_end"] = datetime.strptime(update_data["period_end"], "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="검침 종료일 형식이 올바르지 않습니다 (YYYY-MM-DD)")

        if "charge_month" in update_data and update_data["charge_month"]:
            try:
                update_data["charge_month"] = datetime.strptime(update_data["charge_month"], "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="청구 월 형식이 올바르지 않습니다 (YYYY-MM-DD)")

        # 기간 유효성 검사
        period_start = update_data.get("period_start", utility.period_start)
        period_end = update_data.get("period_end", utility.period_end)
        if period_end <= period_start:
            raise HTTPException(status_code=400, detail="검침 종료일은 시작일보다 늦어야 합니다")

        # 데이터 업데이트
        for field, value in update_data.items():
            setattr(utility, field, value)

        db.commit()
        db.refresh(utility)

        # 로그 생성
        create_database_log(
            db=db,
            table_name="room_utilities",
            record_id=str(utility.id),
            action="UPDATE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            new_values={
                "room_id": str(utility.room_id),
                "utility_type": utility.utility_type,
                "period_start": str(utility.period_start),
                "period_end": str(utility.period_end),
                "usage": float(utility.usage) if utility.usage else None,
                "unit_price": float(utility.unit_price) if utility.unit_price else None,
                "total_amount": float(utility.total_amount) if utility.total_amount else None,
                "charge_month": str(utility.charge_month),
                "memo": utility.memo
            },
            changed_fields=list(update_data.keys()),
            note="部屋光熱費修正"
        )

        return {
            "message": "光熱費が正常に修正されました",
            "id": str(utility.id),
            "utility_type": utility.utility_type,
            "total_amount": float(utility.total_amount) if utility.total_amount else None
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"公共費修正中にエラーが発生しました: {str(e)}"
        )

@app.delete("/rooms/{room_id}/utilities/{utility_id}")
def delete_room_utility_by_room(
    room_id: str,
    utility_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 방의 공과금 삭제"""
    # 방 존재 여부 확인
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 공과금 존재 여부 및 해당 방의 공과금인지 확인
    utility = db.query(RoomUtility).filter(
        RoomUtility.id == utility_id,
        RoomUtility.room_id == room_id
    ).first()
    
    if not utility:
        raise HTTPException(status_code=404, detail="해당 방의 공과금을 찾을 수 없습니다")

    # 삭제 전 값 저장 (로그용)
    old_values = {
        "room_id": str(utility.room_id),
        "utility_type": utility.utility_type,
        "period_start": str(utility.period_start),
        "period_end": str(utility.period_end),
        "usage": float(utility.usage) if utility.usage else None,
        "unit_price": float(utility.unit_price) if utility.unit_price else None,
        "total_amount": float(utility.total_amount) if utility.total_amount else None,
        "charge_month": str(utility.charge_month),
        "memo": utility.memo
    }
    
    try:
        db.delete(utility)
        db.commit()
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="room_utilities",
            record_id=str(utility.id),
            action="DELETE",
            user_id=current_user["id"] if current_user else None,
            old_values=old_values,
            note="部屋光熱費削除"
        )
        
        return {"message": "光熱費が正常に削除されました"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"公共費削除中にエラーが発生しました: {str(e)}"
        )

@app.get("/rooms/{room_id}/utilities")
def get_room_utilities_by_room(
    room_id: str,
    utility_type: Optional[str] = Query(None, description="公共費タイプでフィルタリング"),
    page: int = Query(1, description="ページ番号", ge=1),
    size: int = Query(10, description="ページあたりのアイテム数", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """特定部屋の公共費リストを取得"""
    # 部屋の存在確認
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="部屋が見つかりません")

    # 該当部屋の公共費を取得
    query = db.query(RoomUtility).filter(RoomUtility.room_id == room_id)

    # 公共費タイプでフィルタリング
    if utility_type:
        query = query.filter(RoomUtility.utility_type == utility_type)

    # 全アイテム数を計算
    total_count = query.count()

    # ページネーション適用 (最新順)
    # 페이지네이션 적용 (최신순)
    utilities = query.order_by(RoomUtility.charge_month.desc(), RoomUtility.created_at.desc()).offset((page - 1) * size).limit(size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + size - 1) // size

    # 응답 데이터 준비
    result = []
    for utility in utilities:
        utility_data = {
            "id": str(utility.id),
            "utility_type": utility.utility_type,
            "period_start": utility.period_start.strftime("%Y-%m-%d"),
            "period_end": utility.period_end.strftime("%Y-%m-%d"),
            "usage": float(utility.usage) if utility.usage else None,
            "unit_price": float(utility.unit_price) if utility.unit_price else None,
            "total_amount": float(utility.total_amount) if utility.total_amount else None,
            "charge_month": utility.charge_month.strftime("%Y-%m-%d"),
            "memo": utility.memo,
            "created_at": utility.created_at.strftime("%Y-%m-%d %H:%M:%S")
        }
        result.append(utility_data)

    return {
        "room": {
            "id": str(room.id),
            "room_number": room.room_number,
            "building_name": room.building.name if room.building else None
        },
        "items": result,
        "total": total_count,
        "total_pages": total_pages,
        "current_page": page
    }

@app.get("/room-utilities/{utility_id}/residents-during-period")
def get_residents_during_utility_period(
    utility_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """공과금 기간 동안의 거주자 정보 조회"""
    # 공과금 정보 조회
    utility = db.query(RoomUtility).filter(RoomUtility.id == utility_id).first()
    if not utility:
        raise HTTPException(status_code=404, detail="공과금을 찾을 수 없습니다")

    # 해당 방의 모든 거주 기록 조회
    residents = db.query(Resident).filter(
        Resident.room_id == utility.room_id
    ).all()

    # 공과금 기간 동안의 거주자 정보 계산
    residents_during_period = []
    total_days_in_period = (utility.period_end - utility.period_start).days + 1

    for resident in residents:
        # 거주 기간과 공과금 기간의 겹치는 부분 계산
        resident_start = resident.check_in_date
        resident_end = resident.check_out_date if resident.check_out_date else utility.period_end

        # 겹치는 기간 계산
        overlap_start = max(resident_start, utility.period_start)
        overlap_end = min(resident_end, utility.period_end)

        if overlap_start <= overlap_end:
            # 겹치는 일수 계산
            overlap_days = (overlap_end - overlap_start).days + 1
            
            # 비율 계산
            ratio = overlap_days / total_days_in_period if total_days_in_period > 0 else 0
            
            # 예상 부담금액 계산
            if utility.total_amount:
                estimated_amount = float(utility.total_amount) * ratio
            else:
                estimated_amount = 0

            resident_info = {
                "student_id": str(resident.resident_id),
                "student_name": resident.student.name,
                "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
                "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
                "is_active": resident.is_active,
                "overlap_start": overlap_start.strftime("%Y-%m-%d"),
                "overlap_end": overlap_end.strftime("%Y-%m-%d"),
                "overlap_days": overlap_days,
                "ratio": round(ratio * 100, 2),  # 퍼센트로 표시
                "estimated_amount": round(estimated_amount, 2)
            }
            residents_during_period.append(resident_info)

    # 일수 기준으로 정렬 (많은 일수부터)
    residents_during_period.sort(key=lambda x: x["overlap_days"], reverse=True)

    return {
        "utility": {
            "id": str(utility.id),
            "utility_type": utility.utility_type,
            "period_start": utility.period_start.strftime("%Y-%m-%d"),
            "period_end": utility.period_end.strftime("%Y-%m-%d"),
            "total_amount": float(utility.total_amount) if utility.total_amount else 0,
            "total_days": total_days_in_period
        },
        "residents": residents_during_period,
        "summary": {
            "total_residents": len(residents_during_period),
            "total_ratio": sum(r["ratio"] for r in residents_during_period),
            "total_estimated_amount": sum(r["estimated_amount"] for r in residents_during_period)
        }
    }

@app.post("/room-utilities/{utility_id}/calculate-allocation")
def calculate_utility_allocation(
    utility_id: str,
    allocation_method: str = Query("days_based", description="배분 방법 (days_based 또는 usage_based)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """공과금 배분 계산"""
    # 공과금 정보 조회
    utility = db.query(RoomUtility).filter(RoomUtility.id == utility_id).first()
    if not utility:
        raise HTTPException(status_code=404, detail="공과금을 찾을 수 없습니다")

    # 해당 방의 모든 거주 기록 조회
    residents = db.query(Resident).filter(
        Resident.room_id == utility.room_id
    ).all()

    # 공과금 기간 동안의 거주자 정보 계산
    allocations = []
    total_days_in_period = (utility.period_end - utility.period_start).days + 1
    total_overlap_days = 0

    # 1단계: 각 거주자의 겹치는 일수 계산
    resident_overlaps = []
    for resident in residents:
        resident_start = resident.check_in_date
        resident_end = resident.check_out_date if resident.check_out_date else utility.period_end

        overlap_start = max(resident_start, utility.period_start)
        overlap_end = min(resident_end, utility.period_end)

        if overlap_start <= overlap_end:
            overlap_days = (overlap_end - overlap_start).days + 1
            total_overlap_days += overlap_days
            
            resident_overlaps.append({
                "resident": resident,
                "overlap_days": overlap_days,
                "overlap_start": overlap_start,
                "overlap_end": overlap_end
            })

    # 2단계: 배분 계산
    if allocation_method == "days_based":
        # 일수 기준 배분
        for overlap_info in resident_overlaps:
            resident = overlap_info["resident"]
            overlap_days = overlap_info["overlap_days"]
            
            # 비율 계산
            ratio = overlap_days / total_overlap_days if total_overlap_days > 0 else 0
            
            # 부담금액 계산
            amount = float(utility.total_amount) * ratio if utility.total_amount else 0

            allocation = {
                "student_id": str(resident.resident_id),
                "student_name": resident.student.name,
                "days_used": overlap_days,
                "ratio": round(ratio * 100, 2),
                "amount": round(amount, 2),
                "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
                "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
                "is_active": resident.is_active
            }
            allocations.append(allocation)

    elif allocation_method == "usage_based":
        # 사용량 기준 배분 (현재는 일수 기준과 동일하게 처리)
        # 향후 실제 사용량 데이터가 있으면 여기서 계산 로직 수정
        for overlap_info in resident_overlaps:
            resident = overlap_info["resident"]
            overlap_days = overlap_info["overlap_days"]
            
            ratio = overlap_days / total_overlap_days if total_overlap_days > 0 else 0
            amount = float(utility.total_amount) * ratio if utility.total_amount else 0

            allocation = {
                "student_id": str(resident.resident_id),
                "student_name": resident.student.name,
                "days_used": overlap_days,
                "ratio": round(ratio * 100, 2),
                "amount": round(amount, 2),
                "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
                "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
                "is_active": resident.is_active
            }
            allocations.append(allocation)

    else:
        raise HTTPException(status_code=400, detail="サポートされていない配分方法です")

    # 일수 기준으로 정렬
    allocations.sort(key=lambda x: x["days_used"], reverse=True)

    return {
        "utility": {
            "id": str(utility.id),
            "utility_type": utility.utility_type,
            "period_start": utility.period_start.strftime("%Y-%m-%d"),
            "period_end": utility.period_end.strftime("%Y-%m-%d"),
            "total_amount": float(utility.total_amount) if utility.total_amount else 0,
            "total_days": total_days_in_period
        },
        "allocation_method": allocation_method,
        "allocations": allocations,
        "summary": {
            "total_residents": len(allocations),
            "total_days": total_overlap_days,
            "total_amount": sum(a["amount"] for a in allocations),
            "total_ratio": sum(a["ratio"] for a in allocations)
        }
    }

@app.get("/buildings/monthly-invoice-preview/students/{year}/{month}")
def get_monthly_invoice_preview_by_students(
    year: int,
    month: int,
    company_id: Optional[str] = Query(None, description="특정 회사로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """해당 년도 월의 모든 학생별 청구서 미리보기 데이터 조회 (PDF 다운로드 전 프론트 표시용)"""
    print(f"[DEBUG] get_monthly_invoice_preview 시작 - year: {year}, month: {month}, company_id: {company_id}")
    
    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 전월의 시작/종료일 계산 (7월 데이터 요청 시 6월 거주 기록 조회)
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1
    
    # 전월의 시작/종료일 계산
    if prev_month == 12:
        prev_month_end_date = date(prev_year + 1, 1, 1) - timedelta(days=1)
    else:
        prev_month_end_date = date(prev_year, prev_month + 1, 1) - timedelta(days=1)
    
    prev_month_start_date = date(prev_year, prev_month, 1)
    total_days_in_month = (prev_month_end_date - prev_month_start_date).days + 1

    # 1. 회사별 학생 리스트 조회
    print(f"[DEBUG] 회사별 학생 리스트 조회 시작")
    if company_id:
        # 특정 회사 학생들 조회
        students_query = db.query(Student).filter(Student.company_id == company_id)
    else:
        # 모든 학생 조회
        students_query = db.query(Student)
    
    students = students_query.all()
    print(f"[DEBUG] 조회된 학생 수: {len(students)}")

    # 2. 각 학생별로 해당 월에 거주하는지 확인하고 데이터 수집
    students_data = []
    total_electricity_amount = 0
    total_water_amount = 0
    total_gas_amount = 0
    total_rent_amount = 0
    total_wifi_amount = 0

    for student in students:
        print(f"[DEBUG] 학생 처리 중: {student.name}")
        
        # 해당 학생의 전월 거주 정보 조회 (모든 거주 기록)
        resident_records = db.query(Resident).options(
            joinedload(Resident.room).joinedload(Room.building)
        ).filter(
            Resident.resident_id == student.id,
            Resident.check_in_date <= prev_month_end_date,
            (Resident.check_out_date.is_(None) | (Resident.check_out_date >= prev_month_start_date))
        ).all()

        if not resident_records:
            print(f"[DEBUG] {student.name} - 해당 월에 거주 정보 없음")
            continue

        # 모든 거주 기록에 대해 계산
        total_rent_amount_for_student = 0
        total_wifi_amount_for_student = 0
        total_electricity_amount_for_student = 0
        total_water_amount_for_student = 0
        total_gas_amount_for_student = 0
        all_utilities_data = []
        has_utilities_for_student = False

        print(f"[DEBUG] {student.name} - 거주 기록 수: {len(resident_records)}")
        
        for resident_info in resident_records:
            # 전월에 거주한 일수 계산
            # 실제 입주일과 전월 시작일 중 늦은 날짜를 사용
            if resident_info.check_in_date >= prev_month_start_date:
                check_in = resident_info.check_in_date
            else:
                check_in = prev_month_start_date
            
            # 퇴실일이 None이면 전월 전체 거주로 인식
            if resident_info.check_out_date is None:
                check_out = prev_month_end_date
            else:
                # 실제 퇴실일과 전월 종료일 중 이른 날짜를 사용
                if resident_info.check_out_date <= prev_month_end_date:
                    check_out = resident_info.check_out_date
                else:
                    check_out = prev_month_end_date

            # 해당 월에 거주한 일수 계산
            if check_in <= check_out:
                days_in_month = (check_out - check_in).days + 1
            else:
                days_in_month = 0

            if days_in_month <= 0:
                print(f"[DEBUG] {student.name} - 해당 월에 거주하지 않음")
                continue

            print(f"[DEBUG] {student.name} - check_in: {check_in}, check_out: {check_out}, days_in_month: {days_in_month}")
            print(f"[DEBUG] {student.name} - prev_month_start_date: {prev_month_start_date}, prev_month_end_date: {prev_month_end_date}")

            # 야칭 계산
            print(f"[DEBUG] {student.name} - check_in_date: {resident_info.check_in_date}, check_out_date: {resident_info.check_out_date}")
            
            # 퇴실일이 있는 경우와 없는 경우를 구분하여 계산
            if resident_info.check_out_date is not None:
                # 퇴실일이 있는 경우
                check_in_month = resident_info.check_in_date.month
                check_in_year = resident_info.check_in_date.year
                check_out_month = resident_info.check_out_date.month
                check_out_year = resident_info.check_out_date.year
                
                # check_in_date와 check_out_date가 같은 달이 아닌 경우
                if (check_in_year != check_out_year) or (check_in_month != check_out_month):
                    # 퇴실일 달 기준으로 계산
                    rent_amount = min(days_in_month * 1000, 30000)
                    print(f"[DEBUG] {student.name} - 퇴실일 달 기준 계산: {days_in_month}일 × 1,000 = {rent_amount}엔 (최대 30,000엔)")
                else:
                    # 같은 달인 경우 - 퇴실일이 있으면 항상 일별 계산
                    rent_amount = min(days_in_month * 1000, 30000)
                    print(f"[DEBUG] {student.name} - 같은 달, 퇴실일 있음: {days_in_month}일 × 1,000 = {rent_amount}엔 (최대 30,000엔)")
            else:
                # 퇴실일이 없는 경우
                # 입주일이 1일이고 해당 월 전체를 거주하는 경우만 30,000엔
                if resident_info.check_in_date.day == 1 and days_in_month >= 30:
                    rent_amount = 30000
                    print(f"[DEBUG] {student.name} - 퇴실일 없음, 1일 입주, 전체 월 거주: 30,000엔")
                else:
                    # 그 외의 경우는 일별 계산
                    rent_amount = min(days_in_month * 1000, 30000)
                    print(f"[DEBUG] {student.name} - 퇴실일 없음, 일별 계산: {days_in_month}일 × 1,000 = {rent_amount}엔 (최대 30,000엔)")

            # 와이파이 비용 계산
            if resident_info.check_out_date is not None:
                # 퇴실일이 있는 경우
                check_in_month = resident_info.check_in_date.month
                check_in_year = resident_info.check_in_date.year
                check_out_month = resident_info.check_out_date.month
                check_out_year = resident_info.check_out_date.year
                
                # check_in_date와 check_out_date가 같은 달이 아닌 경우
                if (check_in_year != check_out_year) or (check_in_month != check_out_month):
                    # 퇴실일 달 기준으로 계산
                    wifi_amount = min(int(days_in_month * (700 / 30)), 700)
                    print(f"[DEBUG] {student.name} - 와이파이 퇴실일 달 기준 계산: {days_in_month}일 × (700/30) = {wifi_amount}엔 (최대 700엔)")
                else:
                    # 같은 달인 경우 - 퇴실일이 있으면 항상 일별 계산
                    wifi_amount = min(int(days_in_month * (700 / 30)), 700)
                    print(f"[DEBUG] {student.name} - 와이파이 같은 달, 퇴실일 있음: {days_in_month}일 × (700/30) = {wifi_amount}엔 (최대 700엔)")
            else:
                # 퇴실일이 없는 경우
                # 입주일이 1일이고 해당 월 전체를 거주하는 경우만 700엔 고정
                if resident_info.check_in_date.day == 1 and days_in_month >= 30:
                    wifi_amount = 700
                    print(f"[DEBUG] {student.name} - 와이파이 퇴실일 없음, 1일 입주, 전체 월 거주: 700엔")
                else:
                    # 그 외의 경우는 일별 계산
                    wifi_amount = min(int(days_in_month * (700 / 30)), 700)
                    print(f"[DEBUG] {student.name} - 와이파이 퇴실일 없음, 일별 계산: {days_in_month}일 × (700/30) = {wifi_amount}엔 (최대 700엔)")

            # 3. 공과금 조회 및 계산
            utilities_data = []
            room_electricity_amount = 0
            room_water_amount = 0
            room_gas_amount = 0
            
            print(f"[DEBUG] {student.name} - 현재 거주 기록 계산 완료: rent={rent_amount}, wifi={wifi_amount}")
            has_utilities = False

            if resident_info.room:
                # 해당 방의 해당 월 공과금 조회 (7월 계산으로 입력된 공과금)
                utilities = db.query(RoomUtility).filter(
                    RoomUtility.room_id == resident_info.room.id,
                    RoomUtility.charge_month == date(year, month, 1)
                ).all()

                print(f"[DEBUG] {student.name}의 방 {resident_info.room.room_number} 공과금 개수: {len(utilities)}")

                if utilities:
                    has_utilities = True
                    for utility in utilities:
                        # 해당 유틸리티 기간 내 방 거주자 전체 쿼리
                        all_residents = db.query(Resident).filter(
                            Resident.room_id == resident_info.room.id,
                            Resident.check_in_date <= utility.period_end,
                            (Resident.check_out_date.is_(None) | (Resident.check_out_date >= utility.period_start))
                        ).all()

                        # 전체 person-day 계산
                        total_person_days = 0
                        for r in all_residents:
                            overlap_in = max(r.check_in_date, utility.period_start)
                            overlap_out = min(r.check_out_date or utility.period_end, utility.period_end)
                            overlap_days = (overlap_out - overlap_in).days + 1 if overlap_in <= overlap_out else 0
                            total_person_days += overlap_days

                        # 해당 학생의 공과금 기간 내 거주 일수 계산
                        student_overlap_in = max(resident_info.check_in_date, utility.period_start)
                        student_overlap_out = min(resident_info.check_out_date or utility.period_end, utility.period_end)
                        student_days = (student_overlap_out - student_overlap_in).days + 1 if student_overlap_in <= student_overlap_out else 0

                        # 학생별 부담액 계산
                        if total_person_days > 0 and student_days > 0:
                            # 1일당 요금 계산
                            per_day = float(utility.total_amount) / total_person_days
                            # 학생별 부담액 = 1일당 요금 × 학생 거주 일수
                            student_amount = int(per_day * student_days)
                        else:
                            student_amount = 0

                        utilities_data.append({
                            "utility_type": utility.utility_type,
                            "period_start": utility.period_start.strftime("%Y-%m-%d"),
                            "period_end": utility.period_end.strftime("%Y-%m-%d"),
                            "total_amount": float(utility.total_amount),
                            "student_days": student_days,
                            "total_person_days": total_person_days,
                            "student_amount": student_amount,
                            "room_number": resident_info.room.room_number
                        })
                        
                        # 공과금 유형별로 금액 누적
                        if utility.utility_type == "electricity":
                            room_electricity_amount += student_amount
                        elif utility.utility_type == "water":
                            room_water_amount += student_amount
                        elif utility.utility_type == "gas":
                            room_gas_amount += student_amount
                        
                        print(f"[DEBUG] {student.name} - 방 {resident_info.room.room_number} 유틸리티: {utility.utility_type} = {student_amount}엔")

            # 각 거주 기록별 데이터 누적
            print(f"[DEBUG] {student.name} - 거주 기록 누적 전: rent={total_rent_amount_for_student}, wifi={total_wifi_amount_for_student}")
            print(f"[DEBUG] {student.name} - 현재 거주 기록: rent={rent_amount}, wifi={wifi_amount}")
            
            total_rent_amount_for_student += rent_amount
            total_wifi_amount_for_student += wifi_amount
            total_electricity_amount_for_student += room_electricity_amount
            total_water_amount_for_student += room_water_amount
            total_gas_amount_for_student += room_gas_amount
            if has_utilities:
                has_utilities_for_student = True
                all_utilities_data.extend(utilities_data)
            
            print(f"[DEBUG] {student.name} - 거주 기록 누적 후: rent={total_rent_amount_for_student}, wifi={total_wifi_amount_for_student}")
            
            # 방 이동으로 인한 총액 제한 적용
            total_rent_amount_for_student = min(total_rent_amount_for_student, 30000)
            total_wifi_amount_for_student = min(total_wifi_amount_for_student, 700)
            
            print(f"[DEBUG] {student.name} - 제한 적용 후: rent={total_rent_amount_for_student}, wifi={total_wifi_amount_for_student}")

        # 방번호와 건물명 처리 (복수 거주인 경우)
        if len(resident_records) > 1:
            room_numbers = [r.room.room_number for r in resident_records]
            building_names = [r.room.building.name for r in resident_records]
            room_number = ",".join(room_numbers)
            # 건물 이름이 같으면 중복 제거
            unique_building_names = list(set(building_names))
            building_name = ",".join(unique_building_names)
        else:
            room_number = resident_records[0].room.room_number
            building_name = resident_records[0].room.building.name
        
        # 학생 데이터 구성 (모든 거주 기록을 합산)
        student_data = {
            "student_id": str(student.id),
            "student_name": student.name,
            "student_type": student.student_type,
            "grade_name": student.grade.name if student.grade else None,
            "room_number": room_number,
            "building_name": building_name,
            "days_in_month": sum([((r.check_out_date or prev_month_end_date) - max(r.check_in_date, prev_month_start_date)).days + 1 for r in resident_records if (r.check_out_date or prev_month_end_date) >= max(r.check_in_date, prev_month_start_date)]),
            "rent_amount": total_rent_amount_for_student,
            "wifi_amount": total_wifi_amount_for_student,
            "has_utilities": has_utilities_for_student,
            "utilities": all_utilities_data if has_utilities_for_student else [],
            "electricity_amount": total_electricity_amount_for_student if has_utilities_for_student else 0,
            "water_amount": total_water_amount_for_student if has_utilities_for_student else 0,
            "gas_amount": total_gas_amount_for_student if has_utilities_for_student else 0,
            "total_utilities_amount": (total_electricity_amount_for_student + total_water_amount_for_student + total_gas_amount_for_student) if has_utilities_for_student else 0,
            "rent_wifi_total": total_rent_amount_for_student + total_wifi_amount_for_student,
            "utilities_total": (total_electricity_amount_for_student + total_water_amount_for_student + total_gas_amount_for_student) if has_utilities_for_student else 0,
            "total_amount": total_rent_amount_for_student + total_wifi_amount_for_student + (total_electricity_amount_for_student + total_water_amount_for_student + total_gas_amount_for_student) if has_utilities_for_student else 0
        }

        students_data.append(student_data)
        total_rent_amount += total_rent_amount_for_student
        total_wifi_amount += total_wifi_amount_for_student
        total_electricity_amount += total_electricity_amount_for_student if has_utilities_for_student else 0
        total_water_amount += total_water_amount_for_student if has_utilities_for_student else 0
        total_gas_amount += total_gas_amount_for_student if has_utilities_for_student else 0

        print(f"[DEBUG] {student.name} 데이터 추가 - rent: {total_rent_amount_for_student}, wifi: {total_wifi_amount_for_student}, electricity: {total_electricity_amount_for_student if has_utilities_for_student else 0}, water: {total_water_amount_for_student if has_utilities_for_student else 0}, gas: {total_gas_amount_for_student if has_utilities_for_student else 0}, total: {total_rent_amount_for_student + total_wifi_amount_for_student + (total_electricity_amount_for_student + total_water_amount_for_student + total_gas_amount_for_student) if has_utilities_for_student else 0}")

    print(f"[DEBUG] 최종 결과 - 학생 수: {len(students_data)}, 총 야칭: {total_rent_amount}, 총 와이파이: {total_wifi_amount}, 총 전기: {total_electricity_amount}, 총 수도: {total_water_amount}, 총 가스: {total_gas_amount}")
    
    return {
        "year": year,
        "month": month,
        "billing_period": f"{prev_year}년 {prev_month}월",
        "total_students": len(students_data),
        "students": students_data,
        "summary": {
            "total_electricity_amount": total_electricity_amount,
            "total_water_amount": total_water_amount,
            "total_gas_amount": total_gas_amount,
            "total_utilities_amount": total_electricity_amount + total_water_amount + total_gas_amount,
            "total_rent_amount": total_rent_amount,
            "total_wifi_amount": total_wifi_amount,
            "grand_total": total_rent_amount + total_wifi_amount + total_electricity_amount + total_water_amount + total_gas_amount
        }
    }

@app.get("/rooms/{room_id}/utilities/monthly-allocation/{year}/{month}")
def get_monthly_utility_allocation(
    room_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 방의 월별 공과금 배분 계산"""
    # 방 존재 여부 확인
    room = db.query(Room).options(
        joinedload(Room.building)
    ).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 해당 월의 시작일과 종료일 계산
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # 해당 월의 모든 공과금 조회
    utilities = db.query(RoomUtility).filter(
        RoomUtility.room_id == room_id,
        RoomUtility.charge_month >= start_date,
        RoomUtility.charge_month <= end_date
    ).order_by(RoomUtility.utility_type, RoomUtility.period_start).all()

    if not utilities:
        return {
            "room": {
                "id": str(room.id),
                "room_number": room.room_number,
                "building_name": room.building.name if room.building else None
            },
            "year": year,
            "month": month,
            "utilities": [],
            "summary": {
                "total_utilities": 0,
                "total_amount": 0,
                "total_residents": 0
            }
        }

    # 해당 방의 모든 거주 기록 조회
    residents = db.query(Resident).options(
        joinedload(Resident.student)
    ).filter(Resident.room_id == room_id).all()

    # 각 공과금별 배분 계산
    utility_allocations = []
    total_monthly_amount = 0

    for utility in utilities:
        # 공과금 기간 동안의 거주자 정보 계산
        allocations = []
        total_days_in_period = (utility.period_end - utility.period_start).days + 1
        total_overlap_days = 0

        # 각 거주자의 겹치는 일수 계산
        resident_overlaps = []
        for resident in residents:
            resident_start = resident.check_in_date
            resident_end = resident.check_out_date if resident.check_out_date else utility.period_end

            overlap_start = max(resident_start, utility.period_start)
            overlap_end = min(resident_end, utility.period_end)

            if overlap_start <= overlap_end:
                overlap_days = (overlap_end - overlap_start).days + 1
                total_overlap_days += overlap_days
                
                resident_overlaps.append({
                    "resident": resident,
                    "overlap_days": overlap_days,
                    "overlap_start": overlap_start,
                    "overlap_end": overlap_end
                })

        # 배분 계산 (일수 기준)
        for overlap_info in resident_overlaps:
            resident = overlap_info["resident"]
            overlap_days = overlap_info["overlap_days"]
            
            # 비율 계산
            ratio = overlap_days / total_overlap_days if total_overlap_days > 0 else 0
            
            # 부담금액 계산
            amount = float(utility.total_amount) * ratio if utility.total_amount else 0

            allocation = {
                "student_id": str(resident.resident_id),
                "student_name": resident.student.name,
                "days_used": overlap_days,
                "ratio": round(ratio * 100, 2),
                "amount": round(amount, 2),
                "check_in_date": resident.check_in_date.strftime("%Y-%m-%d"),
                "check_out_date": resident.check_out_date.strftime("%Y-%m-%d") if resident.check_out_date else None,
                "is_active": resident.is_active,
                "overlap_start": overlap_info["overlap_start"].strftime("%Y-%m-%d"),
                "overlap_end": overlap_info["overlap_end"].strftime("%Y-%m-%d")
            }
            allocations.append(allocation)

        # 일수 기준으로 정렬
        allocations.sort(key=lambda x: x["days_used"], reverse=True)

        utility_allocation = {
            "utility_id": str(utility.id),
            "utility_type": utility.utility_type,
            "period_start": utility.period_start.strftime("%Y-%m-%d"),
            "period_end": utility.period_end.strftime("%Y-%m-%d"),
            "total_amount": float(utility.total_amount) if utility.total_amount else 0,
            "total_days": total_days_in_period,
            "charge_month": utility.charge_month.strftime("%Y-%m-%d"),
            "memo": utility.memo,
            "allocations": allocations,
            "summary": {
                "total_residents": len(allocations),
                "total_days": total_overlap_days,
                "total_amount": sum(a["amount"] for a in allocations),
                "total_ratio": sum(a["ratio"] for a in allocations)
            }
        }
        
        utility_allocations.append(utility_allocation)
        total_monthly_amount += float(utility.total_amount) if utility.total_amount else 0

    # 전체 월 요약
    all_students = set()
    for utility in utility_allocations:
        for allocation in utility["allocations"]:
            all_students.add(allocation["student_id"])

    return {
        "room": {
            "id": str(room.id),
            "room_number": room.room_number,
            "building_name": room.building.name if room.building else None
        },
        "year": year,
        "month": month,
        "utilities": utility_allocations,
        "summary": {
            "total_utilities": len(utility_allocations),
            "total_amount": total_monthly_amount,
            "total_residents": len(all_students),
            "utility_types": list(set(u["utility_type"] for u in utility_allocations))
        }
    }

@app.get("/rooms/{room_id}/utilities/monthly-summary/{year}/{month}")
def get_monthly_utility_summary(
    room_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 방의 월별 공과금 요약 정보"""
    # 방 존재 여부 확인
    room = db.query(Room).options(
        joinedload(Room.building)
    ).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")

    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 해당 월의 시작일과 종료일 계산
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # 해당 월의 모든 공과금 조회
    utilities = db.query(RoomUtility).filter(
        RoomUtility.room_id == room_id,
        RoomUtility.charge_month >= start_date,
        RoomUtility.charge_month <= end_date
    ).all()

    # 해당 방의 모든 거주 기록 조회
    residents = db.query(Resident).options(
        joinedload(Resident.student)
    ).filter(Resident.room_id == room_id).all()

    # 공과금별 요약 정보
    utility_summaries = []
    total_monthly_amount = 0

    for utility in utilities:
        # 공과금 기간 동안의 거주자 수와 일수 계산
        total_days_in_period = (utility.period_end - utility.period_start).days + 1
        total_overlap_days = 0
        resident_count = 0

        for resident in residents:
            resident_start = resident.check_in_date
            resident_end = resident.check_out_date if resident.check_out_date else utility.period_end

            overlap_start = max(resident_start, utility.period_start)
            overlap_end = min(resident_end, utility.period_end)

            if overlap_start <= overlap_end:
                overlap_days = (overlap_end - overlap_start).days + 1
                total_overlap_days += overlap_days
                resident_count += 1

        # 평균 일수 계산
        avg_days_per_resident = total_overlap_days / resident_count if resident_count > 0 else 0

        utility_summary = {
            "utility_id": str(utility.id),
            "utility_type": utility.utility_type,
            "period_start": utility.period_start.strftime("%Y-%m-%d"),
            "period_end": utility.period_end.strftime("%Y-%m-%d"),
            "total_amount": float(utility.total_amount) if utility.total_amount else 0,
            "total_days": total_days_in_period,
            "resident_count": resident_count,
            "total_resident_days": total_overlap_days,
            "avg_days_per_resident": round(avg_days_per_resident, 1),
            "amount_per_day": round(float(utility.total_amount) / total_overlap_days, 2) if total_overlap_days > 0 else 0,
            "amount_per_resident": round(float(utility.total_amount) / resident_count, 2) if resident_count > 0 else 0,
            "memo": utility.memo
        }
        
        utility_summaries.append(utility_summary)
        total_monthly_amount += float(utility.total_amount) if utility.total_amount else 0

    return {
        "room": {
            "id": str(room.id),
            "room_number": room.room_number,
            "building_name": room.building.name if room.building else None
        },
        "year": year,
        "month": month,
        "utilities": utility_summaries,
        "summary": {
            "total_utilities": len(utility_summaries),
            "total_amount": total_monthly_amount,
            "utility_types": list(set(u["utility_type"] for u in utility_summaries))
        }
    }

@app.get("/rooms/{room_id}/utilities/monthly-by-students/{year}/{month}")
def get_student_monthly_utility_allocation(
    room_id: str,
    student_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """특정 방, 특정 학생, 특정 월의 공과금 분배 정보 반환 (전달 기준)"""
    # 1. 방/학생 존재 체크
    room = db.query(Room).options(joinedload(Room.building)).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="방을 찾을 수 없습니다")
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    # 2. 전달 기준으로 시작/종료일 계산 (예: 7월 조회 시 6월 1일~6월 30일)
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1
    
    start_date = date(prev_year, prev_month, 1)
    if prev_month == 12:
        end_date = date(prev_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(prev_year, prev_month + 1, 1) - timedelta(days=1)

    # 3. 전달의 모든 공과금 조회

    target_charge_month = date(year, month, 1)
    utilities = db.query(RoomUtility).filter(
        RoomUtility.room_id == room_id,
        RoomUtility.charge_month == target_charge_month
    ).all()

    if not utilities:
        return {
            "student_id": student_id,
            "student_name": student.name,
            "utilities": [],
            "total_amount": 0,
            "detail": f"{year}년 {month}월 조회 시 전달({prev_year}년 {prev_month}월)의 공과금 내역이 없습니다."
        }

    # 4. 학생의 해당 방에 대한 거주 이력 (입주~퇴실)
    resident = db.query(Resident).filter(
        Resident.room_id == room_id,
        Resident.resident_id == student_id,
        Resident.check_in_date <= end_date,
        (Resident.check_out_date.is_(None) | (Resident.check_out_date >= start_date))
    ).first()
    print(f"DEBUG: 날짜!!! - 시작: {end_date}, 종료: {start_date}")
    if not resident:
        return {
            "student_id": student_id,
            "student_name": student.name,
            "utilities": [],
            "total_amount": 0,
            "detail": f"{year}년 {month}월 조회 시 전달({prev_year}년 {prev_month}월)에 거주 이력이 없습니다."
        }

    # 5. 각 공과금별 배분 계산
    utilities_result = []
    total_amount = 0

    for utility in utilities:
        print(f"DEBUG: 공과금 처리 시작 - ID: {utility.id}, 유형: {utility.utility_type}, 총액: {utility.total_amount}")
        print(f"DEBUG: 공과금 기간 - 시작: {utility.period_start}, 종료: {utility.period_end}")
        
        # 해당 유틸리티 기간 내 방 거주자 전체 쿼리
        all_residents = db.query(Resident).filter(
            Resident.room_id == room_id,
            Resident.check_in_date <= utility.period_end,
            (Resident.check_out_date.is_(None) | (Resident.check_out_date >= utility.period_start))
        ).all()

        print(f"DEBUG: 해당 공과금 기간 내 거주자 수: {len(all_residents)}")

        # 이 학생이 실제로 해당 공과금 기간 내에 며칠 있었는지 계산
        stu_in = max(resident.check_in_date, utility.period_start)
        stu_out = min(resident.check_out_date or utility.period_end, utility.period_end)
        days = (stu_out - stu_in).days + 1 if stu_in <= stu_out else 0

        print(f"DEBUG: 학생 {student.name} - 거주기간: {resident.check_in_date}~{resident.check_out_date}")
        print(f"DEBUG: 학생 {student.name} - 겹치는 기간: {stu_in}~{stu_out}, 겹치는 일수: {days}")

        # 전체 person-day 계산
        total_person_days = 0
        for r in all_residents:
            overlap_in = max(r.check_in_date, utility.period_start)
            overlap_out = min(r.check_out_date or utility.period_end, utility.period_end)
            overlap_days = (overlap_out - overlap_in).days + 1 if overlap_in <= overlap_out else 0
            total_person_days += overlap_days
            
            print(f"DEBUG: 거주자 {r.student.name if r.student else 'Unknown'} - 겹치는 일수: {overlap_days}")

        print(f"DEBUG: 전체 person-days: {total_person_days}")

        # 1일당 요금 계산 및 본인 부담액 계산
        per_day = float(utility.total_amount) / total_person_days if total_person_days > 0 else 0
        my_amount = round(per_day * days, 2)

        print(f"DEBUG: 1일당 요금: {per_day:.2f}, 학생 부담액: {my_amount:.2f}")

        utilities_result.append({
            "utility_id": str(utility.id),
            "utility_type": utility.utility_type,
            "period_start": utility.period_start.strftime("%Y-%m-%d"),
            "period_end": utility.period_end.strftime("%Y-%m-%d"),
            "charge_month": utility.charge_month.strftime("%Y-%m-%d"),
            "my_days": days,
            "total_person_days": total_person_days,
            "amount_per_day": round(per_day, 2),
            "my_amount": my_amount,
            "memo": utility.memo
        })
        total_amount += my_amount

    print(f"DEBUG: 총 부담액: {total_amount:.2f}")

    return {
        "student_id": student_id,
        "student_name": student.name,
        "query_month": f"{year}년 {month}월",
        "billing_period": f"{prev_year}년 {prev_month}월",
        "utilities": utilities_result,
        "total_amount": round(total_amount, 2)
    }

@app.get("/buildings/monthly-invoice-preview/rooms/{year}/{month}")
def get_monthly_invoice_preview_by_rooms(
    year: int,
    month: int,
    company_id: Optional[str] = Query(None, description="특정 회사로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """해당 년도 월의 모든 방 공과금 미리보기 데이터 조회 (PDF 다운로드 전 프론트 표시용)"""
    print(f"[DEBUG] get_monthly_invoice_preview 시작 - year: {year}, month: {month}, company_id: {company_id}")
    
    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 전달 기준으로 시작/종료일 계산 (예: 7월 조회 시 6월 1일~6월 30일)
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1
    
    start_date = date(prev_year, prev_month, 1)
    if prev_month == 12:
        end_date = date(prev_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(prev_year + 1, prev_month + 1, 1) - timedelta(days=1)

    target_charge_month = date(year, month, 1)
    
    print(f"[DEBUG] 날짜 계산 완료 - start_date: {start_date}, end_date: {end_date}, target_charge_month: {target_charge_month}")

    # 공과금이 있는 방들 조회 (회사 필터 적용)
    print(f"[DEBUG] 공과금이 있는 방들 조회 시작")
    query = db.query(Room).options(
        joinedload(Room.building)
    ).join(RoomUtility).filter(
        RoomUtility.charge_month == target_charge_month
    )
    
    if company_id:
        print(f"[DEBUG] 회사 필터 적용 - company_id: {company_id}")
        # 회사에 속한 학생들이 거주하는 방들만 필터링
        query = query.join(Resident).join(Student).filter(
            Student.company_id == company_id
        )
    
    rooms_with_utilities = query.distinct().all()
    print(f"[DEBUG] 조회된 방 개수: {len(rooms_with_utilities)}")

    # rooms_with_utilities가 없어도 야칭 정보는 포함해서 반환
    if not rooms_with_utilities:
        print(f"[DEBUG] 공과금이 있는 방이 없음 - 야칭 정보만 조회")
        # 회사에 속한 학생들의 야칭 정보 조회
        rent_data = []
        total_rent_amount = 0
        
        if company_id:
            # 해당 월의 총 일수 계산
            if month == 12:
                month_end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end_date = date(year, month + 1, 1) - timedelta(days=1)
            
            month_start_date = date(year, month, 1)
            total_days_in_month = (month_end_date - month_start_date).days + 1

            # 회사에 속한 학생들의 거주 정보 조회
            print(f"[DEBUG] 회사 학생 거주 정보 조회 시작")
            print(f"[DEBUG] 스타트 날짜: {month_start_date}")
            print(f"[DEBUG] 종료 날짜: {month_end_date}")
            company_students = db.query(Resident).options(
                joinedload(Resident.student)
            ).join(Student).filter(
                Student.company_id == company_id,
                Resident.check_in_date <= month_end_date,
                (Resident.check_out_date.is_(None) | (Resident.check_out_date >= month_start_date))
            ).all()
            print(f"[DEBUG] 조회된 회사 학생 수: {len(company_students)}")
            
            # 각 학생별 야칭 계산
            print(f"[DEBUG] 학생별 야칭 계산 시작 - 총 {len(company_students)}명")
            for resident in company_students:
                student_name = resident.student.name
                print(f"[DEBUG] 학생 처리 중: {student_name}")
                
                # 해당 월에 거주한 일수 계산
                check_in = max(resident.check_in_date, month_start_date)
                
                # 퇴실일이 None이면 해당 월 전체 거주로 인식
                if resident.check_out_date is None:
                    check_in = max(resident.check_in_date, month_start_date)
                    check_out = month_end_date
                else:
                    check_in = resident.check_in_date
                    check_out = min(resident.check_out_date, month_end_date)

                # 해당 월에 거주한 일수 계산
                if check_in <= check_out:
                    days_in_month = (check_out - check_in).days + 1
                else:
                    days_in_month = 0

                print(f"[DEBUG] {student_name} - check_in: {check_in}, check_out: {check_out}, days_in_month: {days_in_month}")

                # 야칭 계산 (30,000엔 / 월)
                rent_per_day = 30000 / total_days_in_month
                rent_amount = rent_per_day * days_in_month
                
                if days_in_month > 0:
                    rent_data.append({
                        "student_id": str(resident.student.id),
                        "student_name": student_name,
                        "room_number": resident.room.room_number if resident.room else "알 수 없음",
                        "building_name": resident.room.building.name if resident.room and resident.room.building else "알 수 없음",
                        "days": days_in_month,
                        "rent_amount": rent_amount
                    })
                    total_rent_amount += rent_amount
                    print(f"[DEBUG] {student_name} 야칭 추가 - days: {days_in_month}, rent_amount: {rent_amount}")
        
        print(f"[DEBUG] 야칭만 있는 결과 반환 - rent_data 개수: {len(rent_data)}, total_rent_amount: {total_rent_amount}")
        return {
            "year": year,
            "month": month,
            "billing_period": f"{prev_year}년 {prev_month}월",
            "total_rooms": 0,
            "rooms": [],
            "rent_data": rent_data,
            "summary": {
                "total_utilities_amount": 0,
                "total_rent_amount": total_rent_amount,
                "grand_total": total_rent_amount
            }
        }

    # 각 방별 공과금 데이터 수집
    print(f"[DEBUG] 공과금이 있는 방 처리 시작 - 총 {len(rooms_with_utilities)}개 방")
    rooms_data = []
    
    for room in rooms_with_utilities:
        print(f"[DEBUG] 방 처리 중: {room.room_number} (빌딩: {room.building.name if room.building else '알 수 없음'})")
        
        # 해당 방의 모든 공과금 조회
        utilities = db.query(RoomUtility).filter(
            RoomUtility.room_id == room.id,
            RoomUtility.charge_month == target_charge_month
        ).all()
        print(f"[DEBUG] 방 {room.room_number}의 공과금 개수: {len(utilities)}")

        # 해당 방의 전월 거주자 조회 (야칭계산용)
        prev_month_residents = db.query(Resident).options(
            joinedload(Resident.student)
        ).filter(
            Resident.room_id == room.id,
            Resident.check_in_date <= end_date,
            (Resident.check_out_date.is_(None) | (Resident.check_out_date >= start_date))
        ).all()
        print(f"[DEBUG] 방 {room.room_number}의 전월 거주자 수: {len(prev_month_residents)}")

        if not prev_month_residents:
            print(f"[DEBUG] 방 {room.room_number}에 전월 거주자가 없음 - 건너뜀")
            continue

        # 각 공과금별 학생별 배분 계산
        utilities_data = []
        print(f"[DEBUG] 방 {room.room_number}의 공과금 처리 시작")
        
        for utility in utilities:
            # 해당 유틸리티 기간 내 방 거주자 전체 쿼리
            all_residents = db.query(Resident).filter(
                Resident.room_id == room.id,
                Resident.check_in_date <= utility.period_end,
                (Resident.check_out_date.is_(None) | (Resident.check_out_date >= utility.period_start))
            ).all()

            # 전체 person-day 계산
            total_person_days = 0
            for r in all_residents:
                overlap_in = max(r.check_in_date, utility.period_start)
                overlap_out = min(r.check_out_date or utility.period_end, utility.period_end)
                overlap_days = (overlap_out - overlap_in).days + 1 if overlap_in <= overlap_out else 0
                total_person_days += overlap_days

            # 1일당 요금 계산 (소수점 버림)
            per_day = int(float(utility.total_amount) / total_person_days) if total_person_days > 0 else 0

            # 각 학생별 부담액 계산
            student_allocations = []
            
            # 모든 학생의 일수와 비율을 먼저 계산
            student_ratios = []
            total_ratio = 0
            
            for resident in prev_month_residents:
                stu_in = max(resident.check_in_date, utility.period_start)
                stu_out = min(resident.check_out_date or utility.period_end, utility.period_end)
                days = (stu_out - stu_in).days + 1 if stu_in <= stu_out else 0
                
                if days > 0:
                    ratio = days / total_person_days
                    total_ratio += ratio
                    student_ratios.append({
                        "resident": resident,
                        "days": days,
                        "ratio": ratio
                    })
            
            # 각 학생별 금액 계산 (1일당 요금 × 일수로 계산)
            total_calculated = 0
            for i, student_info in enumerate(student_ratios):
                if i == len(student_ratios) - 1:
                    # 마지막 학생에게는 남은 금액을 모두 할당하여 총액 보장
                    my_amount = float(utility.total_amount) - total_calculated
                else:
                    # 1일당 요금 × 일수로 계산
                    my_amount = per_day * student_info["days"]
                    total_calculated += my_amount
                
                student_allocations.append({
                    "student_id": str(student_info["resident"].student.id),
                    "student_name": student_info["resident"].student.name,
                    "days": student_info["days"],
                    "amount": my_amount
                })

            utilities_data.append({
                "utility_type": utility.utility_type,
                "period_start": utility.period_start.strftime("%Y-%m-%d"),
                "period_end": utility.period_end.strftime("%Y-%m-%d"),
                "total_amount": float(utility.total_amount),
                "per_day": per_day,
                "total_person_days": total_person_days,
                "student_allocations": student_allocations
            })

        if utilities_data:
            # 야칭 계산 추가
            student_rent_data = {}
            
            # 해당 월의 총 일수 계산
            if month == 12:
                month_end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end_date = date(year, month + 1, 1) - timedelta(days=1)
            
            month_start_date = date(year, month, 1)
            total_days_in_month = (month_end_date - month_start_date).days + 1
            
            # 각 학생별 야칭 계산
            for resident in prev_month_residents:
                student_name = resident.student.name

                # 해당 월에 거주한 일수 계산
                check_in = max(resident.check_in_date, month_start_date)
                
                # 퇴실일이 None이면 해당 월 전체 거주로 인식
                if resident.check_out_date is None:
                    check_in = max(resident.check_in_date, month_start_date)
                    check_out = month_end_date
                else:
                    check_in = resident.check_in_date
                    check_out = min(resident.check_out_date, month_end_date)

                # 해당 월에 거주한 일수 계산
                if check_in <= check_out:
                    days_in_month = (check_out - check_in).days + 1
                else:
                    days_in_month = 0

                # 야칭 계산 (30,000엔 / 월)
                rent_per_day = 30000 / total_days_in_month
                rent_amount = rent_per_day * days_in_month
                
                if days_in_month > 0:
                    student_rent_data[student_name] = {
                        "days": days_in_month,
                        "rent_amount": rent_amount
                    }

            rooms_data.append({
                "room_id": str(room.id),
                "room_number": room.room_number,
                "building_name": room.building.name if room.building else "알 수 없음",
                "building_id": str(room.building.id) if room.building else None,
                "utilities": utilities_data,
                "student_rent_data": student_rent_data,
                "total_utilities_amount": sum(utility["total_amount"] for utility in utilities_data),
                "total_rent_amount": sum(rent_data["rent_amount"] for rent_data in student_rent_data.values())
            })

    return {
        "year": year,
        "month": month,
        "billing_period": f"{prev_year}년 {prev_month}월",
        "total_rooms": len(rooms_data),
        "rooms": rooms_data,
        "summary": {
            "total_utilities_amount": sum(room["total_utilities_amount"] for room in rooms_data),
            "total_rent_amount": sum(room["total_rent_amount"] for room in rooms_data),
            "grand_total": sum(room["total_utilities_amount"] + room["total_rent_amount"] for room in rooms_data)
        }
    }

@app.get("/buildings/download-monthly-invoice/{year}/{month}")
async def download_monthly_invoice(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """해당 년도 월의 모든 방 공과금 PDF 다운로드"""
    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 전달 기준으로 시작/종료일 계산 (예: 7월 조회 시 6월 1일~6월 30일)
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1
    
    start_date = date(prev_year, prev_month, 1)
    if prev_month == 12:
        end_date = date(prev_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(prev_year, prev_month + 1, 1) - timedelta(days=1)

    target_charge_month = date(year, month, 1)

    # 모든 방의 공과금이 있는 방들 조회
    rooms_with_utilities = db.query(Room).options(
        joinedload(Room.building)
    ).join(RoomUtility).filter(
        RoomUtility.charge_month == target_charge_month
    ).distinct().all()

    if not rooms_with_utilities:
        raise HTTPException(
            status_code=404, 
            detail=f"{year}년 {month}월에 청구되는 공과금이 있는 방이 없습니다."
        )

    # 각 방별 공과금 데이터 수집
    rooms_data = []
    
    for room in rooms_with_utilities:
        # 해당 방의 모든 공과금 조회
        utilities = db.query(RoomUtility).filter(
        RoomUtility.room_id == room.id,
            RoomUtility.charge_month == target_charge_month
        ).all()

        # 해당 방의 전월 거주자 조회 (야칭계산용)
        # 전월에 거주했던 사람들 중에서 해당 월에 퇴거하거나 이동하지 않은 사람들
        prev_month_residents = db.query(Resident).options(
            joinedload(Resident.student)
        ).filter(
            Resident.room_id == room.id,
            Resident.check_in_date <= end_date,
            (Resident.check_out_date.is_(None) | (Resident.check_out_date >= start_date))
        ).all()

        # 해당 월에 퇴거하거나 이동하지 않은 사람들 (야칭 30,000엔 청구 대상)
        current_month_residents = db.query(Resident).options(
            joinedload(Resident.student)
        ).filter(
            Resident.room_id == room.id,
            Resident.check_in_date <= end_date,
            (Resident.check_out_date.is_(None) | (Resident.check_out_date >= target_charge_month))
        ).all()

        if not prev_month_residents:
            continue

        # 각 공과금별 학생별 배분 계산
        utilities_data = []
        
        # 디버깅용 로그 추가
        print(f"방 {room.room_number}: prev_month_residents 수 = {len(prev_month_residents)}")
        for resident in prev_month_residents:
            print(f"  - {resident.student.name}: check_in={resident.check_in_date}, check_out={resident.check_out_date}")
            for utility in utilities:
                # 해당 유틸리티 기간 내 방 거주자 전체 쿼리
                all_residents = db.query(Resident).filter(
                Resident.room_id == room.id,
                    Resident.check_in_date <= utility.period_end,
                    (Resident.check_out_date.is_(None) | (Resident.check_out_date >= utility.period_start))
                ).all()

                # 전체 person-day 계산
                total_person_days = 0
                for r in all_residents:
                    overlap_in = max(r.check_in_date, utility.period_start)
                    overlap_out = min(r.check_out_date or utility.period_end, utility.period_end)
                    overlap_days = (overlap_out - overlap_in).days + 1 if overlap_in <= overlap_out else 0
                    total_person_days += overlap_days

            # 1일당 요금 계산 (소수점 버림)
            per_day = int(float(utility.total_amount) / total_person_days) if total_person_days > 0 else 0

            # 각 학생별 부담액 계산
            student_allocations = []
            
            # 모든 학생의 일수와 비율을 먼저 계산
            student_ratios = []
            total_ratio = 0
            
            for resident in prev_month_residents:
                stu_in = max(resident.check_in_date, utility.period_start)
                stu_out = min(resident.check_out_date or utility.period_end, utility.period_end)
                days = (stu_out - stu_in).days + 1 if stu_in <= stu_out else 0
                
                if days > 0:
                    ratio = days / total_person_days
                    total_ratio += ratio
                    student_ratios.append({
                        "resident": resident,
                        "days": days,
                        "ratio": ratio
                    })
            
            # 각 학생별 금액 계산 (1일당 요금 × 일수로 계산)
            total_calculated = 0
            for i, student_info in enumerate(student_ratios):
                if i == len(student_ratios) - 1:
                    # 마지막 학생에게는 남은 금액을 모두 할당하여 총액 보장
                    my_amount = float(utility.total_amount) - total_calculated
                else:
                    # 1일당 요금 × 일수로 계산
                    my_amount = per_day * student_info["days"]
                    total_calculated += my_amount
                
                student_allocations.append({
                    "student_id": str(student_info["resident"].student.id),
                    "student_name": student_info["resident"].student.name,
                    "days": student_info["days"],
                    "amount": my_amount
                })
                
                # 디버깅 로그
                print(f"  공과금 계산: {student_info['resident'].student.name} = {student_info['days']}일, 1일당 요금: {per_day}엔 (소수점 버림), 금액: {my_amount:.2f}엔")

                utilities_data.append({
                  "utility_type": utility.utility_type,
                  "period_start": utility.period_start.strftime("%Y-%m-%d"),
                  "period_end": utility.period_end.strftime("%Y-%m-%d"),
                  "total_amount": float(utility.total_amount),
                  "per_day": per_day,  # 반올림하지 않음
                  "total_person_days": total_person_days,
                  "student_allocations": student_allocations
                })
            
            # 총액 확인 로그
            calculated_total = sum(allocation['amount'] for allocation in student_allocations)
            print(f"  공과금 총액 확인: 원본 {float(utility.total_amount):.2f}엔, 계산된 총액 {calculated_total:.2f}엔, 차이: {float(utility.total_amount) - calculated_total:.2f}엔")

        if utilities_data:
            # 야칭 계산 추가
            student_rent_data = {}
            
            # 해당 월의 총 일수 계산
            if month == 12:
                month_end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end_date = date(year, month + 1, 1) - timedelta(days=1)
            
            month_start_date = date(year, month, 1)
            total_days_in_month = (month_end_date - month_start_date).days + 1
            
            # 각 학생별 야칭 계산
            for resident in prev_month_residents:
                student_name = resident.student.name
                
                # 해당 월에 거주한 일수 계산
                check_in = max(resident.check_in_date, month_start_date)
                
                # 퇴실일이 None이면 해당 월 전체 거주로 인식
                if resident.check_out_date is None:
                    check_in = max(resident.check_in_date, month_start_date)
                    check_out = month_end_date
                else:
                    check_in = resident.check_in_date
                    check_out = min(resident.check_out_date, month_end_date)

                print(f"  체크인: {check_in}, 체크아웃: {check_out}, 원본 체크아웃: {resident.check_out_date}")
                print(f"  원본 체크인: {resident.check_in_date}, 월 시작일: {month_start_date}, 월 종료일: {month_end_date}")
                
                # 퇴실일이 없는 경우: 해당 월 전체 거주로 인식
                if resident.check_out_date is None:
                    resident_days = total_days_in_month
                    print(f"  퇴실일 없음: 해당 월 전체 거주로 인식 ({resident_days}일)")
                else:
                    # 퇴실일이 있는 경우: 실제 거주일수 계산
                    resident_days = (check_out - check_in).days + 1 if check_in <= check_out else 0
                    print(f"  퇴실일 있음: 실제 거주일수 계산 ({resident_days}일)")
                
                print(f"  계산: {resident_days}일 (check_out - check_in = {check_out - check_in})")         
                # 야칭 계산
                if resident_days >= total_days_in_month:
                    # 해당 월에 전부 거주: 30,000엔 고정
                    rent_amount = 30000
                else:
                    # 일부만 거주: 거주일수 × 1,000엔
                    rent_amount = resident_days * 1000
                
                # 와이파이 비용 계산 (한 달 최대 700엔)
                # 비율 기반 계산으로 정확성 향상
                wifi_ratio = resident_days / 30  # 30일 기준 비율
                wifi_amount = min(round(700 * wifi_ratio, 0), 700)  # 최대 700엔, 정수로 반올림
                
                student_rent_data[student_name] = {
                    'days': resident_days,
                    'rent_amount': rent_amount,
                    'wifi_amount': wifi_amount
                }
                print(f"  야칭 계산: {student_name} = {resident_days}일, {rent_amount}엔, 와이파이: {wifi_amount}엔 (계산: 700 × {wifi_ratio:.3f} = {round(700 * wifi_ratio, 0)})")
            
            # utilities_data에 야칭 및 와이파이 정보 추가
            for utility_data in utilities_data:
                for allocation in utility_data['student_allocations']:
                    student_name = allocation['student_name']
                    if student_name in student_rent_data:
                        allocation['rent_amount'] = student_rent_data[student_name]['rent_amount']
                        allocation['rent_days'] = student_rent_data[student_name]['days']
                        allocation['wifi_amount'] = student_rent_data[student_name]['wifi_amount']
            
            rooms_data.append({
                "room_id": str(room.id),
                "room_number": room.room_number,
                "building_name": room.building.name if room.building else None,
            "utilities": utilities_data,
                "student_rent_data": student_rent_data
            })

    # HTML 템플릿 생성
    html_content = generate_utilities_html(rooms_data, year, month, prev_year, prev_month)
    
    # PDF 생성
    pdf_file = HTML(string=html_content).write_pdf()
    
    # 임시 파일로 저장 후 다운로드
    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
        tmp_file.write(pdf_file)
        tmp_file_path = tmp_file.name
    
    return FileResponse(
        tmp_file_path,
        media_type='application/pdf',
        filename=f'光熱費_{year}年_{month}月.pdf'
    )

def generate_utilities_html(rooms_data, year, month, prev_year, prev_month):
    """공과금 HTML 템플릿 생성"""
    from datetime import datetime
    
    # 각 방별로 학생별 부담금액 계산
    processed_rooms = []
    
    for room in rooms_data:
        # 학생별 부담금액 계산 (공과금 + 월세)
        student_totals = {}
        
        # 월세 계산 (한 달 30,000엔, 하루 1,000엔)
        daily_rent = 1000     # 하루 1,000엔
        
        for utility in room['utilities']:
            for allocation in utility['student_allocations']:
                student_name = allocation['student_name']
                if student_name not in student_totals:
                    student_totals[student_name] = {
                        'rent': 0,        # 월세
                        'electricity': 0,  # 전기
                        'water': 0,       # 수도
                        'gas': 0,         # 가스
                        'wifi': 0,        # 와이파이
                        'total': 0        # 합계 (마지막에 계산)
                    }
                
                amount = allocation['amount']
                
                if utility['utility_type'] == 'electricity':
                    student_totals[student_name]['electricity'] = round(amount, 0)
                elif utility['utility_type'] == 'water':
                    student_totals[student_name]['water'] = round(amount, 0)
                elif utility['utility_type'] == 'gas':
                    student_totals[student_name]['gas'] = round(amount, 0)
        
        # 야칭 계산 및 추가
        # student_rent_data에 있는 모든 학생에 대해 야칭 계산
        print(f"방 {room['room_number']}: student_rent_data 확인")
        if 'student_rent_data' in room:
            print(f"  student_rent_data 키: {list(room['student_rent_data'].keys())}")
            for student_name, rent_info in room['student_rent_data'].items():
                print(f"  {student_name}: {rent_info}")
        else:
            print("  student_rent_data가 없습니다!")
        
        if 'student_rent_data' in room:
            for student_name, rent_info in room['student_rent_data'].items():
                # 해당 월의 총 일수 계산
                if month == 12:
                    month_end_date = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    month_end_date = date(year, month + 1, 1) - timedelta(days=1)
                
                month_start_date = date(year, month, 1)
                total_days_in_month = (month_end_date - month_start_date).days + 1
                
                # 야칭 계산
                student_rent_days = rent_info['days']
                if student_rent_days >= total_days_in_month:
                    # 해당 월에 전부 거주: 30,000엔 고정
                    rent_amount = 30000
                else:
                    # 일부만 거주: 거주일수 × 1,000엔
                    rent_amount = student_rent_days * 1000
                
                # student_totals에 해당 학생이 없으면 추가
                if student_name not in student_totals:
                    student_totals[student_name] = {
                        'rent': 0,        # 월세
                        'electricity': 0,  # 전기
                        'water': 0,       # 수도
                        'gas': 0,         # 가스
                        'wifi': 0,        # 와이파이
                        'total': 0        # 합계
                    }
                
                student_totals[student_name]['rent'] = round(rent_amount, 0)
                
                # 와이파이 금액 추가
                if 'wifi_amount' in rent_info:
                    student_totals[student_name]['wifi'] = round(rent_info['wifi_amount'], 0)
                    print(f"  {student_name} 와이파이 추가: {round(rent_info['wifi_amount'], 0)}엔")
        
        # 모든 계산이 끝난 후 total 계산
        for student_name in student_totals:
            total = (student_totals[student_name]['rent'] + 
                    student_totals[student_name]['electricity'] + 
                    student_totals[student_name]['water'] + 
                    student_totals[student_name]['gas'] + 
                    student_totals[student_name]['wifi'])
            student_totals[student_name]['total'] = round(total, 0)
            print(f"  {student_name} 최종 합계: {round(total, 0)}엔 (야칭: {student_totals[student_name]['rent']}, 전기: {student_totals[student_name]['electricity']}, 수도: {student_totals[student_name]['water']}, 가스: {student_totals[student_name]['gas']}, 와이파이: {student_totals[student_name]['wifi']})")
        
        # 방 정보와 학생별 부담금액을 함께 저장
        processed_room = {
            'building_name': room['building_name'],
            'room_number': room['room_number'],
            'student_totals': student_totals
        }
        processed_rooms.append(processed_room)
    
    # Jinja2 템플릿 렌더링
    template = templates.get_template("monthly_utilities.html")
    html_content = template.render(
        year=year,
        month=month,
        prev_year=prev_year,
        prev_month=prev_month,
        created_date=datetime.now().strftime('%Y年%m月%d日'),
        rooms_data=processed_rooms
    )
    
    return html_content

@app.get("/validate-monthly-utilities/{year}/{month}")
def validate_monthly_utilities(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """해당 년도 월의 모든 방 공과금 입력 상태 검증"""
    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 해당 월의 시작일과 종료일 계산
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    target_charge_month = date(year, month, 1)

    # 해당 월에 거주자가 있는 모든 방 조회
    rooms_with_residents = db.query(Room).options(
        joinedload(Room.building)
    ).join(Resident).filter(
        Resident.check_in_date <= end_date,
        (Resident.check_out_date.is_(None) | (Resident.check_out_date >= start_date))
    ).distinct().all()
  
    if not rooms_with_residents:
      return {
        "is_valid": False,
        "message": f"{year}年{month}月に居住者がいる部屋がありません。",
        "missing_rooms": [],
        "total_rooms": 0,
        "valid_rooms": 0,
        "missing_rooms_count": 0
      }

    # 각 방별 공과금 입력 상태 확인
    valid_rooms = []
    missing_rooms = []
    required_utility_types = ['electricity', 'water', 'gas']  # 필수 공과금 유형

    for room in rooms_with_residents:
        # 해당 방의 공과금 조회
        utilities = db.query(RoomUtility).filter(
            RoomUtility.room_id == room.id,
            RoomUtility.charge_month == target_charge_month
        ).all()

        # 입력된 공과금 유형 확인
        input_utility_types = [utility.utility_type for utility in utilities]
        
        # 누락된 공과금 유형 확인
        missing_utility_types = [ut for ut in required_utility_types if ut not in input_utility_types]
        
        room_info = {
            "room_id": str(room.id),
            "room_number": room.room_number,
            "building_name": room.building.name if room.building else None,
            "building_address": room.building.address if room.building else None,
            "input_utilities": input_utility_types,
            "missing_utilities": missing_utility_types,
            "total_residents": db.query(Resident).filter(
                Resident.room_id == room.id,
                Resident.check_in_date <= end_date,
                (Resident.check_out_date.is_(None) | (Resident.check_out_date >= start_date))
            ).count()
        }

        if missing_utility_types:
            missing_rooms.append(room_info)
        else:
            valid_rooms.append(room_info)

    # 검증 결과
    is_valid = len(missing_rooms) == 0
    
    return {
        "is_valid": is_valid,
        "message": f"{year}年{month}月光熱費入力状態: {'完了' if is_valid else '未完了'}",
        "missing_rooms": missing_rooms,
        "valid_rooms": valid_rooms,
        "total_rooms": len(rooms_with_residents),
        "valid_rooms_count": len(valid_rooms),
        "missing_rooms_count": len(missing_rooms),
        "required_utility_types": required_utility_types,
        "year": year,
        "month": month
    }

@app.get("/buildings/download-monthly-invoice/validate/{year}/{month}")
def validate_monthly_invoice(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """월별 인보이스 다운로드 전 검증 - 공과금 입력 상태 확인"""
    # 월 유효성 검사
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월은 1-12 사이의 값이어야 합니다")

    # 해당 월의 시작일과 종료일 계산
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    target_charge_month = date(year, month, 1)

    # 해당 월에 거주자가 있는 모든 방 조회
    rooms_with_residents = db.query(Room).options(
        joinedload(Room.building)
    ).join(Resident).filter(
        Resident.check_in_date <= end_date,
        (Resident.check_out_date.is_(None) | (Resident.check_out_date >= start_date))
    ).distinct().all()

    if not rooms_with_residents:
        return {
            "is_valid": False,
            "can_download": False,
            "message": f"{year}年{month}月に居住者がいる部屋がありません。",
            "missing_rooms": [],
            "total_rooms": 0,
            "valid_rooms": 0,
            "missing_rooms_count": 0,
            "year": year,
            "month": month
        }

    # 각 방별 공과금 입력 상태 확인
    valid_rooms = []
    missing_rooms = []
    required_utility_types = ['electricity', 'water', 'gas']  # 필수 공과금 유형

    for room in rooms_with_residents:
        # 해당 방의 공과금 조회
        utilities = db.query(RoomUtility).filter(
            RoomUtility.room_id == room.id,
            RoomUtility.charge_month == target_charge_month
        ).all()

        # 입력된 공과금 유형 확인
        input_utility_types = [utility.utility_type for utility in utilities]
        
        # 누락된 공과금 유형 확인
        missing_utility_types = [ut for ut in required_utility_types if ut not in input_utility_types]
        
        # 해당 방의 거주자 정보 조회
        residents = db.query(Resident).options(
            joinedload(Resident.student)
        ).filter(
            Resident.room_id == room.id,
            Resident.check_in_date <= end_date,
            (Resident.check_out_date.is_(None) | (Resident.check_out_date >= start_date))
        ).all()
        
        resident_names = [resident.student.name for resident in residents]
        
        room_info = {
            "room_id": str(room.id),
            "room_number": room.room_number,
            "building_name": room.building.name if room.building else None,
            "building_address": room.building.address if room.building else None,
            "input_utilities": input_utility_types,
            "missing_utilities": missing_utility_types,
            "total_residents": len(residents),
            "resident_names": resident_names,
            "missing_utility_names": {
                "electricity": "電気" if "electricity" in missing_utility_types else None,
                "water": "水道" if "water" in missing_utility_types else None,
                "gas": "ガス" if "gas" in missing_utility_types else None
            }
        }

        if missing_utility_types:
            missing_rooms.append(room_info)
        else:
            valid_rooms.append(room_info)

    # 검증 결과
    is_valid = len(missing_rooms) == 0
    can_download = is_valid
    
    return {
        "is_valid": is_valid,
        "can_download": can_download,
        "message": f"{year}年{month}月光熱費入力状態: {'完了' if is_valid else '未完了'}",
        "missing_rooms": missing_rooms,
        "valid_rooms": valid_rooms,
        "total_rooms": len(rooms_with_residents),
        "valid_rooms_count": len(valid_rooms),
        "missing_rooms_count": len(missing_rooms),
        "required_utility_types": required_utility_types,
        "year": year,
        "month": month,
        "download_url": f"/buildings/download-monthly-invoice/{year}/{month}" if can_download else None
    }

@app.get("/elderly")
def get_elderly(
    name: Optional[str] = Query(None, description="고령자 이름으로 검색"),
    name_katakana: Optional[str] = Query(None, description="고령자 이름 카타카나로 검색"),
    gender: Optional[str] = Query(None, description="성별로 검색"),
    care_level: Optional[str] = Query(None, description="요양 등급으로 검색"),
    status: Optional[str] = Query(None, description="상태로 검색"),
    building_id: Optional[str] = Query(None, description="건물 아이디로 검색"),
    room_number: Optional[str] = Query(None, description="방 번호로 검색"),
    sort_by: Optional[str] = Query(None, description="정렬 필드 (name 또는 current_room.room_number)"),
    sort_desc: Optional[bool] = Query(False, description="내림차순 정렬 여부 (true: 내림차순, false: 오름차순)"),
    page: int = Query(1, description="페이지 번호", ge=1),
    page_size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    고령자 목록을 조회합니다.
    """
    # 기본 쿼리 생성
    query = db.query(Elderly).options(
        joinedload(Elderly.current_room).joinedload(Room.building),
        joinedload(Elderly.category),
        joinedload(Elderly.hospitalizations)
    )
    
    # 검색 조건 적용
    if name:
        query = query.filter(Elderly.name.ilike(f"%{name}%"))
    
    if name_katakana:
        query = query.filter(Elderly.name_katakana.ilike(f"%{name_katakana}%"))
    
    if gender:
        query = query.filter(Elderly.gender == gender)
    
    if care_level:
        query = query.filter(Elderly.care_level == care_level)
    
    if status:
        query = query.filter(Elderly.status == status)
    
    if building_id:
        query = query.join(Room, Elderly.current_room_id == Room.id).join(Building, Room.building_id == Building.id)
        query = query.filter(Building.id == building_id)
    
    if room_number:
        query = query.join(Room, Elderly.current_room_id == Room.id)
        query = query.filter(Room.room_number.ilike(f"%{room_number}%"))
    
    # 정렬 적용
    if sort_by:
        if sort_by == "name":
            if sort_desc:
                query = query.order_by(Elderly.name.desc())
            else:
                query = query.order_by(Elderly.name.asc())
        elif sort_by == "current_room.room_number":
            # Room 테이블과 조인하여 방번호로 정렬
            query = query.outerjoin(Room, Elderly.current_room_id == Room.id)
            if sort_desc:
                query = query.order_by(Room.room_number.desc())
            else:
                query = query.order_by(Room.room_number.asc())
    else:
        # 기본 정렬: 생성일 기준 내림차순
        query = query.order_by(Elderly.created_at.desc())
    
    # 전체 개수 계산
    total_count = query.count()
    
    # 페이지네이션 적용
    offset = (page - 1) * page_size
    elderly_list = query.offset(offset).limit(page_size).all()
    
    # 응답 데이터 구성
    elderly_data = []
    for elderly in elderly_list:
        # 입원 상태 확인
        hospitalization_status = "正常"
        latest_hospitalization = None
        
        if elderly.hospitalizations:
            # 가장 최근 입원 기록 찾기
            latest_hospitalization = max(elderly.hospitalizations, key=lambda x: x.date)
            
            # 입원 기록이 있고 퇴원 기록이 없으면 입원중
            admission_records = [h for h in elderly.hospitalizations if h.hospitalization_type == 'admission']
            discharge_records = [h for h in elderly.hospitalizations if h.hospitalization_type == 'discharge']
            
            if admission_records and not discharge_records:
                # 입원 기록만 있고 퇴원 기록이 없으면 입원중
                hospitalization_status = "入院中"
            elif admission_records and discharge_records:
                # 입원과 퇴원 기록이 모두 있으면 최신 기록 확인
                latest_admission = max(admission_records, key=lambda x: x.date)
                latest_discharge = max(discharge_records, key=lambda x: x.date)
                
                if latest_admission.date > latest_discharge.date:
                    hospitalization_status = "入院中"
                else:
                    hospitalization_status = "正常"
        
        elderly_dict = {
        "id": str(elderly.id),
        "name": elderly.name,
            "email": elderly.email,
            "created_at": elderly.created_at,
            "phone": elderly.phone,
            "avatar": elderly.avatar,
        "name_katakana": elderly.name_katakana,
        "gender": elderly.gender,
            "birth_date": elderly.birth_date,
            "status": elderly.status,
            "current_room_id": str(elderly.current_room_id) if elderly.current_room_id else None,
            "care_level": elderly.care_level,
            "hospitalization_status": hospitalization_status,
            "latest_hospitalization": {
                "id": str(latest_hospitalization.id),
                "elderly_id": str(latest_hospitalization.elderly_id),
                "hospitalization_type": latest_hospitalization.hospitalization_type,
                "hospital_name": latest_hospitalization.hospital_name,
                "date": latest_hospitalization.date,
                "last_meal_date": latest_hospitalization.last_meal_date,
                "last_meal_type": latest_hospitalization.last_meal_type,
                "meal_resume_date": latest_hospitalization.meal_resume_date,
                "meal_resume_type": latest_hospitalization.meal_resume_type,
                "note": latest_hospitalization.note
            } if latest_hospitalization else None,
            "current_room": None
        }
        
        # 현재 방 정보 추가
        if elderly.current_room:
            elderly_dict["current_room"] = {
                "id": str(elderly.current_room.id),
                "room_number": elderly.current_room.room_number,
                "floor": elderly.current_room.floor,
                "capacity": elderly.current_room.capacity,
                "rent": elderly.current_room.rent,
                "maintenance": elderly.current_room.maintenance,
                "service": elderly.current_room.service,
                "note": elderly.current_room.note,
                "building": {
                    "id": str(elderly.current_room.building.id),
                    "name": elderly.current_room.building.name,
                    "address": elderly.current_room.building.address,
                    "resident_type": elderly.current_room.building.resident_type
                } if elderly.current_room.building else None
            }
        
        elderly_data.append(elderly_dict)
    
    return {
        "items": elderly_data,
        "total": total_count,
        "total_pages": (total_count + page_size - 1) // page_size
    }

@app.get("/care-items", response_model=List[CareItemResponse])
def get_care_items(
    is_active: Optional[bool] = Query(None, description="활성화 여부로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(CareItem)
    if is_active is not None:
        query = query.filter(CareItem.is_active == is_active)
    items = query.order_by(CareItem.created_at.desc()).all()
    return items


@app.get("/care-meal-prices", response_model=List[CareMealPriceResponse])
def get_care_meal_prices(
    meal_type: Optional[str] = Query(None, description="식사 유형으로 필터링 (breakfast, lunch, dinner)"),
    is_active: Optional[bool] = Query(None, description="활성화 여부로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 식사 가격 목록을 조회합니다.
    """
    query = db.query(CareMealPrice)
    
    if meal_type:
        query = query.filter(CareMealPrice.meal_type == meal_type)
    
    if is_active is not None:
        query = query.filter(CareMealPrice.is_active == is_active)
    
    items = query.order_by(CareMealPrice.created_at.desc()).all()
    return items


@app.post("/care-meal-prices", response_model=CareMealPriceResponse)
def create_care_meal_price(
    meal_price: CareMealPriceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 식사 가격을 생성합니다.
    """
    # 동일한 meal_type이 이미 존재하는지 확인
    existing_price = db.query(CareMealPrice).filter(
        CareMealPrice.meal_type == meal_price.meal_type,
        CareMealPrice.is_active == True
    ).first()
    
    if existing_price:
        raise HTTPException(
            status_code=400,
            detail=f"이미 활성화된 {meal_price.meal_type} 식사 가격이 존재합니다."
        )
    
    db_meal_price = CareMealPrice(
        meal_type=meal_price.meal_type,
        price=meal_price.price,
        is_active=meal_price.is_active
    )
    
    db.add(db_meal_price)
    db.commit()
    db.refresh(db_meal_price)
    
    return db_meal_price


@app.put("/care-meal-prices/{price_id}", response_model=CareMealPriceResponse)
def update_care_meal_price(
    price_id: str,
    meal_price_update: CareMealPriceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 식사 가격을 수정합니다.
    """
    db_meal_price = db.query(CareMealPrice).filter(CareMealPrice.id == price_id).first()
    
    if not db_meal_price:
        raise HTTPException(status_code=404, detail="케어 식사 가격을 찾을 수 없습니다.")
    
    update_data = meal_price_update.dict(exclude_unset=True)
    
    for field, value in update_data.items():
        setattr(db_meal_price, field, value)
    
    db.commit()
    db.refresh(db_meal_price)
    
    return db_meal_price


@app.delete("/care-meal-prices/{price_id}")
def delete_care_meal_price(
    price_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 식사 가격을 삭제합니다.
    """
    db_meal_price = db.query(CareMealPrice).filter(CareMealPrice.id == price_id).first()
    
    if not db_meal_price:
        raise HTTPException(status_code=404, detail="케어 식사 가격을 찾을 수 없습니다.")
    
    db.delete(db_meal_price)
    db.commit()
    
    return {"message": "케어 식사 가격이 성공적으로 삭제되었습니다."}


@app.get("/care-utility-prices", response_model=List[CareUtilityPriceResponse])
def get_care_utility_prices(
    utility_type: Optional[str] = Query(None, description="공과금 유형으로 필터링 (electricity, water, gas)"),
    is_active: Optional[bool] = Query(None, description="활성화 여부로 필터링"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 공과금 가격 목록을 조회합니다.
    """
    query = db.query(CareUtilityPrice)
    
    if utility_type:
        query = query.filter(CareUtilityPrice.utility_type == utility_type)
    
    if is_active is not None:
        query = query.filter(CareUtilityPrice.is_active == is_active)
    
    items = query.order_by(CareUtilityPrice.created_at.desc()).all()
    return items


@app.post("/care-utility-prices", response_model=CareUtilityPriceResponse)
def create_care_utility_price(
    utility_price: CareUtilityPriceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 공과금 가격을 생성합니다.
    """
    # 동일한 utility_type이 이미 존재하는지 확인
    existing_price = db.query(CareUtilityPrice).filter(
        CareUtilityPrice.utility_type == utility_price.utility_type,
        CareUtilityPrice.is_active == True
    ).first()
    
    if existing_price:
        raise HTTPException(
            status_code=400,
            detail=f"이미 활성화된 {utility_price.utility_type} 공과금 가격이 존재합니다."
        )
    
    db_utility_price = CareUtilityPrice(
        utility_type=utility_price.utility_type,
        price_per_unit=utility_price.price_per_unit,
        unit=utility_price.unit,
        is_active=utility_price.is_active
    )
    
    db.add(db_utility_price)
    db.commit()
    db.refresh(db_utility_price)
    
    return db_utility_price


@app.put("/care-utility-prices/{price_id}", response_model=CareUtilityPriceResponse)
def update_care_utility_price(
    price_id: str,
    utility_price_update: CareUtilityPriceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 공과금 가격을 수정합니다.
    """
    db_utility_price = db.query(CareUtilityPrice).filter(CareUtilityPrice.id == price_id).first()
    
    if not db_utility_price:
        raise HTTPException(status_code=404, detail="케어 공과금 가격을 찾을 수 없습니다.")
    
    update_data = utility_price_update.dict(exclude_unset=True)
    
    for field, value in update_data.items():
        setattr(db_utility_price, field, value)
    
    db.commit()
    db.refresh(db_utility_price)
    
    return db_utility_price


@app.delete("/care-utility-prices/{price_id}")
def delete_care_utility_price(
    price_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    케어 공과금 가격을 삭제합니다.
    """
    db_utility_price = db.query(CareUtilityPrice).filter(CareUtilityPrice.id == price_id).first()
    
    if not db_utility_price:
        raise HTTPException(status_code=404, detail="케어 공과금 가격을 찾을 수 없습니다.")
    
    db.delete(db_utility_price)
    db.commit()
    
    return {"message": "케어 공과금 가격이 성공적으로 삭제되었습니다."} 


@app.get("/test")
def test(
    name: Optional[str] = Query(None, description="학생 이름으로 검색"),
    name_katakana: Optional[str] = Query(None, description="학생 이름 카타카나로 검색"),
    nationality: Optional[str] = Query(None, description="학생 국적으로 검색"),
    company: Optional[str] = Query(None, description="회사 이름으로 검색"),
    consultant: Optional[int] = Query(None, description="컨설턴트 번호 이상으로 검색"),
    email: Optional[str] = Query(None, description="이메일로 검색"),
    student_type: Optional[str] = Query(None, description="학생 유형으로 검색"),
    building_name: Optional[str] = Query(None, description="건물 이름으로 검색"),
    room_number: Optional[str] = Query(None, description="방 번호로 검색"),
    status: Optional[str] = Query(None, description="상태로 검색"),
    page: int = Query(1, description="페이지 번호", ge=1),
    page_size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    start = time.time()

    print(f"[TEST] test API 시작 - page: {page}, page_size: {page_size}")
    
    # get_students 함수의 로직을 그대로 실행
    # 기본 쿼리 생성
    query = db.query(Student).options(
        joinedload(Student.company),
        joinedload(Student.grade),
        joinedload(Student.current_room).joinedload(Room.building)
    ).outerjoin(Company).outerjoin(Grade)

    # 방 관련 필터링이 있는 경우 Room과 Building 테이블과 조인
    if building_name or room_number:
        query = query.outerjoin(Room, Student.current_room_id == Room.id).outerjoin(Building, Room.building_id == Building.id)

    # 필터 조건 추가
    if student_type and student_type != 'ALL':
        query = query.filter(Student.student_type.ilike(f"%{student_type}%"))
    if name:
        query = query.filter(Student.name.ilike(f"%{name}%"))
    if nationality:
        query = query.filter(Student.nationality.ilike(f"%{nationality}%"))
    if name_katakana:
        query = query.filter(Student.name_katakana.ilike(f"%{name_katakana}%"))
    if company:
        query = query.filter(Company.name.ilike(f"%{company}%"))
    if consultant is not None:
        query = query.filter(Student.consultant >= consultant)
    if email:
        query = query.filter(Student.email.ilike(f"%{email}%"))
    if building_name:
        query = query.filter(Building.name.ilike(f"%{building_name}%"))
    if room_number:
        query = query.filter(Room.room_number.ilike(f"%{room_number}%"))
    if status:
        query = query.filter(Student.status == status)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    students = query.offset((page - 1) * page_size).limit(page_size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + page_size - 1) // page_size

    # 응답 데이터 준비
    result = []
    for student in students:
        student_data = {
            "id": str(student.id),
            "name": student.name,
            "email": student.email if student.email else "",
            "created_at": student.created_at,
            "company_id": str(student.company_id) if student.company_id else None,
            "consultant": student.consultant,
            "phone": student.phone,
            "avatar": student.avatar,
            "grade_id": str(student.grade_id) if student.grade_id else None,
            "cooperation_submitted_date": student.cooperation_submitted_date,
            "cooperation_submitted_place": student.cooperation_submitted_place,
            "assignment_date": student.assignment_date,
            "ward": student.ward,
            "name_katakana": student.name_katakana,
            "gender": student.gender,
            "birth_date": student.birth_date,
            "nationality": student.nationality,
            "has_spouse": student.has_spouse,
            "japanese_level": student.japanese_level,
            "passport_number": student.passport_number,
            "residence_card_number": student.residence_card_number,
            "residence_card_start": student.residence_card_start,
            "residence_card_expiry": student.residence_card_expiry,
            "resignation_date": student.resignation_date,
            "local_address": student.local_address,
            "address": student.address,
            "experience_over_2_years": student.experience_over_2_years,
            "status": student.status,
            "arrival_type": student.arrival_type,
            "company": {
                "name": student.company.name
            } if student.company else None,
            "grade": {
                "name": student.grade.name
            } if student.grade else None,
            "current_room": {
                "room_number": student.current_room.room_number,
                "building_name": student.current_room.building.name
            } if student.current_room and student.current_room.building else None
        }
        result.append(student_data)

    processing_time = (time.time() - start) * 1000
    
    print(f"[TEST] test API 완료 - 처리 시간: {processing_time:.2f}ms, 총 학생 수: {total_count}, 현재 페이지 학생 수: {len(result)}")
    
    return {
        "processing_time_ms": processing_time,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "page_size": page_size,
        "students": result,
        "filters_applied": {
            "name": name,
            "name_katakana": name_katakana,
            "nationality": nationality,
            "company": company,
            "consultant": consultant,
            "email": email,
            "student_type": student_type,
            "building_name": building_name,
            "room_number": room_number,
            "status": status
        }
    }


@app.get("/test/students")
def test_get_students(
    name: Optional[str] = Query(None, description="학생 이름으로 검색"),
    name_katakana: Optional[str] = Query(None, description="학생 이름 카타카나로 검색"),
    nationality: Optional[str] = Query(None, description="학생 국적으로 검색"),
    company: Optional[str] = Query(None, description="회사 이름으로 검색"),
    consultant: Optional[int] = Query(None, description="컨설턴트 번호 이상으로 검색"),
    email: Optional[str] = Query(None, description="이메일로 검색"),
    student_type: Optional[str] = Query(None, description="학생 유형으로 검색"),
    building_name: Optional[str] = Query(None, description="건물 이름으로 검색"),
    room_number: Optional[str] = Query(None, description="방 번호로 검색"),
    status: Optional[str] = Query(None, description="상태로 검색"),
    page: int = Query(1, description="페이지 번호", ge=1),
    page_size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """테스트용 학생 목록 조회 API - get_students 함수를 내부적으로 호출"""
    start = time.time()
    
    print(f"[TEST] test_get_students 시작 - page: {page}, page_size: {page_size}")
    
    # get_students 함수의 로직을 그대로 실행
    # 기본 쿼리 생성
    query = db.query(Student).options(
        joinedload(Student.company),
        joinedload(Student.grade),
        joinedload(Student.current_room).joinedload(Room.building)
    ).outerjoin(Company).outerjoin(Grade)

    # 방 관련 필터링이 있는 경우 Room과 Building 테이블과 조인
    if building_name or room_number:
        query = query.outerjoin(Room, Student.current_room_id == Room.id).outerjoin(Building, Room.building_id == Building.id)

    # 필터 조건 추가
    if student_type and student_type != 'ALL':
        query = query.filter(Student.student_type.ilike(f"%{student_type}%"))
    if name:
        query = query.filter(Student.name.ilike(f"%{name}%"))
    if nationality:
        query = query.filter(Student.nationality.ilike(f"%{nationality}%"))
    if name_katakana:
        query = query.filter(Student.name_katakana.ilike(f"%{name_katakana}%"))
    if company:
        query = query.filter(Company.name.ilike(f"%{company}%"))
    if consultant is not None:
        query = query.filter(Student.consultant >= consultant)
    if email:
        query = query.filter(Student.email.ilike(f"%{email}%"))
    if building_name:
        query = query.filter(Building.name.ilike(f"%{building_name}%"))
    if room_number:
        query = query.filter(Room.room_number.ilike(f"%{room_number}%"))
    if status:
        query = query.filter(Student.status == status)

    # 전체 항목 수 계산
    total_count = query.count()

    # 페이지네이션 적용
    students = query.offset((page - 1) * page_size).limit(page_size).all()

    # 전체 페이지 수 계산
    total_pages = (total_count + page_size - 1) // page_size

    # 응답 데이터 준비
    result = []
    for student in students:
        student_data = {
            "id": str(student.id),
            "name": student.name,
            "email": student.email if student.email else "",
            "created_at": student.created_at,
            "company_id": str(student.company_id) if student.company_id else None,
            "consultant": student.consultant,
            "phone": student.phone,
            "avatar": student.avatar,
            "grade_id": str(student.grade_id) if student.grade_id else None,
            "cooperation_submitted_date": student.cooperation_submitted_date,
            "cooperation_submitted_place": student.cooperation_submitted_place,
            "assignment_date": student.assignment_date,
            "ward": student.ward,
            "name_katakana": student.name_katakana,
            "gender": student.gender,
            "birth_date": student.birth_date,
            "nationality": student.nationality,
            "has_spouse": student.has_spouse,
            "japanese_level": student.japanese_level,
            "passport_number": student.passport_number,
            "residence_card_number": student.residence_card_number,
            "residence_card_start": student.residence_card_start,
            "residence_card_expiry": student.residence_card_expiry,
            "resignation_date": student.resignation_date,
            "local_address": student.local_address,
            "address": student.address,
            "experience_over_2_years": student.experience_over_2_years,
            "status": student.status,
            "arrival_type": student.arrival_type,
            "company": {
                "name": student.company.name
            } if student.company else None,
            "grade": {
                "name": student.grade.name
            } if student.grade else None,
            "current_room": {
                "room_number": student.current_room.room_number,
                "building_name": student.current_room.building.name
            } if student.current_room and student.current_room.building else None
        }
        result.append(student_data)

    processing_time = (time.time() - start) * 1000
    
    print(f"[TEST] test_get_students 완료 - 처리 시간: {processing_time:.2f}ms, 총 학생 수: {total_count}, 현재 페이지 학생 수: {len(result)}")
    
    return {
        "processing_time_ms": processing_time,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "page_size": page_size,
        "students": result,
        "filters_applied": {
            "name": name,
            "name_katakana": name_katakana,
            "nationality": nationality,
            "company": company,
            "consultant": consultant,
            "email": email,
            "student_type": student_type,
            "building_name": building_name,
            "room_number": room_number,
            "status": status
        }
    }


# Monthly Items API
@app.get("/students/{student_id}/monthly-items")
def get_student_monthly_items(
    student_id: str,
    year: Optional[int] = Query(None, description="년도로 필터링"),
    month: Optional[int] = Query(None, description="월로 필터링 (1-12)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """학생의 월별 관리비 항목 조회"""
    try:
        # 학생 존재 여부 확인
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="学生が見つかりません")
        
        # 기본 쿼리 생성
        query = db.query(BillingMonthlyItem).filter(BillingMonthlyItem.student_id == student_id)
        
        # 필터링
        if year is not None:
            query = query.filter(BillingMonthlyItem.year == year)
        if month is not None:
            query = query.filter(BillingMonthlyItem.month == month)
        
        # 정렬 (생성일 기준 내림차순, 년도, 월, 정렬 순서)
        items = query.order_by(BillingMonthlyItem.created_at.asc(), BillingMonthlyItem.year, BillingMonthlyItem.month, BillingMonthlyItem.sort_order).order_by(BillingMonthlyItem.sort_order).all()
        
        # 항목 이름별로 그룹화
        grouped_items = {}
        for item in items:
            if item.item_name not in grouped_items:
                grouped_items[item.item_name] = {
                    "item_name": item.item_name,
                    "memo": item.memo,
                    "sort_order": item.sort_order,  # sort_order 추가
                    "months": {}
                }
            
            # 1~12월 데이터 초기화 (없는 월은 0으로 설정)
            if len(grouped_items[item.item_name]["months"]) == 0:
                for m in range(1, 13):
                    grouped_items[item.item_name]["months"][m] = {
                        "month": m,
                        "amount": 0,
                        "memo": None,
                        "item_id": None
                    }
            
            # 실제 데이터로 업데이트
            grouped_items[item.item_name]["months"][item.month] = {
                "month": item.month,
                "amount": float(item.amount),
                "memo": item.memo,
                "item_id": str(item.id)
            }
        
        # 그룹화된 데이터를 리스트로 변환
        result_items = []
        for item_name, item_data in grouped_items.items():
            months_list = []
            for month_num in range(1, 13):
                months_list.append(item_data["months"][month_num])
            
            result_items.append({
                "item_name": item_name,
                "memo": item_data["memo"],
                "sort_order": item_data["sort_order"],  # sort_order 추가
                "months": months_list
            })
        
        return {
            "student_id": student_id,
            "student_name": student.name,
            "items": result_items
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"月別管理費項目の取得中にエラーが発生しました: {str(e)}")

@app.post("/students/{student_id}/monthly-items")
def create_student_monthly_items(
    student_id: str,
    item_data: BillingMonthlyItemCreate,
    year: Optional[int] = Query(None, description="생성할 연도 (기본값: 현재 연도)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """학생의 월별 관리비 항목 생성 (1~12월 전체 생성)"""
    try:
        # 학생 존재 여부 확인
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="学生が見つかりません")
        
        # 연도 설정 (기본값: 현재 연도)
        target_year = year if year is not None else datetime.now().year
        
        # 기존 항목이 있는지 확인
        existing_items = db.query(BillingMonthlyItem).filter(
            BillingMonthlyItem.student_id == student_id,
            BillingMonthlyItem.item_name == item_data.item_name,
            BillingMonthlyItem.year == target_year
        ).first()
        
        if existing_items:
            raise HTTPException(status_code=400, detail=f"{target_year}年に既に同じ項目名が存在します。")
        
        # 1~12월까지 항목 생성
        items_to_create = []
        for month in range(1, 13):  # 1~12월
            new_item = BillingMonthlyItem(
                student_id=student_id,
                year=target_year,
                month=month,
                item_name=item_data.item_name,
                amount=0,
                memo=item_data.memo,
                sort_order=0
            )
            items_to_create.append(new_item)
        
        # 데이터베이스에 일괄 삽입
        db.add_all(items_to_create)
        db.commit()
        
        # 데이터베이스 로그 생성 (각 항목별로)
        try:
            for item in items_to_create:
                create_database_log(
                    db=db,
                    table_name="billing_monthly_items",
                    record_id=str(item.id),
                    action="CREATE",
                    user_id=current_user.get("id") if current_user else None,
                    new_values={
                        "student_id": str(item.student_id),
                        "year": item.year,
                        "month": item.month,
                        "item_name": item.item_name,
                        "amount": item.amount,
                        "memo": item.memo,
                        "sort_order": item.sort_order
                    },
                    note=f"월별 관리비 항목 생성 - {item.item_name} ({item.year}년 {item.month}월)"
                )
        except Exception as log_error:
            print(f"로그 생성 중 오류: {log_error}")
        
        return {
            "message": f"{target_year}年の月別管理費項目が正常に作成されました",
            "created_items": len(items_to_create),
            "student_id": student_id,
            "item_name": item_data.item_name,
            "year": target_year
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"月別管理費項目の作成中にエラーが発生しました: {str(e)}")

@app.put("/monthly-items/sort-order")
def update_monthly_items_sort_order(
    sort_order_data: MonthlyItemSortOrderUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """월별 관리비 항목의 정렬 순서 업데이트"""
    try:
        student_id = sort_order_data.student_id
        year = sort_order_data.year
        items = sort_order_data.items
        
        # 학생 존재 여부 확인
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="学生が見つかりません")
        
        # 각 항목의 sort_order 업데이트 (해당 학생의 데이터만)
        updated_count = 0
        log_data = []  # 로그 데이터를 저장할 리스트
        
        for item_data in items:
            item_name = item_data["item_name"]
            new_sort_order = item_data["sort_order"]
            
            # 해당 항목의 모든 월 데이터 조회 (학생 ID, 년도, 항목명으로 필터링)
            monthly_items = db.query(BillingMonthlyItem).filter(
                BillingMonthlyItem.item_name == item_name,
                BillingMonthlyItem.student_id == student_id,
                BillingMonthlyItem.year == year
            ).all()
            
            if monthly_items:
                # 각 월의 데이터에 대해 sort_order 업데이트
                for item in monthly_items:
                    # 기존 값 저장 (로그용)
                    old_sort_order = item.sort_order
                    
                    # sort_order 업데이트
                    item.sort_order = new_sort_order
                    
                    # 로그 데이터 수집
                    log_data.append({
                        "record_id": str(item.id),
                        "old_values": {"sort_order": old_sort_order},
                        "new_values": {"sort_order": new_sort_order},
                        "note": f"月別管理費項目の並び順更新 - {item.item_name} ({item.year}年{item.month}月): {old_sort_order} → {new_sort_order}"
                    })
                    
                    updated_count += 1
            else:
                # 해당 항목이 해당 학생의 데이터가 아니거나 존재하지 않는 경우
                raise HTTPException(
                    status_code=404, 
                    detail=f"項目名 '{item_name}'のデータが見つからないか、該当学生のデータではありません"
                )
        
        # 데이터베이스 커밋
        db.commit()
        
        # 모든 업데이트 완료 후 한 번에 로그 생성
        try:
            for log_item in log_data:
                create_database_log(
                    db=db,
                    table_name="billing_monthly_items",
                    record_id=log_item["record_id"],
                    action="UPDATE",
                    user_id=current_user.get("id") if current_user else None,
                    old_values=log_item["old_values"],
                    new_values=log_item["new_values"],
                    changed_fields=["sort_order"],
                    note=log_item["note"]
                )
        except Exception as log_error:
            print(f"로그 생성 중 오류: {log_error}")
        
        # 데이터베이스 커밋
        db.commit()
        
        return {
            "message": f"{student.name}の{year}年の月別管理費項目の並び順が正常に更新されました",
            "student_id": student_id,
            "student_name": student.name,
            "year": year,
            "updated_count": updated_count,
            "total_items": len(items)
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"並び順の更新中にエラーが発生しました: {str(e)}")

@app.put("/monthly-items/{item_id}")
def update_monthly_item(
    item_id: str,
    item_update: BillingMonthlyItemUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """월별 관리비 항목 수정 (인풋에서 커서가 나가면 저장)"""
    try:
        # 항목 조회
        item = db.query(BillingMonthlyItem).filter(BillingMonthlyItem.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="該当の項目が見つかりません")
        
        # 기존 값 저장 (로그용)
        old_values = {
            "amount": item.amount,
            "memo": item.memo
        }
        
        # 업데이트할 데이터 준비
        update_data = {}
        if item_update.amount is not None:
            item.amount = item_update.amount
        if item_update.memo is not None:
            item.memo = item_update.memo
        
        # 데이터베이스 업데이트
        db.commit()
        
        # 데이터베이스 로그 생성
        try:
            create_database_log(
                db=db,
                table_name="billing_monthly_items",
                record_id=str(item.id),
                action="UPDATE",
                user_id=current_user.get("id") if current_user else None,
                old_values=old_values,
                new_values={
                    "amount": item.amount,
                    "memo": item.memo
                },
                changed_fields=["amount", "memo"] if item_update.amount is not None or item_update.memo is not None else [],
                note=f"月別管理費項目の並び順更新 - {item.item_name}"
            )
        except Exception as e:
            print(f"로그 생성 중 오류: {e}")
        
        return {
            "message": "項目が正常に更新されました",
            "updated_item": item
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"項目更新中にエラーが発生しました: {str(e)}")

@app.delete("/students/{student_id}/monthly-items/{item_name}")
def delete_monthly_item(
    student_id: str,
    item_name: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """학생의 특정 항목 1~12월 데이터 모두 삭제"""
    try:
        # 학생 존재 여부 확인
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다.")
        
        # 해당 항목의 모든 월 데이터 조회
        items = db.query(BillingMonthlyItem).filter(
            BillingMonthlyItem.student_id == student_id,
            BillingMonthlyItem.item_name == item_name
        ).all()
        
        if not items:
            raise HTTPException(status_code=404, detail="해당 항목을 찾을 수 없습니다.")
        
        # 삭제 전 데이터 저장 (로그용)
        deleted_items_data = []
        for item in items:
            deleted_items_data.append({
                "id": str(item.id),
                "student_id": str(item.student_id),
                "year": item.year,
                "month": item.month,
                "item_name": item.item_name,
                "amount": item.amount,
                "memo": item.memo,
                "sort_order": item.sort_order
            })
        
        # 모든 월 데이터 삭제
        deleted_count = len(items)
        for item in items:
            db.delete(item)
        
        db.commit()
        
        # 데이터베이스 로그 생성 (각 삭제된 항목별로)
        try:
            for item_data in deleted_items_data:
                create_database_log(
                    db=db,
                    table_name="billing_monthly_items",
                    record_id=item_data["id"],
                    action="DELETE",
                    user_id=current_user.get("id") if current_user else None,
                    old_values=item_data,
                    note=f"월별 관리비 항목 삭제 - {item_data['item_name']} ({item_data['year']}년 {item_data['month']}월)"
                )
        except Exception as log_error:
            print(f"로그 생성 중 오류: {log_error}")
        
        return {
            "message": f"'{item_name}'項目の{deleted_count}ヶ月データが正常に削除されました",
            "student_id": student_id,
            "item_name": item_name,
            "deleted_count": deleted_count
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"항목 삭제 중 오류가 발생했습니다: {str(e)}")

@app.get("/monthly-items/{item_id}")
def get_monthly_item(
    item_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 월별 관리비 항목 조회"""
    try:
        # 항목 조회
        item = db.query(BillingMonthlyItem).filter(BillingMonthlyItem.id == item_id).first()
        
        if not item:
            raise HTTPException(status_code=404, detail="該当の項目が見つかりません")
        
        return item
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"項目の取得中にエラーが発生しました: {str(e)}")

@app.post("/billing-invoices/generate")
def generate_company_invoice_pdfV2(
    company_id: str,
    year: int,
    month: int,
    student_type: Optional[str] = Query(None, description="학생 타입으로 필터링"),
    memo: Optional[str] = Query(None, description="비고 작성"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """회사별 청구서 생성 (billing_monthly_items 기반) 및 PDF 생성"""
    try:
        # 회사 존재 여부 확인
        company = db.query(Company).filter(Company.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="会社が見つかりません。")
        
        # 해당 회사의 학생들 조회 (학생 타입 필터링)
        query = db.query(Student).filter(Student.company_id == company_id)
        if student_type:
            query = query.filter(Student.student_type == student_type)
        
        students = query.all()
        if not students:
            raise HTTPException(status_code=404, detail="該当会社の学生がありません。")
        
        # 회사별로 하나의 청구서 생성
        created_invoices = []
        invoice_list = []
        total_company_amount = 0
        all_invoice_items = []
        
        # 모든 학생의 월별 항목을 수집
        for student in students:
            # 해당 학생의 해당 월 billing_monthly_items 조회
            monthly_items = db.query(BillingMonthlyItem).filter(
                BillingMonthlyItem.student_id == student.id,
                BillingMonthlyItem.year == year,
                BillingMonthlyItem.month == month
            ).all()
            
            if not monthly_items:
                continue  # 해당 월에 항목이 없으면 건너뛰기
            
            # 총 금액 계산
            total_amount = sum(float(item.amount) for item in monthly_items)
            total_company_amount += total_amount
            
            # 청구서 항목들 준비 (스냅샷)
            for item in monthly_items:
                invoice_item = {
                    "student_id": student.id,
                    "item_name": item.item_name,
                    "amount": item.amount,
                    "memo": item.memo,
                    "sort_order": item.sort_order,
                    "original_item_id": item.id
                }
                all_invoice_items.append(invoice_item)
        
        # 회사별 총액이 있는 경우에만 청구서 생성
        if total_company_amount > 0:
            # 청구서 생성
            invoice = BillingInvoice(
                company_id=company.id,
                year=year,
                month=month,
                total_amount=total_company_amount,
                memo=memo or f"{year}年 {month}月 請求書"
            )
            db.add(invoice)
            db.flush()  # ID 생성
            
            # 청구서 항목들 생성
            invoice_items = []
            for item_data in all_invoice_items:
                invoice_item = BillingInvoiceItem(
                    invoice_id=invoice.id,
                    student_id=item_data["student_id"],
                    item_name=item_data["item_name"],
                    amount=item_data["amount"],
                    memo=item_data["memo"],
                    sort_order=item_data["sort_order"],
                    original_item_id=item_data["original_item_id"]
                )
                invoice_items.append(invoice_item)
            
            db.add_all(invoice_items)
            created_invoices.append(invoice)
            
            # PDF용 데이터 구성 (학생별로 그룹화)
            student_invoice_data = {}
            for item_data in all_invoice_items:
                student_id = item_data["student_id"]
                if student_id not in student_invoice_data:
                    student_invoice_data[student_id] = {
                        "student_name": next((s.name for s in students if s.id == student_id), "Unknown Student"),
                        "invoice_items": [],
                        "total_amount": 0
                    }
                
                amount = int(float(item_data["amount"]))
                if amount > 0:  # 0원이 아닌 항목만 추가
                    student_invoice_data[student_id]["invoice_items"].append({
                        "name": item_data["item_name"] or "Unknown Item",
                        "unit_price": amount,
                        "amount": amount,
                        "memo": item_data["memo"] or ""
                    })
                    student_invoice_data[student_id]["total_amount"] += amount
            
            # 0원이 아닌 항목이 있는 학생만 추가
            for student_data in student_invoice_data.values():
                if student_data["total_amount"] > 0:
                    invoice_data = {
                        "student_name": student_data["student_name"],
                        "invoice_number": str(invoice.id),
                        "invoice_items": student_data["invoice_items"],
                        "total_amount": student_data["total_amount"]
                    }
                    invoice_list.append(invoice_data)
        
        db.commit()
        
        # 학생 타입별 청구서 정보 설정
        def get_invoice_info(student_type):
            if student_type == "SPECIFIED":
                return {
                    "sender_name": "株式会社ワールドワーカー",
                    "sender_address": "〒546-0003 大阪市東住吉区今川四丁目5番9号",
                    "sender_tel": "TEL 06-6760-7830",
                    "sender_fax": "FAX -",
                    "registration_number": "登録番号 -",
                    "recipient_name": f"{company.name}　御中",
                    "recipient_address": company.address or ""
                }
            elif student_type == "GENERAL":
                return {
                    "sender_name": "大阪医療介護協同組合",
                    "sender_address": "〒546-0023 大阪府大阪市東住吉区矢田1-26-7-1階",
                    "sender_tel": "TEL 06-6654-8836",
                    "sender_fax": "FAX 06-6654-8837",
                    "registration_number": "登録番号 T6120005018169",
                    "recipient_name": f"{company.name}　御中",
                    "recipient_address": company.address or ""
                }
            else:  # 기본값 (기존 정보)
                return {
                    "sender_name": "大阪医療介護協同組合",
                    "sender_address": "〒546-0023 大阪府大阪市東住吉区矢田1-26-7-1階",
                    "sender_tel": "TEL 06-6654-8836",
                    "sender_fax": "FAX 06-6654-8837",
                    "registration_number": "登録番号 T6120005018169",
                    "recipient_name": f"{company.name}　御中",
                    "recipient_address": company.address or ""
                }
        
        # 작성일 (오늘 날짜)
        invoice_date = datetime.now().strftime("%Y年%m月%d日")
        
        # 입금 기한 (해당 월의 마지막 날)
        last_day = calendar.monthrange(year, month)[1]
        payment_deadline = f"{year}年 {month}月 {last_day}日"
        
        # 학생 타입별 청구서 정보 가져오기
        invoice_info = get_invoice_info(student_type)
        
        pdf_data = {
            "total_amount": int(total_company_amount),
            "invoice_list": invoice_list,
            "invoice_date": invoice_date,
            "payment_deadline": payment_deadline,
            "sender_name": invoice_info["sender_name"],
            "sender_address": invoice_info["sender_address"],
            "sender_tel": invoice_info["sender_tel"],
            "sender_fax": invoice_info["sender_fax"],
            "registration_number": invoice_info["registration_number"],
            "recipient_name": invoice_info["recipient_name"],
            "memo": memo or ""
        }
        
        # HTML 템플릿 렌더링
        template = templates.get_template("company_invoice.html")
        html_content = template.render(data=pdf_data)
        
        # HTML 내용을 UTF-8로 인코딩하여 PDF 생성
        try:
            pdf_bytes = html_to_pdf_bytes(html_content)
        except UnicodeEncodeError:
            # 인코딩 에러 발생 시 HTML 내용을 안전하게 처리
            safe_html_content = html_content.encode('utf-8', errors='ignore').decode('utf-8')
            pdf_bytes = html_to_pdf_bytes(safe_html_content)
        
        # 파일명 생성 (완전히 영어로 변경)
        filename = f"invoice_{company_id}_{year}_{month}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"請求書作成中にエラーが発生しました: {str(e)}")

@app.post("/billing-invoices/generate/excel")
def generate_company_invoice_excel(
    company_id: str,
    year: int,
    month: int,
    student_type: Optional[str] = Query(None, description="학생 타입으로 필터링"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """회사별 청구서 상세 데이터를 엑셀 파일로 다운로드"""
    try:
        # 회사 존재 여부 확인
        company = db.query(Company).filter(Company.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="会社が見つかりません。")
        
        # 해당 회사의 학생들 조회 (학생 타입 필터링)
        query = db.query(Student).filter(Student.company_id == company_id)
        if student_type:
            query = query.filter(Student.student_type == student_type)
        
        students = query.all()
        if not students:
            raise HTTPException(status_code=404, detail="該当会社の学生がありません。")
        
        # 모든 학생의 월별 항목을 수집
        all_invoice_items = []
        total_company_amount = 0
        
        for student in students:
            # 해당 학생의 해당 월 billing_monthly_items 조회
            monthly_items = db.query(BillingMonthlyItem).filter(
                BillingMonthlyItem.student_id == student.id,
                BillingMonthlyItem.year == year,
                BillingMonthlyItem.month == month
            ).all()
            
            if not monthly_items:
                continue  # 해당 월에 항목이 없으면 건너뛰기
            
            # 총 금액 계산
            total_amount = sum(float(item.amount) for item in monthly_items)
            total_company_amount += total_amount
            
            # 청구서 항목들 준비
            for item in monthly_items:
                invoice_item = {
                    "student_id": student.id,
                    "student_name": student.name,
                    "student_type": student.student_type,
                    "grade_name": student.grade.name if student.grade else None,
                    "item_name": item.item_name,
                    "amount": item.amount,
                    "memo": item.memo,
                    "sort_order": item.sort_order,
                    "original_item_id": item.id
                }
                all_invoice_items.append(invoice_item)
        
        if not all_invoice_items:
            raise HTTPException(status_code=404, detail="該当月の請求項目がありません。")
        
        # 엑셀 데이터 구성 (필요한 컬럼만, 0원 항목 제거)
        excel_data = []
        
        # 헤더 행 추가 (필요한 컬럼만)
        excel_data.append([
            "学生名", "項目名", "金額"
        ])
        
        # 학생별로 그룹화하여 데이터 정리
        student_groups = {}
        for item in all_invoice_items:
            student_id = item["student_id"]
            if student_id not in student_groups:
                student_groups[student_id] = {
                    "student_name": item["student_name"],
                    "items": []
                }
            
            # 0원이 아닌 항목만 추가
            if float(item["amount"]) > 0:
                student_groups[student_id]["items"].append({
                    "item_name": item["item_name"],
                    "amount": item["amount"]
                })
        print(f"엑셀 데이터: {student_groups}")
        # 그룹화된 데이터를 엑셀 형식으로 변환
        for student_id, student_data in student_groups.items():
            if student_data["items"]:  # 항목이 있는 경우만
                # 첫 번째 행: 학생명과 첫 번째 항목
                first_item = student_data["items"][0]
                excel_data.append([
                    student_data["student_name"],
                    first_item["item_name"],
                    first_item["amount"]
                ])
                
                # 나머지 항목들 (학생명은 빈 값으로)
                for item in student_data["items"][1:]:
                    excel_data.append([
                        "",  # 학생명 병합을 위해 빈 값
                        item["item_name"],
                        item["amount"]
                    ])
        
        # 요약 행 추가
        excel_data.append([])  # 빈 행
        excel_data.append([
            "合計", "", total_company_amount
        ])
        
                # 엑셀 파일 생성
        # DataFrame 생성
        df = pd.DataFrame(excel_data[1:], columns=excel_data[0])
        
        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="請求詳細", index=False)
            ws = writer.sheets["請求詳細"]
            
            # 열 너비 자동 조정
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
            
            # 학생명 셀 병합
            current_student_name = None
            merge_start_row = 2  # 헤더 다음 행부터 시작
            merge_end_row = 2
            
            for row_idx in range(2, len(excel_data) + 1):  # 헤더 제외하고 2행부터
                student_name = ws.cell(row=row_idx, column=1).value
                
                if student_name and student_name != "":  # 학생명이 있는 경우
                    if current_student_name is not None and merge_start_row < merge_end_row:
                        # 이전 학생의 셀 병합
                        ws.merge_cells(f'A{merge_start_row}:A{merge_end_row}')
                    
                    # 새로운 학생 시작
                    current_student_name = student_name
                    merge_start_row = row_idx
                    merge_end_row = row_idx
                else:
                    # 같은 학생의 추가 항목
                    merge_end_row = row_idx
            
            # 마지막 학생의 셀 병합
            if current_student_name is not None and merge_start_row < merge_end_row:
                ws.merge_cells(f'A{merge_start_row}:A{merge_end_row}')
            
            # 셀 정렬 설정
            for row in ws.iter_rows(min_row=2, max_row=len(excel_data), min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    
        output.seek(0)

        # 파일명 생성 (슬래시 등 위험문자 제거 권장)
        student_type_filter = f"_{student_type}" if student_type else ""
        display_name = f"請求詳細_{company.name}_{year}年{month}月{student_type_filter}.xlsx"
        # OS/브라우저 안전용: ASCII 대체 이름
        ascii_fallback = "invoice_detail.xlsx"
        # RFC 5987: UTF-8 퍼센트 인코딩
        filename_star = quote(display_name)

        headers = {
            # ASCII 대체 + 실제 유니코드 파일명
            "Content-Disposition": f"attachment; filename={ascii_fallback}; filename*=UTF-8''{filename_star}"
        }

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)}")

@app.get("/billing-invoices/{invoice_id}", response_model=BillingInvoiceResponse)
def get_billing_invoice(
    invoice_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 청구서 조회"""
    try:
        invoice = db.query(BillingInvoice).options(
            joinedload(BillingInvoice.items)
        ).filter(BillingInvoice.id == invoice_id).first()
        
        if not invoice:
            raise HTTPException(status_code=404, detail="請求書が見つかりません。")
        
        return invoice
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"請求書照会中にエラーが発生しました: {str(e)}")

@app.get("/billing-invoices/student/{student_id}")
def get_student_billing_invoices(
    student_id: str,
    year: Optional[int] = Query(None, description="년도로 필터링"),
    month: Optional[int] = Query(None, description="월로 필터링"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """학생이 속한 회사의 청구서 목록 조회 (해당 학생의 항목만 포함)"""
    try:
        # 학생 존재 여부 확인
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="学生が見つかりません。")
        
        # 기본 쿼리 생성 - 해당 학생이 속한 회사의 청구서들
        query = db.query(BillingInvoice).filter(BillingInvoice.company_id == student.company_id)
        
        # 필터링
        if year is not None:
            query = query.filter(BillingInvoice.year == year)
        if month is not None:
            query = query.filter(BillingInvoice.month == month)
        
        # 정렬 (최신순)
        invoices = query.order_by(BillingInvoice.created_at.desc()).all()
        
        # 각 청구서에서 해당 학생의 항목들만 필터링
        result_invoices = []
        for invoice in invoices:
            # 해당 학생의 항목들만 조회
            student_items = db.query(BillingInvoiceItem).filter(
                BillingInvoiceItem.invoice_id == invoice.id,
                BillingInvoiceItem.student_id == student_id
            ).all()
            
            if student_items:  # 해당 학생의 항목이 있는 경우만 포함
                # invoice 객체에 student_items 추가
                invoice.student_items = student_items
                result_invoices.append(invoice)
        
        return {
            "student_id": student_id,
            "student_name": student.name,
            "company_id": student.company_id,
            "invoices": result_invoices
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"請求書一覧照会中にエラーが発生しました: {str(e)}")

@app.delete("/billing-invoices/{invoice_id}")
def delete_billing_invoice(
    invoice_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """청구서 삭제"""
    try:
        invoice = db.query(BillingInvoice).filter(BillingInvoice.id == invoice_id).first()
        if not invoice:
            raise HTTPException(status_code=404, detail="請求書が見つかりません。")
        
        db.delete(invoice)
        db.commit()
        
        return {
            "message": "請求書が正常に削除されました。",
            "deleted_invoice_id": invoice_id
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"請求書削除中にエラーが発生しました: {str(e)}")

# 노인 식사 기록 관련 API들
@app.post("/elderly-meal-records", response_model=ElderlyMealRecordResponse)
def create_elderly_meal_record(
    meal_record: ElderlyMealRecordCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """노인 식사 건너뛴 기록 등록"""
    # 거주자 존재 여부 확인
    resident = db.query(Resident).filter(Resident.id == meal_record.resident_id).first()
    if not resident:
        raise HTTPException(status_code=404, detail="거주자를 찾을 수 없습니다")

    # 날짜 형식 검증
    try:
        if isinstance(meal_record.skip_date, str):
            record_date = datetime.strptime(meal_record.skip_date, "%Y-%m-%d").date()
        else:
            record_date = meal_record.skip_date
    except ValueError:
        raise HTTPException(status_code=400, detail="날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)")

    # 이미 같은 날짜, 같은 식사 유형의 기록이 있는지 확인
    existing_record = db.query(ElderlyMealRecord).filter(
        ElderlyMealRecord.resident_id == meal_record.resident_id,
        ElderlyMealRecord.skip_date == record_date,
        ElderlyMealRecord.meal_type == meal_record.meal_type
    ).first()

    if existing_record:
        # 기존 기록이 있으면 삭제 (토글 기능)
        try:
            # 삭제 전 값 저장 (로그용)
            old_values = {
                "resident_id": str(existing_record.resident_id),
                "date": str(existing_record.skip_date),
                "meal_type": existing_record.meal_type
            }
            
            db.delete(existing_record)
            db.commit()
            
            # 로그 생성
            create_database_log(
                db=db,
                table_name="elderly_meal_records",
                record_id=str(existing_record.id),
                action="DELETE",
                user_id=current_user["id"] if current_user else None,
                old_values=old_values,
                note="노인 식사 기록 토글 삭제 (취소)"
            )
            
            return {
                "message": "식사 기록이 취소되었습니다",
                "action": "cancelled",
                "id": str(existing_record.id),
                "resident_id": str(existing_record.resident_id),
                "skip_date": existing_record.skip_date,
                "meal_type": existing_record.meal_type,
                "created_at": existing_record.created_at,
                "resident": {
                    "id": str(resident.id),
                    "resident_type": resident.resident_type,
                    "elderly": {
                        "id": str(resident.elderly.id),
                        "name": resident.elderly.name
                    } if resident.elderly else None
                }
            }
            
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"식사 기록 취소 중 오류가 발생했습니다: {str(e)}"
            )

    try:
        # 새로운 식사 기록 등록
        new_record = ElderlyMealRecord(
        id=str(uuid.uuid4()),
            resident_id=meal_record.resident_id,
            skip_date=record_date,
            meal_type=meal_record.meal_type
        )
        db.add(new_record)
        db.commit()
        db.refresh(new_record)
        
        # 로그 생성
        create_database_log(
            db=db,
            table_name="elderly_meal_records",
            record_id=str(new_record.id),
            action="CREATE",
            user_id=current_user["id"] if current_user else None,
            new_values={
                "resident_id": str(new_record.resident_id),
                "date": str(new_record.skip_date),
                "meal_type": new_record.meal_type
            },
            note="노인 식사 건너뛴 기록 등록"
        )

        # 응답 데이터 준비
        response_data = {
          "message": "식사 기록이 등록되었습니다",
          "action": "created",
          "id": str(new_record.id),
          "resident_id": str(new_record.resident_id),
          "skip_date": new_record.skip_date,
          "meal_type": new_record.meal_type,
          "created_at": new_record.created_at,
          "resident": {
              "id": str(resident.id),
              "resident_type": resident.resident_type,
              "elderly": {
                  "id": str(resident.elderly.id),
                  "name": resident.elderly.name
              } if resident.elderly else None
          }
        }

        return response_data

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"식사 기록 등록 중 오류가 발생했습니다: {str(e)}"
        )

@app.get("/elderly-meal-records/monthly/{resident_id}/{year}/{month}")
def get_elderly_meal_records_monthly(
    resident_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 거주자의 특정 월 식사 건너뛴 기록 조회 (1일~31일)"""
    # 거주자 존재 여부 확인
    resident = db.query(Resident).options(
        joinedload(Resident.elderly)
    ).filter(Resident.id == resident_id).first()
    if not resident:
        raise HTTPException(status_code=404, detail="거주자를 찾을 수 없습니다")

    # 년도, 월 유효성 검사
    if year < 1900 or year > 2100:
        raise HTTPException(status_code=400, detail="년도가 올바르지 않습니다")
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월이 올바르지 않습니다")

    try:
        # 해당 월의 시작일과 종료일 계산
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # 해당 월의 식사 기록 조회
        records = db.query(ElderlyMealRecord).filter(
            ElderlyMealRecord.resident_id == resident_id,
            ElderlyMealRecord.skip_date >= start_date,
            ElderlyMealRecord.skip_date <= end_date
        ).order_by(ElderlyMealRecord.skip_date, ElderlyMealRecord.meal_type).all()

        # 1일부터 31일까지의 모든 날짜에 대해 데이터 구성
        monthly_data = {}
        for day in range(1, 32):
            current_date = date(year, month, day)
            # 해당 월에 실제로 존재하는 날짜인지 확인
            if current_date.month == month:
                monthly_data[day] = {
                    "date": current_date,
                    "breakfast": False,
                    "lunch": False,
                    "dinner": False,
                    "total_skipped": 0
                }

        # 조회된 기록을 월별 데이터에 반영
        for record in records:
            day = record.date.day
            if day in monthly_data:
                monthly_data[day][record.meal_type] = True
                monthly_data[day]["total_skipped"] += 1

        # 응답 데이터 준비
        result = []
        for day, data in monthly_data.items():
            result.append({
                "day": day,
                "date": str(data["date"]),
                "breakfast": data["breakfast"],
                "lunch": data["lunch"],
                "dinner": data["dinner"],
                "total_skipped": data["total_skipped"]
            })

        # 월별 통계 계산
        total_breakfast_skipped = sum(1 for data in monthly_data.values() if data["breakfast"])
        total_lunch_skipped = sum(1 for data in monthly_data.values() if data["lunch"])
        total_dinner_skipped = sum(1 for data in monthly_data.values() if data["dinner"])
        total_days_with_skips = sum(1 for data in monthly_data.values() if data["total_skipped"] > 0)

        return {
            "resident": {
                "id": str(resident.id),
                "resident_type": resident.resident_type,
                "elderly": {
                    "id": str(resident.elderly.id),
                    "name": resident.elderly.name
                } if resident.elderly else None
            },
            "year": year,
            "month": month,
            "start_date": str(start_date),
            "end_date": str(end_date),
            "daily_records": result,
            "monthly_statistics": {
                "total_breakfast_skipped": total_breakfast_skipped,
                "total_lunch_skipped": total_lunch_skipped,
                "total_dinner_skipped": total_dinner_skipped,
                "total_days_with_skips": total_days_with_skips,
                "total_meals_skipped": total_breakfast_skipped + total_lunch_skipped + total_dinner_skipped
            }
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"날짜 계산 중 오류가 발생했습니다: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"월별 식사 기록 조회 중 오류가 발생했습니다: {str(e)}")

@app.get("/elderly-meal-records/monthly-all/{year}/{month}")
def get_all_elderly_meal_records_monthly(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """전체 노인들의 특정 월 식사 건너뛴 기록 조회 (1일~31일)"""
    # 년도, 월 유효성 검사
    if year < 1900 or year > 2100:
        raise HTTPException(status_code=400, detail="년도가 올바르지 않습니다")
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월이 올바르지 않습니다")

    try:
        # 해당 월의 시작일과 종료일 계산
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # 해당 월의 모든 식사 기록 조회 (노인 거주자만)
        records = db.query(ElderlyMealRecord).options(
            joinedload(ElderlyMealRecord.resident).joinedload(Resident.elderly)
        ).join(Resident).filter(
            ElderlyMealRecord.skip_date >= start_date,
            ElderlyMealRecord.skip_date <= end_date,
            Resident.resident_type == "elderly"
        ).order_by(ElderlyMealRecord.skip_date, ElderlyMealRecord.meal_type).all()

        # 1일부터 31일까지의 모든 날짜에 대해 데이터 구성
        monthly_data = {}
        for day in range(1, 32):
            current_date = date(year, month, day)
            # 해당 월에 실제로 존재하는 날짜인지 확인
            if current_date.month == month:
                monthly_data[day] = {
                    "date": current_date,
                    "breakfast": 0,
                    "lunch": 0,
                    "dinner": 0,
                    "total_skipped": 0,
                    "residents_skipped": set()  # 해당 날짜에 식사를 건너뛴 거주자들
                }

        # 조회된 기록을 월별 데이터에 반영
        for record in records:
            day = record.date.day
            if day in monthly_data:
                monthly_data[day][record.meal_type] += 1
                monthly_data[day]["total_skipped"] += 1
                monthly_data[day]["residents_skipped"].add(str(record.resident_id))

        # 응답 데이터 준비
        result = []
        for day, data in monthly_data.items():
            result.append({
                "day": day,
                "date": str(data["date"]),
                "breakfast": data["breakfast"],
                "lunch": data["lunch"],
                "dinner": data["dinner"],
                "total_skipped": data["total_skipped"],
                "residents_skipped_count": len(data["residents_skipped"])
            })

        # 월별 통계 계산
        total_breakfast_skipped = sum(data["breakfast"] for data in monthly_data.values())
        total_lunch_skipped = sum(data["lunch"] for data in monthly_data.values())
        total_dinner_skipped = sum(data["dinner"] for data in monthly_data.values())
        total_meals_skipped = total_breakfast_skipped + total_lunch_skipped + total_dinner_skipped
        
        # 전체 노인 거주자 수 조회
        total_elderly_residents = db.query(Resident).filter(
            Resident.resident_type == "elderly",
            Resident.is_active == True
        ).count()

        # 식사를 건너뛴 노인 수 (중복 제거)
        all_residents_with_skips = set()
        for data in monthly_data.values():
            all_residents_with_skips.update(data["residents_skipped"])
        
        total_residents_with_skips = len(all_residents_with_skips)

        return {
            "year": year,
            "month": month,
            "start_date": str(start_date),
            "end_date": str(end_date),
            "daily_records": result,
            "monthly_statistics": {
                "total_breakfast_skipped": total_breakfast_skipped,
                "total_lunch_skipped": total_lunch_skipped,
                "total_dinner_skipped": total_dinner_skipped,
                "total_meals_skipped": total_meals_skipped,
                "total_elderly_residents": total_elderly_residents,
                "total_residents_with_skips": total_residents_with_skips,
                "average_meals_skipped_per_resident": round(total_meals_skipped / max(total_elderly_residents, 1), 2)
            }
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"날짜 계산 중 오류가 발생했습니다: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"전체 월별 식사 기록 조회 중 오류가 발생했습니다: {str(e)}")

@app.get("/elderly-meal-records/monthly-summary/{year}/{month}")
def get_elderly_meal_records_monthly_summary(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """전체 노인들의 특정 월 식사 건너뛴 기록 요약 조회 (거주자별 통계)"""
    # 년도, 월 유효성 검사
    if year < 1900 or year > 2100:
        raise HTTPException(status_code=400, detail="년도가 올바르지 않습니다")
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월이 올바르지 않습니다")

    try:
        # 해당 월의 시작일과 종료일 계산
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # 해당 월의 모든 식사 기록 조회
        records = db.query(ElderlyMealRecord).options(
            joinedload(ElderlyMealRecord.resident).joinedload(Resident.elderly)
        ).filter(
            ElderlyMealRecord.skip_date >= start_date,
            ElderlyMealRecord.skip_date <= end_date
        ).all()

        # 거주자별로 데이터 그룹화
        resident_summary = {}
        for record in records:
            resident_id = str(record.resident_id)
            if resident_id not in resident_summary:
                resident_summary[resident_id] = {
                    "resident_id": resident_id,
                    "resident_name": record.resident.elderly.name if record.resident.elderly else "Unknown",
                    "breakfast_skipped": 0,
                    "lunch_skipped": 0,
                    "dinner_skipped": 0,
                    "total_meals_skipped": 0,
                    "days_with_skips": set()
                }
            
            resident_summary[resident_id][f"{record.meal_type}_skipped"] += 1
            resident_summary[resident_id]["total_meals_skipped"] += 1
            resident_summary[resident_id]["days_with_skips"].add(record.date.day)

        # 응답 데이터 준비
        result = []
        total_breakfast_skipped = 0
        total_lunch_skipped = 0
        total_dinner_skipped = 0
        total_residents_with_skips = 0

        for resident_data in resident_summary.values():
            resident_data["days_with_skips"] = len(resident_data["days_with_skips"])
            result.append(resident_data)
            
            total_breakfast_skipped += resident_data["breakfast_skipped"]
            total_lunch_skipped += resident_data["lunch_skipped"]
            total_dinner_skipped += resident_data["dinner_skipped"]
            
            if resident_data["total_meals_skipped"] > 0:
                total_residents_with_skips += 1

        return {
            "year": year,
            "month": month,
            "start_date": str(start_date),
            "end_date": str(end_date),
            "resident_summary": result,
            "overall_statistics": {
                "total_breakfast_skipped": total_breakfast_skipped,
                "total_lunch_skipped": total_lunch_skipped,
                "total_dinner_skipped": total_dinner_skipped,
                "total_meals_skipped": total_breakfast_skipped + total_lunch_skipped + total_dinner_skipped,
                "total_residents_with_skips": total_residents_with_skips,
                "total_residents": len(resident_summary)
            }
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"날짜 계산 중 오류가 발생했습니다: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"월별 요약 조회 중 오류가 발생했습니다: {str(e)}")

@app.get("/elderly-meal-records/monthly-building/{building_id}/{year}/{month}")
def get_elderly_meal_records_monthly_by_building(
    building_id: str,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 빌딩의 노인 거주자들의 월별 식사 건너뛴 기록 조회"""
    # 빌딩 존재 여부 확인
    building = db.query(Building).filter(Building.id == building_id).first()
    if not building:
        raise HTTPException(status_code=404, detail="빌딩을 찾을 수 없습니다")

    # 년도, 월 유효성 검사
    if year < 1900 or year > 2100:
        raise HTTPException(status_code=400, detail="년도가 올바르지 않습니다")
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="월이 올바르지 않습니다")

    try:
        # 해당 월의 시작일과 종료일 계산
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # 해당 빌딩의 노인 거주자들 조회
        elderly_residents = db.query(Resident).options(
            joinedload(Resident.elderly),
            joinedload(Resident.room).joinedload(Room.building)
        ).filter(
            Resident.resident_type == "elderly",
            Resident.is_active == True,
            Resident.room.has(Room.building_id == building_id)
        ).all()

        if not elderly_residents:
            return {
                "building": {
                    "id": str(building.id),
                    "name": building.name
                },
                "year": year,
                "month": month,
                "start_date": str(start_date),
                "end_date": str(end_date),
                "daily_records": [],
                "residents": [],
                "monthly_statistics": {
                    "total_breakfast_skipped": 0,
                    "total_lunch_skipped": 0,
                    "total_dinner_skipped": 0,
                    "total_meals_skipped": 0,
                    "total_elderly_residents": 0,
                    "total_residents_with_skips": 0,
                    "average_meals_skipped_per_resident": 0
                }
            }

        # 해당 월의 식사 기록 조회 (해당 빌딩의 노인 거주자들만)
        resident_ids = [str(resident.id) for resident in elderly_residents]
        records = db.query(ElderlyMealRecord).options(
            joinedload(ElderlyMealRecord.resident).joinedload(Resident.elderly)
        ).filter(
            ElderlyMealRecord.skip_date >= start_date,
            ElderlyMealRecord.skip_date <= end_date,
            ElderlyMealRecord.resident_id.in_(resident_ids)
        ).order_by(ElderlyMealRecord.skip_date, ElderlyMealRecord.meal_type).all()

        # 1일부터 31일까지의 모든 날짜에 대해 데이터 구성
        monthly_data = {}
        for day in range(1, 32):
            current_date = date(year, month, day)
            # 해당 월에 실제로 존재하는 날짜인지 확인
            if current_date.month == month:
                monthly_data[day] = {
                    "date": current_date,
                    "breakfast": 0,
                    "lunch": 0,
                    "dinner": 0,
                    "total_skipped": 0,
                    "residents_skipped": set()
                }

        # 해당 월의 입원 기록 조회 (해당 빌딩의 노인 거주자들만)
        elderly_ids = [str(resident.elderly.id) for resident in elderly_residents if resident.elderly]
        hospitalization_records = db.query(ElderlyHospitalization).filter(
            ElderlyHospitalization.date >= start_date,
            ElderlyHospitalization.date <= end_date,
            ElderlyHospitalization.elderly_id.in_(elderly_ids)
        ).order_by(ElderlyHospitalization.date).all()

        # 조회된 기록을 월별 데이터에 반영
        for record in records:
            day = record.skip_date.day
            if day in monthly_data:
                monthly_data[day][record.meal_type] += 1
                monthly_data[day]["total_skipped"] += 1
                monthly_data[day]["residents_skipped"].add(str(record.resident_id))

        # 응답 데이터 준비
        daily_records = []
        for day, data in monthly_data.items():
            daily_record = {
                "day": day,
                "date": str(data["date"]),
                "breakfast": data["breakfast"],
                "lunch": data["lunch"],
                "dinner": data["dinner"],
                "total_skipped": data["total_skipped"],
                "residents_skipped_count": len(data["residents_skipped"])
            }
            daily_records.append(daily_record)

        # 거주자별 개별 통계 및 일별 상세 데이터 계산
        residents_data = []
        for resident in elderly_residents:
            resident_records = [r for r in records if str(r.resident_id) == str(resident.id)]
            
            # 거주자별 일별 데이터 구성 (1일~31일)
            resident_daily_data = {}
            for day in range(1, 32):
                current_date = date(year, month, day)
                if current_date.month == month:
                    resident_daily_data[day] = {
                        "day": day,
                        "date": str(current_date),
                        "breakfast": False,
                        "lunch": False,
                        "dinner": False,
                        "total_skipped": 0
                    }
            
            # 해당 거주자의 기록을 일별 데이터에 반영
            for record in resident_records:
                day = record.skip_date.day
                if day in resident_daily_data:
                    resident_daily_data[day][record.meal_type] = True
                    resident_daily_data[day]["total_skipped"] += 1
            
            # 일별 데이터를 배열로 변환
            daily_records_array = []
            for day, data in resident_daily_data.items():
                daily_records_array.append(data)
            
            # 거주자별 통계 계산
            breakfast_skipped = len([r for r in resident_records if r.meal_type == "breakfast"])
            lunch_skipped = len([r for r in resident_records if r.meal_type == "lunch"])
            dinner_skipped = len([r for r in resident_records if r.meal_type == "dinner"])
            total_skipped = breakfast_skipped + lunch_skipped + dinner_skipped
            
            # 해당 거주자의 입원 기록 조회
            resident_hospitalizations = []
            if resident.elderly:
                resident_hospitalizations = [h for h in hospitalization_records if str(h.elderly_id) == str(resident.elderly.id)]
            
            residents_data.append({
                "resident_id": str(resident.id),
                "resident_name": resident.elderly.name if resident.elderly else "Unknown",
                "room_number": resident.room.room_number if resident.room else None,
                "breakfast_skipped": breakfast_skipped,
                "lunch_skipped": lunch_skipped,
                "dinner_skipped": dinner_skipped,
                "total_skipped": total_skipped,
                "days_with_skips": len(set(r.skip_date.day for r in resident_records)),
                "daily_records": daily_records_array,  # 1일~31일 상세 데이터
                "hospitalizations": [
                    {
                        "elderly_id": str(h.elderly_id),
                        "hospitalization_type": h.hospitalization_type,
                        "hospital_name": h.hospital_name,
                        "last_meal_date": h.last_meal_date,
                        "last_meal_type": h.last_meal_type,
                        "meal_resume_date": h.meal_resume_date,
                        "meal_resume_type": h.meal_resume_type,
                        "note": h.note
                    } for h in resident_hospitalizations
                ]
            })

        # 월별 통계 계산
        total_breakfast_skipped = sum(data["breakfast"] for data in monthly_data.values())
        total_lunch_skipped = sum(data["lunch"] for data in monthly_data.values())
        total_dinner_skipped = sum(data["dinner"] for data in monthly_data.values())
        total_meals_skipped = total_breakfast_skipped + total_lunch_skipped + total_dinner_skipped
        
        # 식사를 건너뛴 노인 수 (중복 제거)
        all_residents_with_skips = set()
        for data in monthly_data.values():
            all_residents_with_skips.update(data["residents_skipped"])
        
        total_residents_with_skips = len(all_residents_with_skips)
        total_elderly_residents = len(elderly_residents)

        return {
            "building": {
                "id": str(building.id),
                "name": building.name
            },
            "year": year,
            "month": month,
            "start_date": str(start_date),
            "end_date": str(end_date),
            "daily_records": daily_records,
            "residents": residents_data,
            "monthly_statistics": {
                "total_breakfast_skipped": total_breakfast_skipped,
                "total_lunch_skipped": total_lunch_skipped,
                "total_dinner_skipped": total_dinner_skipped,
                "total_meals_skipped": total_meals_skipped,
                "total_elderly_residents": total_elderly_residents,
                "total_residents_with_skips": total_residents_with_skips,
                "average_meals_skipped_per_resident": round(total_meals_skipped / max(total_elderly_residents, 1), 2)
            }
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"날짜 계산 중 오류가 발생했습니다: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"빌딩별 월별 식사 기록 조회 중 오류가 발생했습니다: {str(e)}")

@app.post("/elderly-hospitalizations")
def create_elderly_hospitalization(
    hospitalization_data: ElderlyHospitalizationCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """거주자 입원/퇴원 기록을 생성합니다."""
    try:
        # 노인 존재 여부 확인
        elderly = db.query(Elderly).filter(Elderly.id == hospitalization_data.elderly_id).first()
        if not elderly:
            raise HTTPException(status_code=404, detail="노인을 찾을 수 없습니다")
        
        # 퇴원 등록인 경우 기존 입원 기록 찾기
        if hospitalization_data.hospitalization_type == "discharge":
            existing_hospitalization = db.query(ElderlyHospitalization).filter(
                ElderlyHospitalization.elderly_id == hospitalization_data.elderly_id,
                ElderlyHospitalization.hospitalization_type == "admission",
                ElderlyHospitalization.meal_resume_date.is_(None)  # 퇴원 기록이 없는 입원 기록
            ).order_by(ElderlyHospitalization.date.desc()).first()
            
            if existing_hospitalization:
                # 기존 입원 기록 업데이트
                existing_hospitalization.hospitalization_type = "discharge"
                existing_hospitalization.meal_resume_date = hospitalization_data.meal_resume_date
                existing_hospitalization.meal_resume_type = hospitalization_data.meal_resume_type
                existing_hospitalization.note = hospitalization_data.note
                
                db.commit()
                db.refresh(existing_hospitalization)
                
                return {
                    "message": f"퇴원 기록이 기존 입원 기록에 업데이트되었습니다.",
                    "item": {
                        "id": str(existing_hospitalization.id),
                        "elderly_id": str(existing_hospitalization.elderly_id),
                        "hospitalization_type": existing_hospitalization.hospitalization_type,
                        "hospital_name": existing_hospitalization.hospital_name,
                        "date": existing_hospitalization.date,
                        "meal_resume_date": existing_hospitalization.meal_resume_date,
                        "meal_resume_type": existing_hospitalization.meal_resume_type
                    }
                }
            else:
                raise HTTPException(status_code=404, detail="해당 노인의 미완료 입원 기록을 찾을 수 없습니다")
        
        # 입원 기록 생성 (기존 로직)
        hospitalization = ElderlyHospitalization(
            elderly_id=hospitalization_data.elderly_id,
            hospitalization_type=hospitalization_data.hospitalization_type,
            hospital_name=hospitalization_data.hospital_name,
            date=hospitalization_data.date,
            last_meal_date=hospitalization_data.last_meal_date,
            last_meal_type=hospitalization_data.last_meal_type,
            meal_resume_date=hospitalization_data.meal_resume_date,
            meal_resume_type=hospitalization_data.meal_resume_type,
            note=hospitalization_data.note,
            created_by=current_user.get("name", "Unknown")
        )
        
        db.add(hospitalization)
        db.commit()
        db.refresh(hospitalization)
        
        return {
            "message": f"입원 기록이 정상적으로 생성되었습니다.",
            "item": {
              "id": str(hospitalization.id),
              "elderly_id": str(hospitalization.elderly_id),
              "hospitalization_type": hospitalization.hospitalization_type,
              "hospital_name": hospitalization.hospital_name,
              "date": hospitalization.date
            }
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"입원/퇴원 기록 생성 중 오류가 발생했습니다: {str(e)}")

@app.get("/elderly-hospitalizations/{year}/{month}")
def get_elderly_hospitalizations_by_month(
    year: int,
    month: int,
    elderly_id: Optional[str] = Query(None, description="특정 고령자 ID로 필터링"),
    hospitalization_type: Optional[str] = Query(None, description="입원/퇴원 유형으로 필터링 (admission 또는 discharge)"),
    page: int = Query(1, description="페이지 번호", ge=1),
    page_size: int = Query(10, description="페이지당 항목 수", ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 연도와 달의 병원 입퇴원 기록을 조회합니다."""
    try:
        # 연도와 달 유효성 검사
        if year < 1900 or year > 2100:
            raise HTTPException(status_code=400, detail="유효하지 않은 연도입니다")
        if month < 1 or month > 12:
            raise HTTPException(status_code=400, detail="유효하지 않은 월입니다")
        
        # 해당 월의 시작일과 종료일 계산
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # 기본 쿼리 생성
        query = db.query(ElderlyHospitalization).filter(
            ElderlyHospitalization.date >= start_date,
            ElderlyHospitalization.date <= end_date
        )
        
        # 필터링 조건 추가
        if elderly_id:
            query = query.filter(ElderlyHospitalization.elderly_id == elderly_id)
        
        if hospitalization_type:
            query = query.filter(ElderlyHospitalization.hospitalization_type == hospitalization_type)
        
        # 총 개수 계산
        total_count = query.count()
        
        # 페이지네이션 적용
        offset = (page - 1) * page_size
        hospitalizations = query.order_by(ElderlyHospitalization.date.desc()).offset(offset).limit(page_size).all()
        
        # 응답 데이터 구성
        hospitalization_data = []
        for hospitalization in hospitalizations:
            # 고령자 정보 조회
            elderly = db.query(Elderly).filter(Elderly.id == hospitalization.elderly_id).first()
            
            hospitalization_dict = {
                "id": str(hospitalization.id),
                "elderly_id": str(hospitalization.elderly_id),
                "elderly_name": elderly.name if elderly else "Unknown",
                "elderly_name_katakana": elderly.name_katakana if elderly else None,
                "hospitalization_type": hospitalization.hospitalization_type,
                "hospital_name": hospitalization.hospital_name,
                "date": hospitalization.date,
                "last_meal_date": hospitalization.last_meal_date,
                "last_meal_type": hospitalization.last_meal_type,
                "meal_resume_date": hospitalization.meal_resume_date,
                "meal_resume_type": hospitalization.meal_resume_type,
                "note": hospitalization.note,
                "created_at": hospitalization.created_at,
                "created_by": hospitalization.created_by,
                "elderly": {
                    "id": str(elderly.id),
                    "name": elderly.name,
                    "name_katakana": elderly.name_katakana,
                    "gender": elderly.gender,
                    "care_level": elderly.care_level,
                    "current_room": {
                        "id": str(elderly.current_room.id),
                        "room_number": elderly.current_room.room_number,
                        "building": {
                            "id": str(elderly.current_room.building.id),
                            "name": elderly.current_room.building.name
                        } if elderly.current_room and elderly.current_room.building else None
                    } if elderly.current_room else None
                } if elderly else None
            }
            hospitalization_data.append(hospitalization_dict)
        
        # 월별 통계 계산
        admission_count = len([h for h in hospitalizations if h.hospitalization_type == 'admission'])
        discharge_count = len([h for h in hospitalizations if h.hospitalization_type == 'discharge'])
        unique_elderly_count = len(set(h.elderly_id for h in hospitalizations))
        
        return {
            "year": year,
            "month": month,
            "start_date": str(start_date),
            "end_date": str(end_date),
            "items": hospitalization_data,
            "total": total_count,
            "total_pages": (total_count + page_size - 1) // page_size,
            "current_page": page,
            "page_size": page_size,
            "statistics": {
                "total_hospitalizations": len(hospitalizations),
                "admission_count": admission_count,
                "discharge_count": discharge_count,
                "unique_elderly_count": unique_elderly_count
            }
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"날짜 계산 중 오류가 발생했습니다: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"병원 입퇴원 기록 조회 중 오류가 발생했습니다: {str(e)}")